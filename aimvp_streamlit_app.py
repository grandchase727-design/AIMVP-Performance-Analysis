"""
AIMVP Executive Dashboard - Streamlit Interactive App
Run: streamlit run aimvp_streamlit_app.py
"""
import streamlit as st
import yfinance as yf
import openpyxl
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import os
import warnings
warnings.filterwarnings('ignore')

# ============== Page config ==============
st.set_page_config(
    page_title='AIMVP Executive Dashboard',
    page_icon='📊',
    layout='wide',
    initial_sidebar_state='expanded',
)

# ============== FT (Financial Times) 테마 ==============
# 시그니처: 핑크 페이퍼(#FFF1E5) · 워킹 블랙(#33302E) · 클라렛 액센트(#990F3D) · 세리프 헤드라인
import plotly.io as pio

FT_PINK, FT_BLACK, FT_CLARET, FT_OXFORD, FT_TEAL, FT_GREY = (
    '#FFF1E5', '#33302E', '#990F3D', '#0F5499', '#0D7680', '#66605C')
FT_LINE = 'rgba(51,48,46,0.12)'

# Plotly 전역 템플릿 — 개별 차트가 색/배경을 명시하지 않은 경우에만 적용(시그널/프로파일 의미색은 보존)
pio.templates['ft'] = go.layout.Template(layout=go.Layout(
    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
    colorway=[FT_OXFORD, FT_CLARET, FT_TEAL, '#B86B2E', '#593380', '#1E854E', FT_GREY, '#9E2F50'],
    font=dict(family="-apple-system, 'Segoe UI', Helvetica, Arial, sans-serif",
              color=FT_BLACK, size=12),
    title=dict(font=dict(family="Georgia, 'Source Serif Pro', serif", size=15, color=FT_BLACK)),
    xaxis=dict(gridcolor=FT_LINE, linecolor='rgba(51,48,46,0.45)',
               zerolinecolor='rgba(51,48,46,0.30)', ticks='outside', tickcolor=FT_LINE),
    yaxis=dict(gridcolor=FT_LINE, linecolor='rgba(0,0,0,0)',
               zerolinecolor='rgba(51,48,46,0.30)'),
    legend=dict(font=dict(size=11)),
))
pio.templates.default = 'plotly_white+ft'

st.markdown("""
<style>
/* ── FT 핑크 페이퍼 배경 ── */
.stApp, [data-testid="stAppViewContainer"], [data-testid="stHeader"] { background-color: #FFF1E5; }
[data-testid="stHeader"] { border-bottom: 1px solid rgba(51,48,46,0.10); }
[data-testid="stSidebar"] { background-color: #FBEAD9; border-right: 1px solid #E7D3BD; }
/* ── 헤드라인 세리프 (FT Financier 대용 Georgia) ── */
h1,h2,h3,h4,h5,h6,
[data-testid="stMarkdownContainer"] h1,[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3,[data-testid="stMarkdownContainer"] h4 {
  font-family: Georgia, 'Source Serif Pro', 'Times New Roman', serif !important;
  color: #33302E !important; font-weight: 700; letter-spacing: -0.01em;
}
h1 { font-size: 1.9rem !important; } h2 { font-size: 1.45rem !important; } h3 { font-size: 1.18rem !important; }
/* ── 얇은 룰 ── */
hr { border: none !important; border-top: 1px solid rgba(51,48,46,0.20) !important; margin: 0.7rem 0 !important; }
/* ── 링크: 옥스포드 블루 ── */
a, a:visited { color: #0F5499 !important; }
/* ── 메트릭: 흰 패널 + 얇은 테두리, 값은 세리프 ── */
[data-testid="stMetric"] { background:#FFFFFF; border:1px solid #E7D3BD; border-radius:2px; padding:10px 14px; }
[data-testid="stMetricValue"] { font-family: Georgia, serif; color:#0F5499; font-weight:700; }
[data-testid="stMetricLabel"] { color:#66605C; }
/* ── 탭: 클라렛 활성 인디케이터 ── */
.stTabs [data-baseweb="tab-list"] { border-bottom: 1px solid rgba(51,48,46,0.18); }
.stTabs [aria-selected="true"] { color:#990F3D !important; border-bottom-color:#990F3D !important; }
/* ── 버튼: primary/다운로드 = 클라렛 ── */
.stButton>button { border-radius:2px; }
.stButton>button[kind="primary"], .stDownloadButton>button {
  background:#990F3D !important; color:#FFF1E5 !important; border:1px solid #990F3D !important; }
/* ── 확장 패널 ── */
[data-testid="stExpander"] { border:1px solid #E7D3BD; border-radius:2px; background:rgba(255,255,255,0.45); }
/* ── FT 마스트헤드 ── */
.ft-masthead { border-top:3px solid #33302E; border-bottom:1px solid rgba(51,48,46,0.25);
  padding:10px 0 8px 0; margin:0 0 8px 0; }
.ft-brand { font-family:Georgia,'Source Serif Pro',serif; font-size:2.0rem; font-weight:700;
  color:#33302E; letter-spacing:-0.02em; line-height:1.05; }
.ft-brand .clr { color:#990F3D; }
.ft-tag { font-size:0.80rem; color:#66605C; letter-spacing:0.06em; text-transform:uppercase; margin-top:3px; }
</style>
""", unsafe_allow_html=True)

# ============== Constants ==============
# 스크립트와 같은 폴더의 Excel을 사용 (이식성). 환경변수 AIMVP_EXCEL_PATH로 오버라이드 가능.
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.environ.get(
    'AIMVP_EXCEL_PATH',
    os.path.join(_APP_DIR, 'AIMVP 포트폴리오 비중추이.xlsx'),
)
FUND_START = datetime(2022, 3, 1)

# 정적 벤치마크 현금 proxy = 한국 CD(91일물) 금리 (연율). 조정 가능 상수.
# (TIGER CD금리 ETF 357870.KS는 2023-06 이후만 존재 → 전체 기간 커버 위해 상수 proxy.
#  2022~26 CD 91일 평균 ~3.0~3.5% 수준 / 357870.KS 실현 CAGR ~3.2%)
CD91_ANNUAL = 0.035

TICKER_MAP = {
    'acwi us equity':'ACWI','efa us equity':'EFA','iwm us equity':'IWM',
    'qqq us equity':'QQQ','spy us equity':'SPY','vwo us equity':'VWO',
    'pdbc us equity':'PDBC','dbmf us equity':'DBMF',
    'gto us equity':'GTO','hyg us equity':'HYG','lqd us equity':'LQD',
    'bkln us equity':'BKLN','ief us equity':'IEF','embd us equity':'EMBD',
    'ewa us equity':'EWA','ewc us equity':'EWC','enzl us equity':'ENZL',
    'ewg us equity':'EWG','norw us equity':'NORW','ewo us equity':'EWO',
    'ewh us equity':'EWH','ews us equity':'EWS','ewi us equity':'EWI',
    'ewn us equity':'EWN','ewj us equity':'EWJ','mchi us equity':'MCHI',
    'ewk us equity':'EWK','ewq us equity':'EWQ','ewu us equity':'EWU',
    'ewl us equity':'EWL','ewz us equity':'EWZ','inda us equity':'INDA',
    'ewy us equity':'EWY','eden us equity':'EDEN','eis us equity':'EIS',
    'ewd us equity':'EWD','ewp us equity':'EWP',
    'phyl us equity':'PHYL',  # PGIM Active HY Bond (added 2026-06)
    'a357870':'_CASH_KR','357870 ks equity':'_CASH_KR',
    'a114820':'_KRBOND','114820 ks equity':'_KRBOND',
    'a451540':'_KRBOND','451540 ks equity':'_KRBOND',
    'usdkrw curncy':'_CASH_USD','cash_krw':'_CASH_KR',
}

TICKER_NAMES = {
    'acwi us equity': 'iShares MSCI ACWI',
    'efa us equity':  'iShares MSCI EAFE',
    'iwm us equity':  'iShares Russell 2000',
    'qqq us equity':  'Invesco QQQ',
    'spy us equity':  'SPDR S&P 500',
    'vwo us equity':  'Vanguard FTSE EM',
    'pdbc us equity': 'Invesco Commodity',
    'dbmf us equity': 'IMGP DBi Managed Futures',
    'gto us equity':  'Invesco Total Return Bond',
    'hyg us equity':  'iShares US HY',
    'lqd us equity':  'iShares IG Corp Bond',
    'bkln us equity': 'Invesco Senior Loan',
    'ief us equity':  'iShares 7-10Y Treasury',
    'embd us equity': 'Global X EM Bond',
    'ewa us equity':  'iShares MSCI Australia',
    'ewc us equity':  'iShares MSCI Canada',
    'enzl us equity': 'iShares MSCI NZ',
    'ewg us equity':  'iShares MSCI Germany',
    'norw us equity': 'Global X Norway',
    'ewo us equity':  'iShares MSCI Austria',
    'ewh us equity':  'iShares MSCI HK',
    'ews us equity':  'iShares MSCI Singapore',
    'ewi us equity':  'iShares MSCI Italy',
    'ewn us equity':  'iShares MSCI Netherlands',
    'ewj us equity':  'iShares MSCI Japan',
    'mchi us equity': 'iShares MSCI China',
    'ewk us equity':  'iShares MSCI Belgium',
    'ewq us equity':  'iShares MSCI France',
    'ewu us equity':  'iShares MSCI UK',
    'ewl us equity':  'iShares MSCI Switzerland',
    'ewz us equity':  'iShares MSCI Brazil',
    'inda us equity': 'iShares MSCI India',
    'ewy us equity':  'iShares MSCI Korea',
    'eden us equity': 'iShares MSCI Denmark',
    'eis us equity':  'iShares MSCI Israel',
    'ewd us equity':  'iShares MSCI Sweden',
    'ewp us equity':  'iShares MSCI Spain',
    'phyl us equity': 'PGIM Active HY Bond',
    'a357870':         'TIGER CD금리 (KR)',
    '357870 ks equity':'TIGER CD금리 (KR)',
    'a114820':         'TIGER 국채3년 (KR)',
    '114820 ks equity':'TIGER 국채3년 (KR)',
    'a451540':         'TIGER 종합채권 (KR)',
    '451540 ks equity':'TIGER 종합채권 (KR)',
    'usdkrw curncy':   'USD Cash',
    'cash_krw':        'KRW Cash',
}

# 캡처 비대칭(Base·주식only에서 하락>상승) 원인 분석 본문 — 대시보드 expander에서 사용
_CAPTURE_RCA_MD = """
### 한 줄 답
하락장 특유의 위험(다운사이드 베타 급등)이 **아니라**, 주식 슬리브가 ACWI 대비 **베타 ≈ 1.04 + 지속적인 작은 음(−)의 알파(≈ −0.14%/월)** 를 가지며, **합산법(aggregate) 캡처 공식이 이 음의 알파를 상승월엔 빼고 하락월엔 더하기** 때문입니다.

### 1. 메커니즘 — 공식의 분모 부호 뒤집힘 (산술적 원인)
슬리브를 `슬리브월수익 = β·ACWI월수익 + α` 로 보면, 합산법 정의상:

- **상승 캡처** = β + α·(n_up / ΣACWI_up) — ΣACWI_up = **+77.3 (양수)** → 음의 α가 **빼짐**
- **하락 캡처** = β + α·(n_dn / ΣACWI_dn) — ΣACWI_dn = **−35.1 (음수)** → 같은 음의 α가 **더해짐**

β=1.04, α=−0.147 대입 → 상승 **1.002** / 하락 **1.073** (관측 1.002 / 1.070 일치). 원시 합계 2×2 연립 → **β=1.038, α=−0.139%/월**. 상승–하락 격차 0.068은 **100% α 항에서 발생**하고 β는 격차에 0 기여 — 베타는 캡처의 *수준*을 1.04로 올릴 뿐, *비대칭*은 오직 음의 알파가 만듭니다.

### 2. 경제적 원천 — 음의 알파는 어디서 오나

| 구성 | 상승월 | 하락월 |
|---|---|---|
| **미국 대형/기술 초과보유** (SPY 35% + QQQ 11% 가 ACWI 36% 위에 얹힘 → 실효 미국비중 > ACWI ~60%) · β>1 | 도움(상승 증폭) | **드래그: QQQ −1.80, SPY −1.73** (전체 압도) |
| **EM·단일국가 오버레이** (VWO·中·印·韓·EFA·日) · 미국 주도장 부진 | **드래그: VWO −1.14, 中 −0.83, 印 −0.32** (상승 못 따라감) | 방어 쿠션 없음 |

- **상승월:** 미국 β>1 이득을 EM 부진이 거의 정확히 상쇄 → 캡처 ≈ 1.00 (β 1.04에 못 미침)
- **하락월:** 고베타 미국주가 더 빠지고 EM은 쿠션이 안 됨 → 캡처 ≈ 1.07

### 3. 아닌 것 (기각된 가설)
- **하락장 베타 급등 / 컨벡시티 ✗** — 일별 베타가 상승일 **1.040** vs 하락일 **1.042** (사실상 동일). 최대 드래그 월은 −8% 폭락월이 아니라 ACWI −0.55%인 **2025-02** → 컨벡시티와 정반대.
- **단일월 / 표본 아티팩트 ✗** — 2025-02을 빼도 하락 캡처 **1.0445 > 상승 1.0023**. 8개 하락월 중 무엇을 빼도 [1.045, 1.090], **부호 불변**. 중립형(n_dn=9)도 독립적으로 동일 패턴.

### 4. 견고성과 한계
- ✅ **방향(하락>상승)은 견고** — 상수 베타 + 음의 알파로 완전 재현, 레짐 의존 항 불필요.
- ⚠️ **수치 크기(1.07/1.09)는 취약** — 2025-02 한 달이 하락-초과의 ~38%. "1.07"보다 **"구조적으로 1보다 약간 큼"** 으로 해석 권장.
- ⚠️ **레짐 종속** — 음의 알파는 *2022~26 미국/AI 주도장* 산물. SPY 부진·EM 주도 레짐에선 비대칭이 좁혀지거나 역전 가능.

### 5. 재검증 (현재 데이터)
- 적극형 Base·주식only: 상승 **1.0023** / 하락 **1.0701** (n_up 20 / n_dn 8).  중립형: **1.016 / 1.091** (n_dn 9).
- 두 방정식에서 푼 α: 상승식 **−0.133%/월**, 하락식 **−0.147%/월** (β=1.037) — 부호·크기 일치.
- LOO(2025-02 제외) 하락 **1.0445 > 상승 1.0023** — 부호 유지.
- ※ **소계행 파싱 제외(엔진 정정) 후 재산출 — 주식only는 주식 슬리브만 필터하므로 소계 영향이 없어 값 불변.** 결론 동일.
"""

# ETF look-through to country exposure (% of each ETF allocated to country)
# Multi-country ETFs use representative composition; single-country ETFs = 100%
LOOKTHROUGH = {
    # ── Single-country country ETFs ──
    'spy us equity':  {'미국': 100},
    'qqq us equity':  {'미국': 100},
    'iwm us equity':  {'미국': 100},
    'ewa us equity':  {'호주': 100},
    'ewc us equity':  {'캐나다': 100},
    'enzl us equity': {'뉴질랜드': 100},
    'ewg us equity':  {'독일': 100},
    'norw us equity': {'노르웨이': 100},
    'ewo us equity':  {'오스트리아': 100},
    'ewh us equity':  {'홍콩': 100},
    'ews us equity':  {'싱가포르': 100},
    'ewi us equity':  {'이탈리아': 100},
    'ewn us equity':  {'네덜란드': 100},
    'ewj us equity':  {'일본': 100},
    'mchi us equity': {'중국': 100},
    'ewk us equity':  {'벨기에': 100},
    'ewq us equity':  {'프랑스': 100},
    'ewu us equity':  {'영국': 100},
    'ewl us equity':  {'스위스': 100},
    'ewz us equity':  {'브라질': 100},
    'inda us equity': {'인도': 100},
    'ewy us equity':  {'한국': 100},
    'eden us equity': {'덴마크': 100},
    'eis us equity':  {'이스라엘': 100},
    'ewd us equity':  {'스웨덴': 100},
    'ewp us equity':  {'스페인': 100},
    # ── Multi-country ETFs (approximate composition) ──
    'acwi us equity': {  # MSCI ACWI (Global)
        '미국': 63.0, '일본': 5.0, '영국': 3.5, '중국': 3.0,
        '캐나다': 2.7, '프랑스': 2.6, '독일': 2.2, '스위스': 2.0,
        '대만': 2.0, '호주': 1.8, '인도': 1.7, '네덜란드': 1.4,
        '한국': 1.3, '기타': 7.8,
    },
    'efa us equity': {  # MSCI EAFE (Developed ex-US)
        '일본': 22.0, '영국': 16.0, '프랑스': 11.0, '스위스': 9.0,
        '독일': 9.0, '호주': 7.0, '네덜란드': 5.0, '스웨덴': 4.0,
        '덴마크': 3.0, '홍콩': 3.0, '이탈리아': 2.7, '스페인': 2.5,
        '싱가포르': 1.5, '기타': 4.3,
    },
    'vwo us equity': {  # FTSE Emerging Markets
        '중국': 33.0, '인도': 19.0, '대만': 17.0, '브라질': 5.5,
        '한국': 4.3, '사우디': 3.6, '남아공': 3.4, '멕시코': 2.4,
        '말레이시아': 1.4, '인도네시아': 1.4, '태국': 1.3, '기타': 6.7,
    },
    # 원자재/헤지펀드/채권/현금은 country lookthrough 제외 (지역성 무관)
}


def compute_country_breakdown(weights, top_n=10):
    """Look-through ETF holdings to country exposure. Returns top N (% of equity sleeve)."""
    country = {}
    eq_sleeve_total = 0
    for tk, w in weights.items():
        if tk not in LOOKTHROUGH:
            continue
        eq_sleeve_total += w
        for c, pct in LOOKTHROUGH[tk].items():
            country[c] = country.get(c, 0) + w * pct / 100
    if eq_sleeve_total == 0:
        return [], 0
    # Normalize to equity sleeve = 100%
    country_normalized = {c: v / eq_sleeve_total * 100 for c, v in country.items()}
    sorted_countries = sorted(country_normalized.items(), key=lambda x: x[1], reverse=True)
    return sorted_countries[:top_n], eq_sleeve_total


def compute_country_contribution(weights, etf_returns_pct):
    """국가별 기여(bps) — 주식 슬리브 100% 정규화 기준.

    etf_returns_pct: {ticker: r%} — 외부에서 계산된 period return (%).
    각 국가 c 기여 = ∑_ETF (slv_w_etf × etf_country_share_c × r_etf) / 100  (bps)
    """
    eq_sleeve_total = sum(w for tk, w in weights.items() if tk in LOOKTHROUGH)
    if eq_sleeve_total == 0:
        return {}
    contribs = {}
    for tk, w in weights.items():
        if tk not in LOOKTHROUGH:
            continue
        slv_w = w / eq_sleeve_total * 100  # %
        r = etf_returns_pct.get(tk, 0.0)
        for c, share in LOOKTHROUGH[tk].items():
            contribs[c] = contribs.get(c, 0) + slv_w * share * r / 100  # bps
    return contribs


# ETF sector look-through (approximate GICS sector composition, % of each ETF)
SECTOR_LOOKTHROUGH = {
    # ── Broad ETFs ──
    'spy us equity': {  # S&P 500
        '정보기술': 32.0, '금융': 13.0, '헬스케어': 11.0, '임의소비재': 10.0,
        '통신서비스': 9.0, '산업재': 8.0, '필수소비재': 5.5, '에너지': 4.0,
        '유틸리티': 2.5, '부동산': 2.5, '소재': 2.5,
    },
    'qqq us equity': {  # Nasdaq-100 (tech-heavy)
        '정보기술': 50.0, '통신서비스': 16.0, '임의소비재': 16.0,
        '헬스케어': 6.0, '필수소비재': 6.0, '산업재': 4.0,
        '금융': 1.5, '유틸리티': 0.5,
    },
    'iwm us equity': {  # Russell 2000 (small cap)
        '산업재': 18.0, '금융': 16.0, '헬스케어': 16.0, '정보기술': 12.0,
        '임의소비재': 10.0, '부동산': 6.0, '에너지': 6.0, '소재': 5.0,
        '필수소비재': 4.0, '통신서비스': 4.0, '유틸리티': 3.0,
    },
    'acwi us equity': {  # MSCI ACWI Global
        '정보기술': 26.0, '금융': 16.0, '헬스케어': 11.0, '임의소비재': 11.0,
        '산업재': 11.0, '통신서비스': 8.0, '필수소비재': 6.0, '에너지': 4.0,
        '소재': 3.5, '유틸리티': 2.5, '부동산': 2.0,
    },
    'efa us equity': {  # MSCI EAFE (Developed ex-US)
        '금융': 21.0, '산업재': 17.0, '헬스케어': 12.0, '임의소비재': 11.0,
        '정보기술': 9.0, '필수소비재': 8.0, '소재': 7.0, '통신서비스': 5.0,
        '에너지': 5.0, '유틸리티': 3.0, '부동산': 2.0,
    },
    'vwo us equity': {  # FTSE Emerging Markets
        '정보기술': 23.0, '금융': 22.0, '임의소비재': 13.0, '통신서비스': 10.0,
        '산업재': 7.0, '소재': 7.0, '필수소비재': 6.0, '에너지': 5.0,
        '유틸리티': 3.0, '부동산': 3.0, '헬스케어': 1.0,
    },
    # ── Country ETFs (top countries used by AIMVP) ──
    'ewj us equity': {  # Japan
        '산업재': 23.0, '임의소비재': 19.0, '정보기술': 15.0, '금융': 11.0,
        '헬스케어': 9.0, '통신서비스': 8.0, '필수소비재': 6.0, '소재': 5.0,
        '유틸리티': 2.0, '에너지': 2.0,
    },
    'mchi us equity': {  # China
        '임의소비재': 27.0, '금융': 23.0, '통신서비스': 19.0, '정보기술': 8.0,
        '산업재': 7.0, '헬스케어': 6.0, '필수소비재': 4.0, '소재': 3.0,
        '에너지': 2.0, '부동산': 1.0,
    },
    'ewc us equity': {  # Canada
        '금융': 32.0, '에너지': 17.0, '산업재': 12.0, '정보기술': 11.0,
        '소재': 11.0, '임의소비재': 4.0, '통신서비스': 4.0, '필수소비재': 4.0,
        '헬스케어': 2.0, '부동산': 2.0, '유틸리티': 1.0,
    },
    'ewy us equity': {  # Korea
        '정보기술': 36.0, '임의소비재': 13.0, '금융': 13.0, '산업재': 10.0,
        '통신서비스': 10.0, '헬스케어': 7.0, '소재': 6.0, '필수소비재': 3.0,
        '에너지': 1.0, '유틸리티': 1.0,
    },
    'ewu us equity': {  # United Kingdom
        '금융': 21.0, '필수소비재': 16.0, '에너지': 13.0, '헬스케어': 13.0,
        '산업재': 11.0, '임의소비재': 8.0, '소재': 7.0, '통신서비스': 4.0,
        '유틸리티': 3.0, '정보기술': 1.0, '부동산': 3.0,
    },
    'ewl us equity': {  # Switzerland
        '헬스케어': 35.0, '금융': 20.0, '산업재': 13.0, '필수소비재': 12.0,
        '소재': 7.0, '정보기술': 4.0, '임의소비재': 3.0, '통신서비스': 3.0,
        '부동산': 2.0, '유틸리티': 1.0,
    },
    'ewg us equity': {  # Germany
        '산업재': 25.0, '금융': 18.0, '임의소비재': 14.0, '소재': 12.0,
        '정보기술': 10.0, '헬스케어': 8.0, '필수소비재': 5.0, '통신서비스': 3.0,
        '에너지': 2.0, '유틸리티': 3.0,
    },
    'ewa us equity': {  # Australia
        '금융': 30.0, '소재': 25.0, '헬스케어': 9.0, '산업재': 8.0,
        '부동산': 6.0, '에너지': 5.0, '임의소비재': 5.0, '필수소비재': 4.0,
        '정보기술': 3.0, '통신서비스': 3.0, '유틸리티': 2.0,
    },
    'enzl us equity': {  # New Zealand
        '헬스케어': 18.0, '유틸리티': 16.0, '산업재': 16.0, '통신서비스': 12.0,
        '임의소비재': 10.0, '필수소비재': 9.0, '금융': 7.0, '소재': 5.0,
        '부동산': 4.0, '정보기술': 3.0,
    },
    'norw us equity': {  # Norway
        '에너지': 35.0, '금융': 22.0, '산업재': 12.0, '필수소비재': 8.0,
        '소재': 7.0, '통신서비스': 5.0, '정보기술': 4.0, '헬스케어': 3.0,
        '임의소비재': 2.0, '부동산': 2.0,
    },
    'ewo us equity': {  # Austria
        '금융': 35.0, '에너지': 18.0, '산업재': 14.0, '소재': 10.0,
        '유틸리티': 8.0, '통신서비스': 5.0, '부동산': 4.0, '임의소비재': 3.0,
        '필수소비재': 2.0, '정보기술': 1.0,
    },
    'ewh us equity': {  # Hong Kong
        '금융': 35.0, '부동산': 20.0, '임의소비재': 12.0, '산업재': 10.0,
        '통신서비스': 8.0, '유틸리티': 5.0, '필수소비재': 4.0, '에너지': 3.0,
        '소재': 2.0, '정보기술': 1.0,
    },
    'ews us equity': {  # Singapore
        '금융': 45.0, '산업재': 15.0, '부동산': 13.0, '통신서비스': 10.0,
        '임의소비재': 6.0, '필수소비재': 4.0, '소재': 3.0, '유틸리티': 2.0,
        '정보기술': 1.0, '헬스케어': 1.0,
    },
    'ewi us equity': {  # Italy
        '금융': 38.0, '에너지': 13.0, '유틸리티': 13.0, '산업재': 10.0,
        '임의소비재': 8.0, '필수소비재': 6.0, '통신서비스': 4.0, '소재': 4.0,
        '헬스케어': 3.0, '정보기술': 1.0,
    },
    'ewn us equity': {  # Netherlands
        '정보기술': 30.0, '헬스케어': 18.0, '금융': 15.0, '소재': 10.0,
        '산업재': 8.0, '필수소비재': 7.0, '에너지': 5.0, '통신서비스': 3.0,
        '임의소비재': 2.0, '부동산': 2.0,
    },
    'ewk us equity': {  # Belgium
        '필수소비재': 32.0, '금융': 16.0, '헬스케어': 14.0, '산업재': 10.0,
        '부동산': 8.0, '소재': 6.0, '통신서비스': 5.0, '임의소비재': 4.0,
        '유틸리티': 3.0, '정보기술': 2.0,
    },
    'ewq us equity': {  # France
        '임의소비재': 22.0, '산업재': 18.0, '금융': 15.0, '헬스케어': 12.0,
        '정보기술': 10.0, '필수소비재': 8.0, '에너지': 5.0, '소재': 4.0,
        '통신서비스': 3.0, '유틸리티': 3.0,
    },
    'ewz us equity': {  # Brazil
        '금융': 28.0, '소재': 20.0, '에너지': 15.0, '필수소비재': 10.0,
        '유틸리티': 9.0, '산업재': 6.0, '임의소비재': 5.0, '통신서비스': 4.0,
        '헬스케어': 2.0, '부동산': 1.0,
    },
    'inda us equity': {  # India
        '금융': 24.0, '정보기술': 13.0, '에너지': 10.0, '소재': 9.0,
        '임의소비재': 9.0, '필수소비재': 8.0, '산업재': 8.0, '헬스케어': 7.0,
        '통신서비스': 6.0, '유틸리티': 4.0, '부동산': 2.0,
    },
    'eden us equity': {  # Denmark
        '헬스케어': 50.0, '산업재': 15.0, '금융': 12.0, '필수소비재': 8.0,
        '에너지': 5.0, '정보기술': 4.0, '소재': 3.0, '임의소비재': 2.0,
        '통신서비스': 1.0,
    },
    'eis us equity': {  # Israel
        '정보기술': 35.0, '금융': 22.0, '헬스케어': 12.0, '산업재': 10.0,
        '소재': 7.0, '통신서비스': 5.0, '필수소비재': 4.0, '에너지': 3.0,
        '임의소비재': 2.0,
    },
    'ewd us equity': {  # Sweden
        '산업재': 30.0, '금융': 18.0, '임의소비재': 12.0, '정보기술': 12.0,
        '헬스케어': 8.0, '소재': 6.0, '필수소비재': 5.0, '통신서비스': 4.0,
        '부동산': 3.0, '에너지': 2.0,
    },
    'ewp us equity': {  # Spain
        '금융': 30.0, '유틸리티': 22.0, '산업재': 13.0, '에너지': 10.0,
        '통신서비스': 8.0, '임의소비재': 6.0, '필수소비재': 4.0, '헬스케어': 3.0,
        '소재': 2.0, '부동산': 2.0,
    },
}


def compute_sector_breakdown(weights, top_n=10):
    """Look-through ETF holdings to sector exposure. Returns top N (% of equity sleeve)."""
    sectors = {}
    eq_sleeve_total = 0
    for tk, w in weights.items():
        if tk not in SECTOR_LOOKTHROUGH:
            continue
        eq_sleeve_total += w
        for s, pct in SECTOR_LOOKTHROUGH[tk].items():
            sectors[s] = sectors.get(s, 0) + w * pct / 100
    if eq_sleeve_total == 0:
        return [], 0
    sectors_norm = {s: v / eq_sleeve_total * 100 for s, v in sectors.items()}
    return sorted(sectors_norm.items(), key=lambda x: x[1], reverse=True)[:top_n], eq_sleeve_total


def compute_sector_contribution(weights, etf_returns_pct):
    """섹터별 기여(bps) — 주식 슬리브 100% 정규화 기준.

    각 섹터 s 기여 = ∑_ETF (slv_w_etf × etf_sector_share_s × r_etf) / 100  (bps)
    """
    eq_sleeve_total = sum(w for tk, w in weights.items() if tk in SECTOR_LOOKTHROUGH)
    if eq_sleeve_total == 0:
        return {}
    contribs = {}
    for tk, w in weights.items():
        if tk not in SECTOR_LOOKTHROUGH:
            continue
        slv_w = w / eq_sleeve_total * 100
        r = etf_returns_pct.get(tk, 0.0)
        for s, share in SECTOR_LOOKTHROUGH[tk].items():
            contribs[s] = contribs.get(s, 0) + slv_w * share * r / 100  # bps
    return contribs


# ============================================================
# 🎯 Factor Look-Through (Approach A: MSCI Style Curated)
# ============================================================
# 출처: MSCI Style Factor / Morningstar Style Box / 3rd-party 분석 기반 근사치.
# 각 ETF의 6개 표준 팩터(Value/Growth/Quality/Momentum/Size/Low Vol) 노출도(%).
# 합 100% 정규화 (단, 실제 팩터 노출은 중첩되므로 직관적 share 표현 목적).
FACTOR_LOOKTHROUGH_MSCI = {
    # ── 광역 ETFs ──
    'spy us equity':  {'Value':18,'Growth':22,'Quality':25,'Momentum':13,'Size':7, 'Low Vol':15},  # S&P 500 large blend
    'qqq us equity':  {'Value':5, 'Growth':42,'Quality':23,'Momentum':18,'Size':4, 'Low Vol':8 },  # Nasdaq-100 growth/tech
    'iwm us equity':  {'Value':20,'Growth':18,'Quality':9, 'Momentum':10,'Size':35,'Low Vol':8 },  # Russell 2000 small cap
    'acwi us equity': {'Value':21,'Growth':19,'Quality':22,'Momentum':12,'Size':9, 'Low Vol':17},  # MSCI ACWI global
    'efa us equity':  {'Value':28,'Growth':14,'Quality':18,'Momentum':10,'Size':8, 'Low Vol':22},  # MSCI EAFE developed ex-US
    'vwo us equity':  {'Value':26,'Growth':22,'Quality':12,'Momentum':12,'Size':16,'Low Vol':12},  # FTSE EM
    # ── Country ETFs (MSCI Style Factor 추정) ──
    'ewj us equity':  {'Value':30,'Growth':14,'Quality':16,'Momentum':9, 'Size':9, 'Low Vol':22},  # Japan: value/low vol
    'mchi us equity': {'Value':22,'Growth':27,'Quality':12,'Momentum':13,'Size':9, 'Low Vol':17},  # China: tech/value mix
    'ewu us equity':  {'Value':32,'Growth':11,'Quality':18,'Momentum':8, 'Size':6, 'Low Vol':25},  # UK: value/dividend/low vol
    'ewg us equity':  {'Value':28,'Growth':16,'Quality':18,'Momentum':11,'Size':9, 'Low Vol':18},  # Germany: value/cyclical
    'ewa us equity':  {'Value':32,'Growth':12,'Quality':14,'Momentum':11,'Size':10,'Low Vol':21},  # Australia: value/materials
    'ewc us equity':  {'Value':30,'Growth':14,'Quality':17,'Momentum':10,'Size':9, 'Low Vol':20},  # Canada: value/financials
    'inda us equity': {'Value':18,'Growth':28,'Quality':17,'Momentum':15,'Size':10,'Low Vol':12},  # India: growth
    'ewy us equity':  {'Value':24,'Growth':25,'Quality':16,'Momentum':12,'Size':10,'Low Vol':13},  # Korea: growth/tech
    'ewh us equity':  {'Value':30,'Growth':15,'Quality':16,'Momentum':10,'Size':8, 'Low Vol':21},  # Hong Kong: value
    'ews us equity':  {'Value':31,'Growth':14,'Quality':17,'Momentum':9, 'Size':8, 'Low Vol':21},  # Singapore: value/financials
    'ewi us equity':  {'Value':30,'Growth':14,'Quality':16,'Momentum':11,'Size':10,'Low Vol':19},  # Italy: value
    'ewn us equity':  {'Value':18,'Growth':24,'Quality':22,'Momentum':14,'Size':7, 'Low Vol':15},  # Netherlands: quality/growth
    'ewk us equity':  {'Value':28,'Growth':14,'Quality':18,'Momentum':9, 'Size':8, 'Low Vol':23},  # Belgium: value/low vol
    'ewq us equity':  {'Value':22,'Growth':20,'Quality':21,'Momentum':12,'Size':8, 'Low Vol':17},  # France: quality/growth
    'ewl us equity':  {'Value':18,'Growth':17,'Quality':26,'Momentum':11,'Size':7, 'Low Vol':21},  # Switzerland: quality/low vol
    'ewz us equity':  {'Value':32,'Growth':14,'Quality':12,'Momentum':13,'Size':11,'Low Vol':18},  # Brazil: value
    'ewd us equity':  {'Value':22,'Growth':22,'Quality':20,'Momentum':12,'Size':9, 'Low Vol':15},  # Sweden: quality/growth
    'ewp us equity':  {'Value':30,'Growth':13,'Quality':16,'Momentum':10,'Size':9, 'Low Vol':22},  # Spain: value/low vol
    'ewo us equity':  {'Value':32,'Growth':13,'Quality':15,'Momentum':10,'Size':11,'Low Vol':19},  # Austria: value
    'enzl us equity': {'Value':22,'Growth':16,'Quality':19,'Momentum':9, 'Size':9, 'Low Vol':25},  # New Zealand: low vol/quality
    'norw us equity': {'Value':32,'Growth':13,'Quality':15,'Momentum':10,'Size':9, 'Low Vol':21},  # Norway: value/energy
    'eden us equity': {'Value':18,'Growth':24,'Quality':24,'Momentum':13,'Size':7, 'Low Vol':14},  # Denmark: quality/growth
    'eis us equity':  {'Value':20,'Growth':25,'Quality':18,'Momentum':13,'Size':9, 'Low Vol':15},  # Israel: growth/tech
}

FACTOR_ORDER = ['Value', 'Growth', 'Quality', 'Momentum', 'Size', 'Low Vol']


def compute_factor_breakdown_msci(weights):
    """MSCI 큐레이션 기반 팩터 노출도 (% of equity sleeve)."""
    factors = {}
    eq_sleeve_total = 0
    for tk, w in weights.items():
        if tk not in FACTOR_LOOKTHROUGH_MSCI:
            continue
        eq_sleeve_total += w
        for f, pct in FACTOR_LOOKTHROUGH_MSCI[tk].items():
            factors[f] = factors.get(f, 0) + w * pct / 100
    if eq_sleeve_total == 0:
        return {}, 0
    return {f: v / eq_sleeve_total * 100 for f, v in factors.items()}, eq_sleeve_total


def compute_factor_contribution_msci(weights, etf_returns_pct):
    """팩터별 기여(bps, 슬리브 기준) — MSCI 큐레이션 기반.

    각 팩터 f 기여 = ∑_ETF (slv_w × etf_factor_share × r_etf) / 100  (bps)
    """
    eq_sleeve_total = sum(w for tk, w in weights.items() if tk in FACTOR_LOOKTHROUGH_MSCI)
    if eq_sleeve_total == 0:
        return {}
    contribs = {}
    for tk, w in weights.items():
        if tk not in FACTOR_LOOKTHROUGH_MSCI:
            continue
        slv_w = w / eq_sleeve_total * 100
        r = etf_returns_pct.get(tk, 0.0)
        for f, share in FACTOR_LOOKTHROUGH_MSCI[tk].items():
            contribs[f] = contribs.get(f, 0) + slv_w * share * r / 100  # bps
    return contribs


# ============================================================
# 📐 Factor Look-Through (Approach B: Fama-French Regression)
# ============================================================
# 6개 팩터 = MKT, VAL, SIZE, QUAL, MOM, LOWVOL — 모두 long-short proxy ETFs.
#   MKT     = SPY (시장)
#   VAL     = VLUE - SPY  (Value minus Market)
#   SIZE    = IWM  - SPY  (Small minus Big)
#   Quality = QUAL - SPY  (Quality minus Market)
#   Momentum= MTUM - SPY  (Momentum minus Market)
#   Low Vol = USMV - SPY  (Low Vol minus Market)
FF_FACTOR_SYMS = {
    'MKT':      ('SPY', None),    # Market: SPY direct
    'Value':    ('VLUE', 'SPY'),  # long-short
    'Size':     ('IWM',  'SPY'),
    'Quality':  ('QUAL', 'SPY'),
    'Momentum': ('MTUM', 'SPY'),
    'Low Vol':  ('USMV', 'SPY'),
}
FF_FACTOR_ORDER = ['MKT', 'Value', 'Size', 'Quality', 'Momentum', 'Low Vol']


def compute_ff_factor_returns_daily(px):
    """일별 long-short 팩터 수익률 (DataFrame)."""
    out = {}
    if 'SPY' not in px.columns:
        return None
    spy_ret = px['SPY'].pct_change()
    for f_name, (long_sym, short_sym) in FF_FACTOR_SYMS.items():
        if long_sym not in px.columns:
            continue
        long_ret = px[long_sym].pct_change()
        if short_sym is None:
            out[f_name] = long_ret
        else:
            if short_sym not in px.columns:
                continue
            out[f_name] = long_ret - px[short_sym].pct_change()
    if not out:
        return None
    return pd.DataFrame(out).dropna()


@st.cache_data(ttl=3600, show_spinner=False)
def compute_ff_betas_for_etfs(px_signature, etf_symbols_tuple, _px):
    """각 ETF에 대해 FF 6-factor 회귀 → {sym: {factor: beta}} 반환.

    Y = ETF 일별 수익률
    X = [1, MKT, Value, Size, Quality, Momentum, Low Vol]
    OLS via np.linalg.lstsq.
    """
    factor_rets = compute_ff_factor_returns_daily(_px)
    if factor_rets is None or len(factor_rets) < 60:
        return {}
    cols = [c for c in FF_FACTOR_ORDER if c in factor_rets.columns]
    results = {}
    for sym in etf_symbols_tuple:
        if sym not in _px.columns:
            continue
        etf_ret = _px[sym].pct_change()
        merged = pd.concat([etf_ret.rename('Y'), factor_rets[cols]],
                           axis=1).dropna()
        if len(merged) < 60:
            continue
        Y = merged['Y'].values
        X = np.column_stack([np.ones(len(Y))] + [merged[c].values for c in cols])
        try:
            betas, *_ = np.linalg.lstsq(X, Y, rcond=None)
        except Exception:
            continue
        if np.any(np.isnan(betas)):
            continue
        results[sym] = {'alpha': float(betas[0])}
        for i, c in enumerate(cols):
            results[sym][c] = float(betas[i + 1])
    return results


def compute_portfolio_ff_betas(weights, betas_map):
    """포트폴리오 net factor beta = ∑ (slv_w × etf_beta).
    Returns dict {factor: net_beta} where slv_w is sleeve fraction.
    """
    eq_sleeve_total = sum(w for tk, w in weights.items() if tk in FACTOR_LOOKTHROUGH_MSCI)
    if eq_sleeve_total == 0:
        return {}
    out = {f: 0.0 for f in FF_FACTOR_ORDER}
    for tk, w in weights.items():
        if tk not in FACTOR_LOOKTHROUGH_MSCI:
            continue
        sym = TICKER_MAP.get(tk)
        if sym is None or sym not in betas_map:
            continue
        slv_w_frac = w / eq_sleeve_total  # fraction (0~1)
        for f in FF_FACTOR_ORDER:
            out[f] += slv_w_frac * betas_map[sym].get(f, 0.0)
    return out


def compute_factor_contribution_ff(weights, betas_map, factor_period_returns_pct):
    """팩터 기여(bps) = portfolio_net_beta × factor_period_return × 100.

    factor_period_returns_pct: {factor: r%} — 해당 기간 누적 팩터 수익률 (%).
    """
    net_betas = compute_portfolio_ff_betas(weights, betas_map)
    return {f: net_betas.get(f, 0.0) * factor_period_returns_pct.get(f, 0.0) * 100
            for f in FF_FACTOR_ORDER}


# Bond ETF metadata: (Effective Duration in years, Yield to Maturity in %)
BOND_METADATA = {
    # KR bonds
    'a357870':           {'name': 'TIGER CD금리 (KR)',   'duration': 0.1, 'ytm': 3.5},
    '357870 ks equity':  {'name': 'TIGER CD금리 (KR)',   'duration': 0.1, 'ytm': 3.5},
    'a114820':           {'name': 'TIGER 국채3년 (KR)',   'duration': 2.7, 'ytm': 3.4},
    '114820 ks equity':  {'name': 'TIGER 국채3년 (KR)',   'duration': 2.7, 'ytm': 3.4},
    'a451540':           {'name': 'TIGER 종합채권 (KR)',  'duration': 5.0, 'ytm': 3.8},
    '451540 ks equity':  {'name': 'TIGER 종합채권 (KR)',  'duration': 5.0, 'ytm': 3.8},
    # US bonds
    'gto us equity':  {'name': 'Invesco Total Return Bond',  'duration': 5.5, 'ytm': 5.0},
    'hyg us equity':  {'name': 'iShares US HY',              'duration': 3.2, 'ytm': 7.5},
    'lqd us equity':  {'name': 'iShares IG Corp Bond',       'duration': 8.4, 'ytm': 5.5},
    'bkln us equity': {'name': 'Invesco Senior Loan',        'duration': 0.3, 'ytm': 6.5},
    'ief us equity':  {'name': 'iShares 7-10Y Treasury',     'duration': 7.5, 'ytm': 4.5},
    'embd us equity': {'name': 'Global X EM Bond',           'duration': 7.0, 'ytm': 8.0},
    'phyl us equity': {'name': 'PGIM Active HY Bond',        'duration': 3.5, 'ytm': 7.2},
}


import calendar as _calendar

def is_last_business_day_of_month(d):
    """True if date d is the last weekday (Mon-Fri) of its month."""
    last_day = _calendar.monthrange(d.year, d.month)[1]
    for day in range(last_day, 0, -1):
        weekday = d.replace(day=day).weekday()
        if weekday < 5:
            return d.day == day
    return False


def _period_ret(series, start_ts, end_ts):
    s = series[(series.index >= start_ts) & (series.index <= end_ts)].dropna()
    if len(s) < 2: return 0.0
    return (s.iloc[-1] / s.iloc[0] - 1) * 100


def _get_period_data(px, report_date, period='weekly'):
    end_ts = pd.Timestamp(report_date)
    avail = px.index[px.index <= end_ts]
    if len(avail) < 2: return None
    end_ts = avail[-1]
    if period == 'monthly':
        month_start = end_ts.replace(day=1)
        in_month = avail[avail >= month_start]
        if len(in_month) == 0: return None
        start_ts = in_month[0]
    else:
        if len(avail) < 6: return None
        start_ts = avail[-6]
    def safe(col): return _period_ret(px[col], start_ts, end_ts) if col in px.columns else 0.0
    return {
        'start': start_ts, 'end': end_ts,
        'spy':  safe('SPY'),  'qqq': safe('QQQ'),
        'acwi': safe('ACWI'), 'efa': safe('EFA'), 'vwo': safe('VWO'),
        'usdkrw': safe('USDKRW=X'),
        # Country ETFs (for country attribution narrative)
        'ewc': safe('EWC'), 'ewj': safe('EWJ'), 'mchi': safe('MCHI'),
        'ewu': safe('EWU'), 'ewl': safe('EWL'), 'ewy': safe('EWY'),
        # Bond ETFs (for bond narrative)
        'lqd': safe('LQD'), 'ief': safe('IEF'),
        'hyg': safe('HYG'), 'phyl': safe('PHYL'),
        'embd': safe('EMBD'),
    }


def make_weekly_commentary(data, signal, signal_date, agg_itd, neu_itd, agg_wtd, neu_wtd):
    if data is None:
        return '※ 데이터가 충분하지 않아 commentary를 생성할 수 없습니다.'
    s = data['start'].strftime('%m/%d'); e = data['end'].strftime('%m/%d')
    assets = [('S&P 500', data['spy']), ('나스닥100', data['qqq']),
              ('선진국(ex-US)', data['efa']), ('신흥국', data['vwo'])]
    assets.sort(key=lambda x: x[1], reverse=True)
    best, worst = assets[0], assets[-1]
    avg = sum(r for _, r in assets) / len(assets)
    sentiment = ('강세' if avg > 1.5 else '강보합' if avg > 0
                 else '약보합' if avg > -1.5 else '약세')
    fx = data['usdkrw']
    fx_msg = (f'원/달러 환율은 {fx:+.2f}% 변동하며 원화 {"약세 압력이 지속" if fx > 0.3 else "강세 흐름" if fx < -0.3 else "안정세"}'
              f'를 보였고, USD 보유 자산은 {"환차익" if fx > 0 else "환손실"}이 발생했습니다.')
    # Tech vs broad market spread
    tech_gap = data['qqq'] - data['spy']
    tech_msg = (f'기술주 vs 광역시장 스프레드(QQQ−SPY)는 {tech_gap:+.2f}%p로 '
                f'{"성장주 우위" if tech_gap > 0.5 else "광역시장 우위" if tech_gap < -0.5 else "균형 흐름"}'
                f'을 나타냈습니다.')
    rule_msg = ('Bull 시그널 하 50% SPY→QQQ 스왑 룰이 정상 작동' if signal == 'Bull'
                else 'Base/Bear 시그널로 룰 미적용')
    text = (
        f'**📅 주간 시장 요약 ({s} ~ {e}, 5영업일)**\n\n'
        f'지난주 글로벌 주식시장은 평균 **{avg:+.2f}%**로 **{sentiment}** 흐름을 나타냈습니다. '
        f'지수별로는 **{best[0]}** {best[1]:+.2f}%로 가장 우수했고, **{worst[0]}** {worst[1]:+.2f}%로 부진했습니다. '
        f'세부적으로 S&P 500 {data["spy"]:+.2f}%, 나스닥100 {data["qqq"]:+.2f}%, '
        f'선진국 ex-US(EFA) {data["efa"]:+.2f}%, 신흥국(VWO) {data["vwo"]:+.2f}%, MSCI ACWI {data["acwi"]:+.2f}%를 기록했습니다. '
        f'{tech_msg} {fx_msg} '
        f'AIMVP 시그널은 **{signal}** 상태(발효일 {signal_date})를 유지하고 있으며, {rule_msg}되었습니다. '
        f'주간(WTD) 운용 성과는 적극형 {agg_wtd:+.2f}%, 중립형 {neu_wtd:+.2f}%를 기록했고, '
        f'펀드 출범 이후 누적은 적극형 {agg_itd:+.2f}%, 중립형 {neu_itd:+.2f}%로 안정적인 흐름을 이어가고 있습니다.'
    )
    return text


def make_monthly_commentary(data, signal, signal_date,
                            agg_eq_pct, neu_eq_pct,
                            countries_held, bonds_held=None):
    """월간 펀드 매니저 commentary — 조회월 ETF 가격·종목 데이터만으로 자동 생성.

    포맷 참고: MVP시리즈 펀드 매니저 Comment 형식 (구조만 차용).
    외부 거시 이벤트(지정학·정책 등) 가정 없이 ETF 수익률 데이터에서 직접 도출.
    """
    bonds_held = bonds_held or set()
    if data is None:
        return '※ 데이터가 충분하지 않아 월간 commentary를 생성할 수 없습니다.'

    month_label = f"{data['end'].year}년 {data['end'].month}월"
    s_md = f"{data['start'].month}/{data['start'].day}"
    e_md = f"{data['end'].month}/{data['end'].day}"

    # === 시장 톤 (data-driven) ===
    avg_eq = (data['spy'] + data['qqq'] + data['acwi'] + data['efa'] + data['vwo']) / 5
    tech_gap = data['qqq'] - data['spy']

    if avg_eq > 2:
        market_tone = '글로벌 증시는 전반적으로 강세 흐름을 보였습니다'
    elif avg_eq > 0.5:
        market_tone = '글로벌 증시는 상승세를 나타냈습니다'
    elif avg_eq > -0.5:
        market_tone = '글로벌 증시는 제한적 등락 속 횡보 흐름을 보였습니다'
    elif avg_eq > -2:
        market_tone = '글로벌 증시는 약세 압력이 확대되었습니다'
    else:
        market_tone = '글로벌 증시는 큰 폭의 조정 흐름을 보였습니다'

    # US narrative (SPY vs QQQ split, no policy/sector assumption)
    if data['qqq'] > 1 and tech_gap > 0.5:
        us_narrative = (f'미국 증시는 기술주 강세 흐름(QQQ {data["qqq"]:+.2f}%, '
                        f'S&P 500 {data["spy"]:+.2f}%)이 두드러졌고')
    elif data['spy'] > 0 and data['qqq'] > 0:
        us_narrative = (f'미국 증시는 대형주 전반의 견조한 상승(S&P 500 {data["spy"]:+.2f}%, '
                        f'QQQ {data["qqq"]:+.2f}%) 흐름이 이어졌고')
    elif data['spy'] < -2:
        us_narrative = (f'미국 증시는 큰 폭의 조정(S&P 500 {data["spy"]:+.2f}%, '
                        f'QQQ {data["qqq"]:+.2f}%) 흐름을 보였고')
    elif data['spy'] < 0 and data['qqq'] < 0:
        us_narrative = (f'미국 증시는 약세 흐름(S&P 500 {data["spy"]:+.2f}%, '
                        f'QQQ {data["qqq"]:+.2f}%)을 나타냈고')
    else:
        us_narrative = (f'미국 증시는 혼조세(S&P 500 {data["spy"]:+.2f}%, '
                        f'QQQ {data["qqq"]:+.2f}%)를 보였고')

    # International narrative
    efa_r = data['efa']; vwo_r = data['vwo']
    if efa_r > 0.5 and vwo_r > 0.5:
        intl_narrative = (f'선진국(EFA {efa_r:+.2f}%)과 신흥국(VWO {vwo_r:+.2f}%) 증시도 '
                          f'동반 상승하였습니다')
    elif efa_r > 0 and vwo_r > 0:
        intl_narrative = (f'선진국(EFA {efa_r:+.2f}%)과 신흥국(VWO {vwo_r:+.2f}%) 증시 '
                          f'모두 상승하였습니다')
    elif efa_r > 0:
        intl_narrative = (f'선진국(EFA {efa_r:+.2f}%) 증시는 상승하였으나 '
                          f'신흥국(VWO {vwo_r:+.2f}%)은 상대적으로 부진하였습니다')
    elif vwo_r > 0:
        intl_narrative = (f'신흥국(VWO {vwo_r:+.2f}%)이 상대적으로 견조한 반면 '
                          f'선진국(EFA {efa_r:+.2f}%)은 약세를 시현하였습니다')
    else:
        intl_narrative = (f'선진국(EFA {efa_r:+.2f}%)과 신흥국(VWO {vwo_r:+.2f}%) 증시 '
                          f'모두 약세를 시현하였습니다')

    # US 핵심 ETF 성과 (S&P 500 ETF / QQQ)
    us_avg = (data['spy'] + data['qqq']) / 2
    if us_avg > 1:
        us_etf_sent = (f'주식 포트폴리오 내 주요 편입 대상인 **S&P 500 ETF**(SPY {data["spy"]:+.2f}%)와 '
                       f'**나스닥 100 ETF**(QQQ {data["qqq"]:+.2f}%)는 양호한 성과를 시현하며 '
                       f'포트폴리오 성과에 긍정적으로 기여하였습니다')
    elif us_avg > 0:
        us_etf_sent = (f'주식 포트폴리오 내 주요 편입 대상인 **S&P 500 ETF**(SPY {data["spy"]:+.2f}%)와 '
                       f'**나스닥 100 ETF**(QQQ {data["qqq"]:+.2f}%)는 소폭 상승하며 '
                       f'포트폴리오 성과에 기여하였습니다')
    elif us_avg > -1:
        us_etf_sent = (f'주식 포트폴리오 내 주요 편입 대상인 **S&P 500 ETF**(SPY {data["spy"]:+.2f}%)와 '
                       f'**나스닥 100 ETF**(QQQ {data["qqq"]:+.2f}%)는 혼조 흐름을 보였습니다')
    else:
        us_etf_sent = (f'주식 포트폴리오 내 주요 편입 대상인 **S&P 500 ETF**(SPY {data["spy"]:+.2f}%)와 '
                       f'**나스닥 100 ETF**(QQQ {data["qqq"]:+.2f}%)는 부진한 성과를 기록하며 '
                       f'포트폴리오 성과에 부정적으로 작용하였습니다')

    # 글로벌 ETFs (ACWI/EFA/EM)
    glob_avg = (data['acwi'] + efa_r + vwo_r) / 3
    glob_dir = '상승' if glob_avg > 0 else '하락'
    glob_impact = '긍정적' if glob_avg > 0 else '부정적'
    global_etf_sent = (
        f'또한 전반적인 글로벌 증시에 투자하는 **MSCI ACWI ETF**(ACWI {data["acwi"]:+.2f}%), '
        f'미국 외 선진국 증시 **MSCI EAFE ETF**(EFA {efa_r:+.2f}%), '
        f'**이머징 증시 ETF**(VWO {vwo_r:+.2f}%)도 {glob_dir}하며 '
        f'포트폴리오 성과에 {glob_impact}으로 기여하였습니다'
    )

    # 국가 배분 narrative — 실제 보유 국가 ETF의 best/worst 자동 추출
    country_returns = []
    for ctry, ret_key in [('캐나다', 'ewc'), ('일본', 'ewj'), ('중국', 'mchi'),
                          ('영국', 'ewu'), ('스위스', 'ewl'), ('한국', 'ewy')]:
        if ctry in countries_held:
            country_returns.append((ctry, data.get(ret_key, 0)))

    if country_returns:
        country_returns.sort(key=lambda x: x[1], reverse=True)
        positives = [c for c in country_returns if c[1] > 0]
        negatives = [c for c in country_returns if c[1] <= 0]
        if positives and not negatives:
            top_names = ', '.join(f'{c}({r:+.2f}%)' for c, r in positives[:3])
            country_sentence = (f'국가 배분 전략 관점에서 전술적으로 편입한 '
                                f'**{top_names}** 증시 ETF는 포트폴리오 성과를 '
                                f'더해준 요인으로 나타났습니다.')
        elif positives and negatives:
            top_names = ', '.join(f'{c}({r:+.2f}%)' for c, r in positives[:2])
            bot_names = ', '.join(f'{c}({r:+.2f}%)' for c, r in negatives[-2:])
            country_sentence = (f'국가 배분 측면에서 **{top_names}** 증시 ETF는 성과 기여 요인으로 '
                                f'작용하였으며, {bot_names} 증시는 상대적으로 부진하여 '
                                f'일부 차감 요인으로 나타났습니다.')
        else:
            bot_names = ', '.join(f'{c}({r:+.2f}%)' for c, r in negatives[:3])
            country_sentence = (f'국가 배분 측면에서는 **{bot_names}** 증시 ETF가 '
                                f'전반적으로 약세를 보이며 차감 요인으로 작용하였습니다.')
    else:
        country_sentence = ''

    # ── 채권 narrative (실제 보유 채권만 언급) ──
    has_hyg     = 'hyg us equity'  in bonds_held
    has_phyl    = 'phyl us equity' in bonds_held
    has_lqd     = 'lqd us equity'  in bonds_held
    has_embd    = 'embd us equity' in bonds_held
    has_kr_long = 'a451540' in bonds_held  # TIGER 종합채권
    has_kr_cd   = 'a357870' in bonds_held  # TIGER CD금리

    ief_r = data.get('ief', 0)

    # 채권 macro tone (IEF as 7-10Y rate proxy)
    if ief_r > 1:
        bond_macro = f'중장기 국채 금리는 큰 폭으로 하락(IEF {ief_r:+.2f}%)하며 채권 가격은 강세를 시현하였고'
    elif ief_r > 0.3:
        bond_macro = f'중장기 국채 금리는 소폭 하락(IEF {ief_r:+.2f}%)하며 채권 가격은 강보합 흐름을 보였고'
    elif ief_r > -0.3:
        bond_macro = f'중장기 국채 금리는 좁은 박스권에서 등락(IEF {ief_r:+.2f}%)을 보였고'
    elif ief_r > -1:
        bond_macro = f'중장기 국채 금리는 상승 압력(IEF {ief_r:+.2f}%)이 이어지며 채권 가격은 약보합세를 보였고'
    else:
        bond_macro = f'중장기 국채 금리는 큰 폭의 상승(IEF {ief_r:+.2f}%)을 보이며 채권 가격은 약세를 기록하였고'

    # 채권 변동성 (절대값 기반)
    bond_vol = ('높은 금리 변동성 장세로 전개되었습니다'
                if abs(ief_r) > 1.5 else '제한적 변동성 흐름을 이어갔습니다')

    # KR 종합채권 (중장기 듀레이션, IEF 방향과 유사하게 작용)
    if has_kr_long:
        if ief_r > 0.5:
            kr_long_sent = (f'이러한 금리 흐름에 따라 중장기 듀레이션을 보유한 '
                            f'**TIGER 종합채권(AA-이상)액티브 ETF**는 양호한 성과를 시현하였습니다. ')
        elif ief_r < -0.5:
            kr_long_sent = (f'이러한 금리 상승 압력으로 중장기 듀레이션을 보유한 '
                            f'**TIGER 종합채권(AA-이상)액티브 ETF**는 다소 부진한 성과를 기록하였습니다. ')
        else:
            kr_long_sent = (f'중장기 듀레이션을 보유한 **TIGER 종합채권(AA-이상)액티브 ETF**는 '
                            f'박스권 흐름을 이어갔습니다. ')
    else:
        kr_long_sent = ''

    # 크레딧·EM 채권 ETFs (보유 종목만 평균)
    credit_rets = []
    credit_etf_labels = []
    if has_lqd:
        credit_rets.append(data.get('lqd', 0))
        credit_etf_labels.append(f'**미국 투자등급회사채 ETF**(LQD {data.get("lqd", 0):+.2f}%)')
    hy_labels = []
    if has_hyg:
        credit_rets.append(data.get('hyg', 0))
        hy_labels.append(f'HYG {data.get("hyg", 0):+.2f}%')
    if has_phyl:
        credit_rets.append(data.get('phyl', 0))
        hy_labels.append(f'PHYL {data.get("phyl", 0):+.2f}%')
    if hy_labels:
        credit_etf_labels.append(f'**미국 투기등급회사채 ETF**({", ".join(hy_labels)})')
    if has_embd:
        credit_rets.append(data.get('embd', 0))
        credit_etf_labels.append(f'**이머징 채권 ETF**(EMBD {data.get("embd", 0):+.2f}%)')

    credit_avg = sum(credit_rets) / len(credit_rets) if credit_rets else 0
    if credit_etf_labels:
        if credit_avg > 0.3:
            credit_outcome = '양호한 성과를 시현하며 채권 포트폴리오 성과에 긍정적으로 기여하였습니다'
        elif credit_avg < -0.3:
            credit_outcome = '부진한 성과를 기록하며 채권 포트폴리오 성과에 부정적으로 작용하였습니다'
        else:
            credit_outcome = '중립적인 흐름을 보이며 채권 포트폴리오 성과 안정성에 기여하였습니다'
        credit_sentence = f'{" 및 ".join(credit_etf_labels)}는 {credit_outcome}. '
    else:
        credit_sentence = ''

    # KR CD (단기금리, 안정성)
    kr_cd_sent = (
        '또한 **TIGER CD금리투자KIS(합성) ETF**의 일정 비중을 유지한 점은 '
        '꾸준한 단기금리 수익을 통해 채권 포트폴리오의 변동성을 완화하는 요인으로 작용하였습니다. '
        if has_kr_cd else ''
    )

    # FX impact
    fx = data['usdkrw']
    if fx > 1:
        fx_tone = (f'원달러 환율은 월초 대비 큰 폭으로 상승({fx:+.2f}%)하며 '
                   f'USD 표시 자산의 환차익이 발생해 전체 포트폴리오 성과에 긍정적으로 작용하였습니다')
    elif fx > 0.3:
        fx_tone = (f'원달러 환율은 월초 대비 상승({fx:+.2f}%)하며 USD 표시 자산의 환차익이 발생해 '
                   f'전체 포트폴리오 성과에 소폭 긍정적으로 작용하였습니다')
    elif fx < -1:
        fx_tone = (f'원달러 환율은 월초 대비 큰 폭으로 하락({fx:+.2f}%)하며 '
                   f'USD 표시 자산의 환손실이 발생해 전체 포트폴리오 성과에 부정적으로 작용하였습니다')
    elif fx < -0.3:
        fx_tone = (f'원달러 환율은 월초 대비 하락({fx:+.2f}%)하며 USD 표시 자산의 환손실이 발생해 '
                   f'전체 포트폴리오 성과에 소폭 부정적으로 작용하였습니다')
    else:
        fx_tone = (f'원달러 환율은 월초 대비 큰 변동 없이({fx:+.2f}%) '
                   f'안정적인 흐름을 유지하였습니다')

    # === Compose ===
    text = (
        f'**📰 MVP시리즈 펀드 매니저 Comment ({month_label})**\n\n'
        f'**ETF AI MVP (적극), ETF AI MVP (중립)**\n\n'
        f'ETF AI MVP (적극), ETF AI MVP (중립) 펀드는 위험자산의 비중을 각각 '
        f'**{agg_eq_pct:.1f}%, {neu_eq_pct:.1f}%**로 운용하였으며, AIMVP 시그널은 '
        f'**{signal}** 상태(발효일 {signal_date})를 유지하였습니다. '
        f'{month_label}({s_md}~{e_md}) {market_tone}. '
        f'{us_narrative}, {intl_narrative}. '
        f'{us_etf_sent}. {global_etf_sent}. {country_sentence}\n\n'
        f'채권 시장의 경우 {bond_macro} {bond_vol} '
        f'{kr_long_sent}'
        f'{credit_sentence}'
        f'{kr_cd_sent}'
        f'한편 {fx_tone}.'
    )
    return text


def compute_bond_metrics(weights):
    """Compute weighted average duration/YTM and bucket breakdowns for bond holdings.
    Returns dict with: wavg_dur, wavg_ytm, dur_buckets, ytm_buckets, details (list)."""
    bond_total = 0
    wsum_dur = 0
    wsum_ytm = 0
    details = []
    for tk, w in weights.items():
        meta = BOND_METADATA.get(tk)
        if not meta or w <= 0: continue
        bond_total += w
        wsum_dur += w * meta['duration']
        wsum_ytm += w * meta['ytm']
        details.append({
            'name': meta['name'],
            'weight': w,
            'duration': meta['duration'],
            'ytm': meta['ytm'],
        })
    if bond_total == 0:
        return None
    wavg_dur = wsum_dur / bond_total
    wavg_ytm = wsum_ytm / bond_total

    # Duration buckets
    dur_buckets = {'단기 (0~3년)': 0, '중기 (3~7년)': 0, '장기 (7년+)': 0}
    for d in details:
        share = d['weight'] / bond_total * 100
        if d['duration'] < 3:
            dur_buckets['단기 (0~3년)'] += share
        elif d['duration'] < 7:
            dur_buckets['중기 (3~7년)'] += share
        else:
            dur_buckets['장기 (7년+)'] += share

    # YTM buckets
    ytm_buckets = {'저금리 (0~4%)': 0, '중금리 (4~6%)': 0, '고금리 (6~8%)': 0, '초고금리 (8%+)': 0}
    for d in details:
        share = d['weight'] / bond_total * 100
        if d['ytm'] < 4:
            ytm_buckets['저금리 (0~4%)'] += share
        elif d['ytm'] < 6:
            ytm_buckets['중금리 (4~6%)'] += share
        elif d['ytm'] < 8:
            ytm_buckets['고금리 (6~8%)'] += share
        else:
            ytm_buckets['초고금리 (8%+)'] += share

    # Sort details by weight desc
    details.sort(key=lambda x: x['weight'], reverse=True)

    return {
        'bond_total':  bond_total,
        'wavg_dur':    wavg_dur,
        'wavg_ytm':    wavg_ytm,
        'dur_buckets': dur_buckets,
        'ytm_buckets': ytm_buckets,
        'details':     details,
    }


TICKER_CATEGORY = {
    'acwi us equity': '광역 주식', 'efa us equity': '광역 주식',
    'iwm us equity': '광역 주식', 'qqq us equity': '광역 주식',
    'spy us equity': '광역 주식', 'vwo us equity': '광역 주식',
    'ewa us equity':'국가 주식','ewc us equity':'국가 주식','enzl us equity':'국가 주식',
    'ewg us equity':'국가 주식','norw us equity':'국가 주식','ewo us equity':'국가 주식',
    'ewh us equity':'국가 주식','ews us equity':'국가 주식','ewi us equity':'국가 주식',
    'ewn us equity':'국가 주식','ewj us equity':'국가 주식','mchi us equity':'국가 주식',
    'ewk us equity':'국가 주식','ewq us equity':'국가 주식','ewu us equity':'국가 주식',
    'ewl us equity':'국가 주식','ewz us equity':'국가 주식','inda us equity':'국가 주식',
    'ewy us equity':'국가 주식','eden us equity':'국가 주식','eis us equity':'국가 주식',
    'ewd us equity':'국가 주식','ewp us equity':'국가 주식',
    'pdbc us equity':'원자재','dbmf us equity':'헤지펀드',
    'gto us equity':'채권','hyg us equity':'채권','lqd us equity':'채권',
    'bkln us equity':'채권','ief us equity':'채권','embd us equity':'채권',
    'phyl us equity':'채권',
    'a357870':'채권','357870 ks equity':'채권',
    'a114820':'채권','114820 ks equity':'채권',
    'a451540':'채권','451540 ks equity':'채권',
    'usdkrw curncy':'현금','cash_krw':'현금',
}


def std_ticker(t):
    if not t: return None
    return str(t).strip().lower()

def parse_date(v):
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    return str(v).split(' ')[0]


# Excel 내 소계(subtotal) 행 — 실제 보유종목이 아니므로 파싱에서 제외
# (구 시트 2022~23에 '주식합/채권합/헤지펀드합'이 보유처럼 섞여 있어 정규화 왜곡 유발)
SUBTOTAL_LABELS = {'주식합', '채권합', '헤지펀드합', '현금합', '원자재합',
                   '대안합', '주식소계', '채권소계', '합계', '소계', '총합'}


@st.cache_data(ttl=3600)
def parse_holdings(profile='적극'):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    def parse_phase(ws, sig_row, date_row, asset_rows):
        sigs = list(ws[sig_row]); dates = list(ws[date_row])
        asset_data = {}; tickers = []
        for ri in asset_rows:
            row = list(ws[ri])
            tk = std_ticker(row[2].value) or std_ticker(row[1].value)
            if not tk or tk in SUBTOTAL_LABELS: continue
            asset_data[tk] = [c.value for c in row]
            tickers.append(tk)
        out = []
        for i in range(3, len(sigs)):
            sv = sigs[i].value; dv = dates[i].value
            if sv not in ('Bull','Base','Bear'): continue
            w = {}
            for tk in tickers:
                v = asset_data[tk][i]
                if isinstance(v,(int,float)) and v > 0:
                    w[tk] = float(v)
            out.append((parse_date(dv), sv, w))
        return out

    if profile == '적극':
        ws1 = wb['AIMVP포트폴리오현황(국가별자산배분이전)']
        ph1 = parse_phase(ws1, 3, 5, range(6, 22))
        ws2 = wb['AIMVP포트폴리오현황(국가별자산배분이후)']
        # 2026-06-01: PHYL row inserted at row 43 → range extended by 1
        ph2 = parse_phase(ws2, 4, 6, range(8, 49))
    else:
        ws1 = wb['AIMVP포트폴리오현황(국가별자산배분이전)']
        ph1 = parse_phase(ws1, 29, 31, range(33, 49))
        ws2 = wb['AIMVP포트폴리오현황(국가별자산배분이후)']
        # 2026-06-01: PHYL inserts shifted 중립 section: signal 53→54, date 55→56, assets 57→58, end 97→99
        ph2 = parse_phase(ws2, 54, 56, range(58, 99))

    seen = set(); combined = []
    for r in ph1 + ph2:
        if r[0] not in seen:
            combined.append(r); seen.add(r[0])
    combined.sort(key=lambda x: x[0])
    return combined


@st.cache_data(ttl=3600)
def fetch_prices(end_date):
    yf_tickers = set()
    for prof in ['적극', '중립']:
        for _,_,w in parse_holdings(prof):
            for tk in w:
                sym = TICKER_MAP.get(tk)
                if sym and not sym.startswith('_'):
                    yf_tickers.add(sym)
    # 투자유니버스 시트의 모든 ETF — 보유 이력 없어도 Universe 페이지 return 산출용
    try:
        wb_u = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        if '투자유니버스' in wb_u.sheetnames:
            ws_u = wb_u['투자유니버스']
            for i in range(3, ws_u.max_row + 1):
                ticker_v = ws_u.cell(i, 4).value
                if not ticker_v:
                    continue
                tk_u = std_ticker(ticker_v)
                sym_u = TICKER_MAP.get(tk_u)
                if sym_u and not sym_u.startswith('_'):
                    yf_tickers.add(sym_u)
    except Exception:
        pass
    yf_tickers |= {'SPY', 'QQQ', 'ACWI', 'USDKRW=X'}  # add FX for USD cash
    # Factor proxy ETFs (for Fama-French style regression)
    yf_tickers |= {'VLUE', 'IWM', 'QUAL', 'MTUM', 'USMV'}
    # Benchmark proxies (중립형: Bloomberg Global Aggregate Bond)
    yf_tickers |= {'BNDW'}  # Vanguard Total World Bond — Bloomberg Global Agg proxy
    # 5Y CAGR 산출을 위해 start를 5년+여유 만큼 앞당김 (기존 분석 영향 없음 — 모두 date slice 사용)
    px = yf.download(sorted(yf_tickers), start='2021-01-01', end=end_date,
                     auto_adjust=True, progress=False)['Close']
    px = px.ffill()
    KRBOND_DAILY = (1.035)**(1/252) - 1

    # USD cash (KRW-denominated fund holding USD) → tracks USDKRW exchange rate
    if 'USDKRW=X' in px.columns:
        px['_CASH_USD'] = px['USDKRW=X'].copy()
    else:
        px['_CASH_USD'] = 1.0  # fallback

    # KRW cash → no return (flat)
    px['_CASH_KR'] = 1.0

    # KR bonds (proxy) — small interest accrual
    px['_KRBOND'] = (1 + KRBOND_DAILY) ** np.arange(len(px))
    return px


def simulate_portfolio(rebals, px, swap_ratio=0.0, swap_during='Bull', end_date=None):
    """연속 일별 포트폴리오 수익률 시계열 (fraction).

    각 리밸 weights를 [리밸일, 다음 리밸일) 동안 적용하되, **리밸 경계의 일별 수익률을
    누락 없이 연속 체인**한다. (구버전은 윈도우별 pct_change().dropna() 후 concat 하여
    매 리밸 경계의 turn-of-month 수익률 1일을 통째로 누락 → 누적이 체계적으로 과소계상됐음.)

    규약: rebalance-at-close — 리밸일 **당일** 수익률(전일→리밸일)은 **직전 비중**으로 산출하고,
    리밸일 **다음 거래일**부터 새 비중 적용. (단일자산 100% 시 직접 가격비율과 정확히 일치.)
    """
    if not rebals:
        return pd.Series(dtype=float)
    end_ts = pd.Timestamp(end_date) if end_date else px.index.max()
    start_ts = pd.Timestamp(rebals[0][0])
    idx = px.index[(px.index >= start_ts) & (px.index <= end_ts)]
    if len(idx) < 2:
        return pd.Series(dtype=float)

    # 각 리밸 시점의 정규화된 sym-weight (swap 적용)
    weight_rows = {}
    for date_str, signal, weights in rebals:
        rb = pd.Timestamp(date_str)
        if rb > end_ts:
            break
        w_adj = dict(weights)
        if signal == swap_during and 'spy us equity' in w_adj and swap_ratio > 0:
            spy_w = w_adj.get('spy us equity', 0)
            qqq_w = w_adj.get('qqq us equity', 0)
            w_adj['spy us equity'] = spy_w * (1 - swap_ratio)
            w_adj['qqq us equity'] = qqq_w + spy_w * swap_ratio
        mapped = {}
        for tk, w in w_adj.items():
            sym = TICKER_MAP.get(tk)
            if sym is None or sym not in px.columns:
                continue
            mapped[sym] = mapped.get(sym, 0) + w
        tot = sum(mapped.values())
        if tot <= 0:
            continue
        weight_rows[rb] = {k: v / tot for k, v in mapped.items()}
    if not weight_rows:
        return pd.Series(dtype=float)

    syms = sorted({s for d in weight_rows.values() for s in d})
    # 리밸일 기준 weight 행렬 → 거래일로 ffill → shift(1) (리밸 당일 수익률엔 직전 비중)
    W = pd.DataFrame(0.0, index=sorted(weight_rows), columns=syms)
    for rb, dct in weight_rows.items():
        for s, w in dct.items():
            W.loc[rb, s] = w
    W = W.reindex(idx, method='ffill').shift(1)
    R = px[syms].reindex(idx).pct_change().fillna(0)
    port_daily = (W * R).sum(axis=1)
    return port_daily.iloc[1:]  # 첫날(앵커) 제거 — 나머지는 누락 없는 연속 수익률


def compute_metrics(daily_rets):
    cum = (1 + daily_rets).cumprod()
    today = cum.index[-1]; today_val = cum.iloc[-1]
    wtd_idx = cum.index[cum.index >= (today - pd.Timedelta(days=7))]
    wtd_ret = (today_val / cum.loc[wtd_idx[0]] - 1) * 100 if len(wtd_idx) > 1 else 0
    month_start = today.replace(day=1)
    mtd_idx = cum.index[cum.index >= month_start]
    mtd_ret = (today_val / cum.loc[mtd_idx[0]] - 1) * 100 if len(mtd_idx) > 1 else 0
    year_start = today.replace(month=1, day=1)
    ytd_idx = cum.index[cum.index >= year_start]
    ytd_ret = (today_val / cum.loc[ytd_idx[0]] - 1) * 100 if len(ytd_idx) > 1 else 0
    itd_ret = (today_val - 1) * 100
    ann_ret = (1 + daily_rets.mean())**252 - 1
    ann_vol = daily_rets.std() * np.sqrt(252)
    sharpe = ann_ret / ann_vol if ann_vol > 0 else 0
    mdd = ((cum / cum.cummax()) - 1).min() * 100
    downside = daily_rets[daily_rets < 0].std() * np.sqrt(252)
    sortino = ann_ret / downside if downside > 0 else 0
    return {
        'WTD': wtd_ret, 'MTD': mtd_ret, 'YTD': ytd_ret, 'ITD': itd_ret,
        'CAGR': ann_ret*100, 'Vol': ann_vol*100,
        'Sharpe': sharpe, 'Sortino': sortino, 'MaxDD': mdd,
    }


def compute_holdings_perf(rebals, px, end_date_str):
    """Holdings-level performance: 1D / 1W / MTD total return + contribution."""
    latest = rebals[-1]
    rebal_date = latest[0]
    weights = latest[2]
    end_ts = pd.Timestamp(end_date_str)

    # Determine reference dates
    px_up_to = px[px.index <= end_ts]
    if len(px_up_to) < 2:
        return pd.DataFrame()
    today = px_up_to.index[-1]
    d1 = px_up_to.index[-2] if len(px_up_to) >= 2 else today

    # 1W back: ~5 business days
    week_ago_idx = px_up_to.index[px_up_to.index <= (today - pd.Timedelta(days=7))]
    d1w = week_ago_idx[-1] if len(week_ago_idx) > 0 else px_up_to.index[0]

    # MTD: first business day of current month
    month_start = today.replace(day=1)
    mtd_idx = px_up_to.index[px_up_to.index >= month_start]
    d_mtd = mtd_idx[0] if len(mtd_idx) > 0 else today

    rows = []
    for tk, w in weights.items():
        sym = TICKER_MAP.get(tk)
        if sym is None: continue
        if sym not in px.columns: continue
        try:
            p_now  = px[sym].loc[today]
            p_1d   = px[sym].loc[d1]
            p_1w   = px[sym].loc[d1w]
            p_mtd  = px[sym].loc[d_mtd]
            r_1d   = (p_now / p_1d  - 1) * 100 if p_1d  > 0 else 0
            r_1w   = (p_now / p_1w  - 1) * 100 if p_1w  > 0 else 0
            r_mtd  = (p_now / p_mtd - 1) * 100 if p_mtd > 0 else 0
        except (KeyError, ZeroDivisionError):
            r_1d = r_1w = r_mtd = 0

        # Contribution = weight × return (in bps)
        c_1d  = w * r_1d  / 100 * 100  # bps
        c_1w  = w * r_1w  / 100 * 100
        c_mtd = w * r_mtd / 100 * 100

        rows.append({
            'Ticker': sym if not sym.startswith('_') else tk.upper(),
            '종목명': TICKER_NAMES.get(tk, tk),
            '구분': TICKER_CATEGORY.get(tk, '기타'),
            '비중(%)': round(w, 2),
            '1D Return(%)':  round(r_1d, 2),
            '1W Return(%)':  round(r_1w, 2),
            'MTD Return(%)': round(r_mtd, 2),
            '1D 기여(bps)':  round(c_1d, 1),
            '1W 기여(bps)':  round(c_1w, 1),
            'MTD 기여(bps)': round(c_mtd, 1),
        })
    df = pd.DataFrame(rows)
    if len(df) > 0:
        df = df.sort_values('비중(%)', ascending=False).reset_index(drop=True)
    return df, rebal_date, today.strftime('%Y-%m-%d')


def compute_comparison(rebals, px, end_date_str):
    """현재 리밸런스 vs 전월(직전) 리밸런스 비교: 비중 + MTD 기여."""
    if len(rebals) < 2:
        return None
    cur_rebal = rebals[-1]
    prev_rebal = rebals[-2]
    end_ts = pd.Timestamp(end_date_str)
    px_up = px[px.index <= end_ts]
    if len(px_up) < 2:
        return None
    today = px_up.index[-1]

    # 현재 MTD: 이번 달 1일 ~ 오늘
    cur_month_start = today.replace(day=1)
    cur_mtd_idx = px_up.index[px_up.index >= cur_month_start]
    cur_mtd_start = cur_mtd_idx[0] if len(cur_mtd_idx) > 0 else today

    # 전월 보유 기간: 직전 리밸 → 현재 리밸
    prev_start = pd.Timestamp(prev_rebal[0])
    prev_end_target = pd.Timestamp(cur_rebal[0])
    prev_active_idx = px_up.index[(px_up.index >= prev_start) & (px_up.index <= prev_end_target)]
    if len(prev_active_idx) < 2:
        return None
    prev_p0 = prev_active_idx[0]
    prev_p1 = prev_active_idx[-1]

    all_tk = set(cur_rebal[2].keys()) | set(prev_rebal[2].keys())
    rows = []
    for tk in all_tk:
        sym = TICKER_MAP.get(tk)
        if sym is None: continue
        if sym not in px.columns: continue
        w_cur  = cur_rebal[2].get(tk, 0)
        w_prev = prev_rebal[2].get(tk, 0)
        try:
            r_cur  = (px[sym].loc[today]  / px[sym].loc[cur_mtd_start] - 1) * 100
        except Exception:
            r_cur = 0
        try:
            r_prev = (px[sym].loc[prev_p1] / px[sym].loc[prev_p0]       - 1) * 100
        except Exception:
            r_prev = 0
        # 핵심 attribution (현재 MTD return으로 통일하여 비중 효과 분리)
        c_cur_w_cur  = w_cur  * r_cur / 100 * 100   # 현재 비중 × 현재 MTD return
        c_cur_w_prev = w_prev * r_cur / 100 * 100   # 전월 비중 × 현재 MTD return (counter-factual)
        rebal_impact = c_cur_w_cur - c_cur_w_prev   # = (w_cur - w_prev) × r_cur → 비중 조절 효과
        rows.append({
            'Ticker': sym if not sym.startswith('_') else tk.upper(),
            '종목명': TICKER_NAMES.get(tk, tk),
            '구분': TICKER_CATEGORY.get(tk, '기타'),
            '전월 비중(%)': round(w_prev, 2),
            '현재 비중(%)': round(w_cur, 2),
            '비중 변화(%p)': round(w_cur - w_prev, 2),
            'MTD(bps)': round(r_cur * 100, 1),  # 종목 MTD total return in bps
            'MTD 기여(bps, 현재비중)': round(c_cur_w_cur, 1),
            'MTD 기여(bps, 전월비중)': round(c_cur_w_prev, 1),
            '리밸런싱 효과(bps)': round(rebal_impact, 1),
        })
    df = pd.DataFrame(rows)
    if df.empty:
        return None
    CAT_ORDER = ['광역 주식', '국가 주식', '원자재', '헤지펀드', '채권', '현금', '기타']
    df['_cat'] = df['구분'].map({c: i for i, c in enumerate(CAT_ORDER)}).fillna(99)
    df = (df.sort_values(['_cat', '현재 비중(%)'], ascending=[True, False])
            .drop(columns='_cat').reset_index(drop=True))
    return {
        'df': df,
        'cur_rebal':  cur_rebal[0],  'cur_signal':  cur_rebal[1],
        'prev_rebal': prev_rebal[0], 'prev_signal': prev_rebal[1],
        'cur_period':  f'{cur_mtd_start.strftime("%Y-%m-%d")} ~ {today.strftime("%Y-%m-%d")}',
        'prev_period': f'{prev_p0.strftime("%Y-%m-%d")} ~ {prev_p1.strftime("%Y-%m-%d")}',
    }


def compute_rebal_effect_history(rebals, px, end_date_str):
    """운용시작일부터 조회일까지 각 리밸런스 시점의 비중 조절 효과(bps) 시계열.

    각 리밸 i 시점에서 ∑_t (w_{i,t} - w_{i-1,t}) × r_{i→i+1, t} (t = 종목).
    마지막 리밸은 다음 리밸 대신 end_date_str까지 partial 측정.
    """
    if len(rebals) < 2:
        return pd.DataFrame()
    end_ts = pd.Timestamp(end_date_str)
    px_up = px[px.index <= end_ts]
    if len(px_up) < 2:
        return pd.DataFrame()
    rows = []
    for i in range(1, len(rebals)):
        prev_rebal = rebals[i-1]
        cur_rebal = rebals[i]
        d_cur = pd.Timestamp(cur_rebal[0])
        if d_cur > end_ts:
            break
        if i + 1 < len(rebals):
            d_next = min(pd.Timestamp(rebals[i+1][0]), end_ts)
        else:
            d_next = end_ts
        idx = px_up.index[(px_up.index >= d_cur) & (px_up.index <= d_next)]
        if len(idx) < 2:
            continue
        p0 = idx[0]; p1 = idx[-1]
        all_tk = set(cur_rebal[2].keys()) | set(prev_rebal[2].keys())
        total_impact = 0.0
        for tk in all_tk:
            sym = TICKER_MAP.get(tk)
            if sym is None or sym not in px.columns:
                continue
            w_cur = cur_rebal[2].get(tk, 0)
            w_prev = prev_rebal[2].get(tk, 0)
            try:
                r = (px[sym].loc[p1] / px[sym].loc[p0] - 1) * 100
            except Exception:
                r = 0
            total_impact += (w_cur - w_prev) * r  # %p × % = bps
        rows.append({
            'date': d_cur,
            'signal': cur_rebal[1],
            'impact_bps': total_impact,
            'partial': i + 1 >= len(rebals),
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values('date').reset_index(drop=True)
        df['cumulative_bps'] = df['impact_bps'].cumsum()
    return df


EQUITY_SLEEVE_CATS    = {'광역 주식', '국가 주식'}
BOND_SLEEVE_CATS      = {'채권'}
DEFENSIVE_SLEEVE_CATS = {'채권', '현금'}
BROAD_EQ_TICKERS      = {'spy us equity', 'qqq us equity', 'iwm us equity',
                         'acwi us equity', 'efa us equity', 'vwo us equity'}


def compute_signal_asset_breakdown(rebals, px, end_date_str):
    """시그널을 독립 이벤트로 취급, 시그널×자산별 평균 비중/return/기여 산출."""
    if len(rebals) < 2:
        return None
    end_ts = pd.Timestamp(end_date_str)
    px_up = px[px.index <= end_ts]
    rows = []
    for i in range(len(rebals)):
        rb_date = pd.Timestamp(rebals[i][0])
        if rb_date > end_ts:
            break
        if i + 1 < len(rebals):
            next_date = min(pd.Timestamp(rebals[i+1][0]), end_ts)
        else:
            next_date = end_ts
        idx = px_up.index[(px_up.index >= rb_date) & (px_up.index <= next_date)]
        if len(idx) < 2:
            continue
        p0, p1 = idx[0], idx[-1]
        signal = rebals[i][1]
        weights = rebals[i][2]
        for tk, w in weights.items():
            sym = TICKER_MAP.get(tk)
            if sym is None or sym not in px.columns:
                continue
            try:
                r = (px[sym].loc[p1] / px[sym].loc[p0] - 1) * 100
            except Exception:
                r = 0
            rows.append({
                'signal': signal,
                'ticker': tk,
                'sym': sym.upper() if not sym.startswith('_') else tk.upper(),
                'name': TICKER_NAMES.get(tk, tk),
                'category': TICKER_CATEGORY.get(tk, '기타'),
                'weight': w,
                'return_pct': r,
                'contrib_bps': w * r,  # w(%) × r(%) = bps (portfolio level)
            })
    return pd.DataFrame(rows)


def compute_acwi_cf_alpha(rebals, px, end_date_str, sleeve='broad'):
    """슬리브 실제 수익률 vs ACWI 100% counter-factual 알파.

    sleeve='broad': 광역주식만 (SPY/QQQ/IWM/ACWI/EFA/VWO)
    sleeve='equity_all': 전체 주식 (광역 + 국가)
    슬리브 비중은 sleeve-normalize → 100%로 보고, CF = 100% ACWI.
    """
    if sleeve == 'broad':
        sleeve_tks = BROAD_EQ_TICKERS
    else:
        sleeve_tks = set(LOOKTHROUGH.keys())  # 모든 주식 (광역+국가)
    if len(rebals) < 2:
        return None
    end_ts = pd.Timestamp(end_date_str)
    px_up = px[px.index <= end_ts]
    if 'ACWI' not in px.columns:
        return None
    rows = []
    for i in range(len(rebals)):
        rb_date = pd.Timestamp(rebals[i][0])
        if rb_date > end_ts:
            break
        if i + 1 < len(rebals):
            next_date = min(pd.Timestamp(rebals[i+1][0]), end_ts)
        else:
            next_date = end_ts
        idx = px_up.index[(px_up.index >= rb_date) & (px_up.index <= next_date)]
        if len(idx) < 2:
            continue
        p0, p1 = idx[0], idx[-1]
        weights = rebals[i][2]
        sleeve_total = sum(w for tk, w in weights.items() if tk in sleeve_tks)
        if sleeve_total <= 0:
            continue
        # 실제 슬리브 수익률 (sleeve-normalized)
        actual_ret = 0.0
        for tk, w in weights.items():
            if tk not in sleeve_tks:
                continue
            sym = TICKER_MAP.get(tk)
            if sym is None or sym not in px.columns:
                continue
            try:
                r = (px[sym].loc[p1] / px[sym].loc[p0] - 1) * 100
                actual_ret += (w / sleeve_total) * r
            except Exception:
                pass
        try:
            r_acwi = (px['ACWI'].loc[p1] / px['ACWI'].loc[p0] - 1) * 100
        except Exception:
            r_acwi = 0
        rows.append({
            'date': rb_date,
            'signal': rebals[i][1],
            'sleeve_weight': sleeve_total,
            'actual_return': actual_ret,
            'acwi_return': r_acwi,
            'alpha': actual_ret - r_acwi,
        })
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df['cum_actual'] = ((1 + df['actual_return'] / 100).cumprod() - 1) * 100
    df['cum_acwi'] = ((1 + df['acwi_return'] / 100).cumprod() - 1) * 100
    df['cum_alpha'] = df['cum_actual'] - df['cum_acwi']
    return df


def compute_benchmark_daily_return(profile_name, px, idx):
    """Profile별 정적 벤치마크 일별 수익률(fraction).

    Base 시그널 weights = 정적 벤치마크 (주식 ACWI / 채권 BNDW / 현금 한국 CD 91일물 금리).
    적극형 Base: 주식 75 + 채권 10 + 현금 15
    중립형 Base: 주식 45 + 채권 40 + 현금 15
    ※ 현금은 USDKRW(FX) → 한국 CD(91일) 금리 proxy(연 CD91_ANNUAL)로 변경 (2026-06-25).
      CD 91일 ETF(357870.KS)는 2023-06 이후만 존재해 전체 기간 미달 → 상수 proxy 사용.
    """
    KRBOND_DAILY = (1.035) ** (1 / 252) - 1
    CD91_DAILY = (1 + CD91_ANNUAL) ** (1 / 252) - 1

    if 'ACWI' not in px.columns:
        return pd.Series(0.0, index=idx)
    acwi_ret = px['ACWI'].reindex(idx).pct_change().fillna(0)
    if 'BNDW' in px.columns:
        bndw_ret = px['BNDW'].reindex(idx).pct_change().fillna(0)
    else:
        bndw_ret = pd.Series(KRBOND_DAILY, index=idx)
    # 현금 = 한국 CD 91일물 금리 일별 누적 (상수 accrual)
    cd_ret = pd.Series(CD91_DAILY, index=idx)
    if len(cd_ret) > 0:
        cd_ret.iloc[0] = 0.0  # 첫날은 0 (기간 진입 앵커, ACWI/BNDW와 동일하게 pct_change 첫값 0)

    if profile_name in ('적극형', '적극'):
        bench = 0.75 * acwi_ret + 0.10 * bndw_ret + 0.15 * cd_ret
    else:  # 중립
        bench = 0.45 * acwi_ret + 0.40 * bndw_ret + 0.15 * cd_ret
    return bench


def compute_signal_win_lose(rebals, px, end_date_str, profile_name,
                              alpha_threshold_bps=30):
    """각 리밸 시점에서 시그널 선택의 win/mild/lose 평가.

    Method 1: Alpha = portfolio return − profile benchmark return (bps).
              ±alpha_threshold_bps 기준 3-tier 분류.
    Method 2: 시그널 의도(Bull/Base/Bear) vs ACWI period regime 일치 여부.
    """
    if len(rebals) < 2:
        return None
    end_ts = pd.Timestamp(end_date_str)
    px_up = px[px.index <= end_ts]
    if len(px_up) < 2:
        return None
    rows = []
    for i in range(len(rebals)):
        rb_date = pd.Timestamp(rebals[i][0])
        if rb_date > end_ts:
            break
        if i + 1 < len(rebals):
            next_date = min(pd.Timestamp(rebals[i+1][0]), end_ts)
        else:
            next_date = end_ts
        idx = px_up.index[(px_up.index >= rb_date) & (px_up.index <= next_date)]
        if len(idx) < 2:
            continue
        p0, p1 = idx[0], idx[-1]

        # 실제 portfolio 수익률 (period total %)
        weights = rebals[i][2]
        actual_ret = 0.0
        for tk, w in weights.items():
            sym = TICKER_MAP.get(tk)
            if sym is None or sym not in px.columns:
                continue
            try:
                r = (px[sym].loc[p1] / px[sym].loc[p0] - 1) * 100
                actual_ret += w / 100 * r  # bps 단위? no, % * (w/100) = %
            except Exception:
                pass

        # 벤치마크 수익률 (period total %)
        bench_daily = compute_benchmark_daily_return(profile_name, px, idx)
        bench_ret = ((1 + bench_daily).prod() - 1) * 100

        # ACWI period return for intent check
        try:
            acwi_ret = (px['ACWI'].loc[p1] / px['ACWI'].loc[p0] - 1) * 100
        except Exception:
            acwi_ret = 0

        alpha_bps = (actual_ret - bench_ret) * 100  # %p → bps

        signal = rebals[i][1]

        # Method 1: alpha 3-tier
        if alpha_bps > alpha_threshold_bps:
            m1_label = '🏆 좋은 선택'
            m1_score = 1
        elif alpha_bps < -alpha_threshold_bps:
            m1_label = '❌ 나쁜 선택'
            m1_score = -1
        else:
            m1_label = '⚪ 마일드'
            m1_score = 0

        # Method 2: signal intent vs ACWI regime
        if signal == 'Bull':
            if acwi_ret > 1.5:
                m2_label = '🏆 의도 적중'; m2_score = 1
            elif acwi_ret < -1.5:
                m2_label = '❌ 의도 빗나감'; m2_score = -1
            else:
                m2_label = '⚪ 중립'; m2_score = 0
        elif signal == 'Base':
            if abs(acwi_ret) <= 2:
                m2_label = '🏆 의도 적중'; m2_score = 1
            elif abs(acwi_ret) > 4:
                m2_label = '❌ 의도 빗나감'; m2_score = -1
            else:
                m2_label = '⚪ 중립'; m2_score = 0
        else:  # Bear
            if acwi_ret < -2:
                m2_label = '🏆 의도 적중'; m2_score = 1
            elif acwi_ret > 1:
                m2_label = '❌ 의도 빗나감'; m2_score = -1
            else:
                m2_label = '⚪ 중립'; m2_score = 0

        # 4-quadrant 분류
        if m1_score > 0 and m2_score > 0:
            quadrant = '🏆 확신의 승리'
        elif m1_score > 0 and m2_score <= 0:
            quadrant = '🍀 운 좋은 승리'
        elif m1_score <= 0 and m2_score > 0:
            quadrant = '😐 억울한 손실'
        else:
            quadrant = '❌ 확실한 실패'

        rows.append({
            'date': rb_date,
            'signal': signal,
            'actual_return': actual_ret,
            'benchmark_return': bench_ret,
            'alpha_bps': alpha_bps,
            'acwi_return': acwi_ret,
            'method1': m1_label,
            'method1_score': m1_score,
            'method2': m2_label,
            'method2_score': m2_score,
            'quadrant': quadrant,
        })
    return pd.DataFrame(rows)


def compute_broad_etf_contributions(rebals, px, end_date_str):
    """광역주식 각 ETF의 슬리브 기준 기여(bps) — sleeve-normalized 비중 × return."""
    if len(rebals) < 2:
        return None
    end_ts = pd.Timestamp(end_date_str)
    px_up = px[px.index <= end_ts]
    rows = []
    for i in range(len(rebals)):
        rb_date = pd.Timestamp(rebals[i][0])
        if rb_date > end_ts:
            break
        if i + 1 < len(rebals):
            next_date = min(pd.Timestamp(rebals[i+1][0]), end_ts)
        else:
            next_date = end_ts
        idx = px_up.index[(px_up.index >= rb_date) & (px_up.index <= next_date)]
        if len(idx) < 2:
            continue
        p0, p1 = idx[0], idx[-1]
        weights = rebals[i][2]
        sleeve_total = sum(w for tk, w in weights.items() if tk in BROAD_EQ_TICKERS)
        if sleeve_total <= 0:
            continue
        for tk in BROAD_EQ_TICKERS:
            sym = TICKER_MAP.get(tk)
            if sym is None or sym not in px.columns:
                continue
            w = weights.get(tk, 0)
            slv_w = w / sleeve_total * 100 if w > 0 else 0
            try:
                r = (px[sym].loc[p1] / px[sym].loc[p0] - 1) * 100
            except Exception:
                r = 0
            contrib_bps = slv_w * r  # 슬리브 기준 bps
            rows.append({
                'date': rb_date,
                'signal': rebals[i][1],
                'ticker': tk,
                'sym': sym,
                'sleeve_weight_pct': slv_w,
                'return_pct': r,
                'contrib_bps': contrib_bps,
            })
    return pd.DataFrame(rows)


def compute_downside_hedge_history(rebals, px, end_date_str):
    """주식 슬리브 음(-) 수익 발생 시 채권+현금 슬리브의 완충 효과 시계열.

    각 리밸 i → i+1 보유 기간 동안:
      eq_contrib  = ∑_{t ∈ 주식 슬리브} w_{i,t} × r_{i→i+1, t}  (bps)
      def_contrib = ∑_{t ∈ 채권+현금}     w_{i,t} × r_{i→i+1, t}  (bps)
      hedge_ratio = def_contrib / |eq_contrib|  (주식 음수 period 한정)
    """
    if len(rebals) < 2:
        return pd.DataFrame()
    end_ts = pd.Timestamp(end_date_str)
    px_up = px[px.index <= end_ts]
    if len(px_up) < 2:
        return pd.DataFrame()
    rows = []
    for i in range(1, len(rebals)):
        prev_rebal = rebals[i-1]
        cur_rebal = rebals[i]
        d_cur = pd.Timestamp(cur_rebal[0])
        if d_cur > end_ts:
            break
        if i + 1 < len(rebals):
            d_next = min(pd.Timestamp(rebals[i+1][0]), end_ts)
        else:
            d_next = end_ts
        idx = px_up.index[(px_up.index >= d_cur) & (px_up.index <= d_next)]
        if len(idx) < 2:
            continue
        p0 = idx[0]; p1 = idx[-1]
        eq_contrib = 0.0; def_contrib = 0.0
        eq_weight = 0.0;  def_weight = 0.0
        for tk, w in cur_rebal[2].items():
            sym = TICKER_MAP.get(tk)
            if sym is None or sym not in px.columns:
                continue
            try:
                r = (px[sym].loc[p1] / px[sym].loc[p0] - 1) * 100
            except Exception:
                r = 0
            cat = TICKER_CATEGORY.get(tk)
            if cat in EQUITY_SLEEVE_CATS:
                eq_contrib += w * r
                eq_weight += w
            elif cat in DEFENSIVE_SLEEVE_CATS:
                def_contrib += w * r
                def_weight += w
        rows.append({
            'date': d_cur,
            'signal': cur_rebal[1],
            'equity_weight': eq_weight,
            'defensive_weight': def_weight,
            'equity_contrib_bps': eq_contrib,
            'defensive_contrib_bps': def_contrib,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values('date').reset_index(drop=True)
        df['is_downside'] = df['equity_contrib_bps'] < 0
        df['hedge_ratio'] = np.where(
            df['equity_contrib_bps'] < 0,
            df['defensive_contrib_bps'] / df['equity_contrib_bps'].abs() * 100,
            np.nan,
        )
        df['net_contrib_bps'] = df['equity_contrib_bps'] + df['defensive_contrib_bps']
    return df


def compute_sleeve_rebal_effect_history(rebals, px, end_date_str, sleeve='equity'):
    """슬리브(주식/채권) 내부 비중 조절이 만든 알파(bps) 시계열.

    슬리브 i 시점 효과 = ∑_{t ∈ 슬리브} (w_{i,t} - w_{i-1,t}) × r_{i→i+1, t}.
    같은 슬리브 내 상대 비중 조절(예: SPY → QQQ, HYG → IEF)의 정합성을 측정.
    """
    cats = EQUITY_SLEEVE_CATS if sleeve == 'equity' else BOND_SLEEVE_CATS
    if len(rebals) < 2:
        return pd.DataFrame()
    end_ts = pd.Timestamp(end_date_str)
    px_up = px[px.index <= end_ts]
    if len(px_up) < 2:
        return pd.DataFrame()
    rows = []
    for i in range(1, len(rebals)):
        prev_rebal = rebals[i-1]
        cur_rebal = rebals[i]
        d_cur = pd.Timestamp(cur_rebal[0])
        if d_cur > end_ts:
            break
        if i + 1 < len(rebals):
            d_next = min(pd.Timestamp(rebals[i+1][0]), end_ts)
        else:
            d_next = end_ts
        idx = px_up.index[(px_up.index >= d_cur) & (px_up.index <= d_next)]
        if len(idx) < 2:
            continue
        p0 = idx[0]; p1 = idx[-1]
        all_tk = set(cur_rebal[2].keys()) | set(prev_rebal[2].keys())
        impact = 0.0
        for tk in all_tk:
            if TICKER_CATEGORY.get(tk) not in cats:
                continue
            sym = TICKER_MAP.get(tk)
            if sym is None or sym not in px.columns:
                continue
            w_cur = cur_rebal[2].get(tk, 0)
            w_prev = prev_rebal[2].get(tk, 0)
            try:
                r = (px[sym].loc[p1] / px[sym].loc[p0] - 1) * 100
            except Exception:
                r = 0
            impact += (w_cur - w_prev) * r
        rows.append({
            'date': d_cur,
            'signal': cur_rebal[1],
            'impact_bps': impact,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values('date').reset_index(drop=True)
        df['cumulative_bps'] = df['impact_bps'].cumsum()
        df['hit'] = (df['impact_bps'] > 0).astype(int)
        df['cum_n'] = range(1, len(df) + 1)
        df['cum_hits'] = df['hit'].cumsum()
        df['cum_hit_rate'] = df['cum_hits'] / df['cum_n'] * 100
    return df


def compute_bull_hit_rate(rebals, px):
    bull = [(i, r) for i, r in enumerate(rebals) if r[1] == 'Bull']
    spy_p = px['SPY'].dropna(); qqq_p = px['QQQ'].dropna()
    hits = 0; n = 0; excesses = []; details = []
    for _, r in bull:
        d = pd.Timestamp(r[0])
        s = spy_p[spy_p.index >= d]; q = qqq_p[qqq_p.index >= d]
        if len(s) < 22 or len(q) < 22: continue
        s_ret = s.iloc[21] / s.iloc[0] - 1
        q_ret = q.iloc[21] / q.iloc[0] - 1
        excess = q_ret - s_ret
        excesses.append(excess)
        is_hit = excess > 0
        if is_hit: hits += 1
        n += 1
        details.append({
            'date': r[0], 'SPY 1m': f'{s_ret*100:+.2f}%',
            'QQQ 1m': f'{q_ret*100:+.2f}%',
            '초과': f'{excess*100:+.2f}%p', '적중': 'O' if is_hit else 'X'
        })
    return hits, n, np.mean(excesses)*100 if excesses else 0, details


# ============== Sidebar ==============
# Initialize session state (must be before widget instantiation)
if 'report_date_key' not in st.session_state:
    st.session_state['report_date_key'] = datetime(2026, 5, 22).date()
if 'last_update' not in st.session_state:
    st.session_state['last_update'] = None

# ----- Update section (top of sidebar) -----
st.sidebar.markdown('### 🔄 Update')
update_col1, update_col2 = st.sidebar.columns([1, 1.4])
with update_col1:
    if st.button('📥 Update', use_container_width=True, type='primary',
                 help='최신 데이터로 업데이트 (가격 캐시 비우기 + Report Date를 오늘로 설정)'):
        st.cache_data.clear()
        st.session_state['report_date_key'] = datetime.now().date()
        st.session_state['last_update'] = datetime.now()
        st.toast('✓ 최신 데이터로 업데이트되었습니다', icon='🔄')
        st.rerun()
with update_col2:
    if st.session_state['last_update']:
        st.caption(f'최근 업데이트\n{st.session_state["last_update"].strftime("%m-%d %H:%M")}')
    else:
        st.caption('업데이트 이력 없음')

st.sidebar.markdown('---')
st.sidebar.title('⚙️ Dashboard 설정')

report_date = st.sidebar.date_input(
    '📅 Report Date (발행일)',
    key='report_date_key',
    min_value=datetime(2022, 3, 1),
    max_value=datetime(2030, 12, 31),
)
end_date_str = (datetime.combine(report_date, datetime.min.time())).strftime('%Y-%m-%d')

profiles = st.sidebar.multiselect(
    '📊 표시할 Risk Profile',
    options=['적극형', '중립형'],
    default=['적극형', '중립형'],
)

swap_scenario = st.sidebar.select_slider(
    '🔄 SPY → QQQ 스왑 비율 시나리오 (What-if)',
    options=[0, 25, 50, 75, 100],
    value=50,
    format_func=lambda x: f'{x}%',
)
st.sidebar.caption(f'※ 현행 룰: 50% / 시나리오: {swap_scenario}%')

show_benchmark = st.sidebar.checkbox('벤치마크 표시 (SPY, ACWI)', value=True)

st.sidebar.markdown('---')
st.sidebar.markdown('### 📋 시스템 정보')
st.sidebar.info(
    f'**발행 주기:** 매주 금요일\n\n'
    f'**Fund Inception:** {FUND_START.strftime("%Y-%m-%d")}\n\n'
    f'**Phase:** 1 MVP\n\n'
    f'**데이터 소스:** Yahoo Finance'
)

# ============== Main ==============
st.markdown(
    '<div class="ft-masthead">'
    '<div class="ft-brand">AIMVP <span class="clr">Performance</span></div>'
    '<div class="ft-tag">Asset Investment Multi-Vehicle Portfolio — Executive Dashboard</div>'
    '</div>',
    unsafe_allow_html=True)
col_a, col_b, col_c = st.columns(3)
with col_a:
    st.caption(f'**Report Date:** {report_date.strftime("%Y-%m-%d (%a)")}')
with col_b:
    st.caption(f'**Fund Inception:** {FUND_START.strftime("%Y-%m-%d")}')
with col_c:
    st.caption(f'**Phase 1 MVP** | 매주 금요일 발행')

st.markdown('---')

# Load data
with st.spinner('데이터 로딩 중...'):
    agg_rebals_all = parse_holdings('적극')
    neu_rebals_all = parse_holdings('중립')

    # ─── Filter rebals by Report Date ───
    # Only include rebalances on/before Report Date so that "latest portfolio"
    # = most recent rebal as of the selected Report Date, "previous" = the rebal before
    def filter_rebals_by_date(rebals, date_str):
        cutoff = pd.Timestamp(date_str)
        return [r for r in rebals if pd.Timestamp(r[0]) <= cutoff]

    agg_rebals = filter_rebals_by_date(agg_rebals_all, end_date_str)
    neu_rebals = filter_rebals_by_date(neu_rebals_all, end_date_str)
    if not agg_rebals or not neu_rebals:
        st.error(f'⚠ Report Date ({end_date_str})에 해당하는 리밸런싱 데이터가 없습니다. '
                 f'더 최근 날짜를 선택해 주세요.')
        st.stop()

    px = fetch_prices(end_date_str)

# ─── Guard: yfinance가 빈/불완전 응답을 줄 때(레이트리밋 등) 명확히 안내 ───
# (방치 시 compute_metrics 등 하류에서 불명확한 IndexError로 대시보드 전체가 중단됨)
if (px is None or px.empty
        or 'SPY' not in px.columns or px['SPY'].dropna().empty
        or 'ACWI' not in px.columns or px['ACWI'].dropna().empty):
    st.error('⚠ 가격 데이터를 불러오지 못했습니다 (Yahoo Finance 응답이 비어 있음 — '
             '일시적 레이트리밋 가능성). 잠시 후 사이드바의 🔄 **Update** 버튼으로 '
             '캐시를 비우고 다시 시도해 주세요.')
    st.stop()

# Simulate
agg_actual = simulate_portfolio(agg_rebals, px, swap_ratio=0.0, end_date=end_date_str)
agg_scenario = simulate_portfolio(agg_rebals, px, swap_ratio=swap_scenario/100, end_date=end_date_str)
neu_actual = simulate_portfolio(neu_rebals, px, swap_ratio=0.0, end_date=end_date_str)
neu_scenario = simulate_portfolio(neu_rebals, px, swap_ratio=swap_scenario/100, end_date=end_date_str)

spy_daily = px['SPY'].pct_change().dropna()
acwi_daily = px['ACWI'].pct_change().dropna()
spy_window = spy_daily[spy_daily.index >= pd.Timestamp(agg_rebals[0][0])]
acwi_window = acwi_daily[acwi_daily.index >= pd.Timestamp(agg_rebals[0][0])]

# Metrics
m_agg = compute_metrics(agg_actual)
m_agg_sc = compute_metrics(agg_scenario)
m_neu = compute_metrics(neu_actual)
m_neu_sc = compute_metrics(neu_scenario)
m_spy = compute_metrics(spy_window)
m_acwi = compute_metrics(acwi_window)

# Signal
last = agg_rebals[-1]
current_signal = last[1]
sig_color = {'Bull': '🟢', 'Base': '🟡', 'Bear': '🔴'}[current_signal]
spy_w = last[2].get('spy us equity', 0)
qqq_w = last[2].get('qqq us equity', 0)
rule_applied = current_signal == 'Bull'

# Hit rate
hits, n_bull, avg_excess, hit_details = compute_bull_hit_rate(agg_rebals, px)

# ============== 📰 Commentary (top of dashboard) ==============
weekly_data  = _get_period_data(px, end_date_str, period='weekly')
monthly_data = _get_period_data(px, end_date_str, period='monthly')

is_month_end = is_last_business_day_of_month(report_date)

# Top country / sector for monthly commentary context
_top_country = '미국'
_top_sector  = '정보기술'
try:
    _ctop, _ = compute_country_breakdown(agg_rebals[-1][2], top_n=1)
    if _ctop: _top_country = _ctop[0][0]
    _stop, _ = compute_sector_breakdown(agg_rebals[-1][2], top_n=1)
    if _stop: _top_sector = _stop[0][0]
except Exception:
    pass

# 현재 equity 비중 (적극/중립)
_agg_eq_pct = sum(v for k, v in agg_rebals[-1][2].items() if k in SECTOR_LOOKTHROUGH)
_neu_eq_pct = sum(v for k, v in neu_rebals[-1][2].items() if k in SECTOR_LOOKTHROUGH)

# Currently held country ETFs (>0 in current 적극 portfolio)
_held_countries = set()
_country_tk_to_name = {
    'ewc us equity': '캐나다', 'ewj us equity': '일본', 'mchi us equity': '중국',
    'ewu us equity': '영국',  'ewl us equity': '스위스', 'ewy us equity': '한국',
    'ewg us equity': '독일', 'ewa us equity': '호주', 'ewz us equity': '브라질',
    'inda us equity': '인도',
}
for tk, w in agg_rebals[-1][2].items():
    if w > 0 and tk in _country_tk_to_name:
        _held_countries.add(_country_tk_to_name[tk])

# Currently held bond ETFs (>0 across both 적극 and 중립 portfolios — union)
_bond_tickers = {
    'hyg us equity', 'phyl us equity', 'lqd us equity', 'embd us equity',
    'ief us equity', 'bkln us equity', 'gto us equity',
    'a357870', 'a451540', 'a114820',
    '357870 ks equity', '114820 ks equity',
}
_held_bonds = set()
for prof_rebals in (agg_rebals, neu_rebals):
    for tk, w in prof_rebals[-1][2].items():
        if w > 0 and tk in _bond_tickers:
            _held_bonds.add(tk)

with st.container(border=True):
    st.markdown('## 📰 Commentary')

    # ─── 생성 방식 토글 ───
    gen_mode = st.radio(
        '생성 방식',
        ['📝 Rule-based (single function)', '🧩 Swarm Agent (modular)'],
        horizontal=True, key='commentary_gen_mode',
        help='Rule-based: 단일 함수에서 모든 narrative 생성.  '
             'Swarm Agent: 5 specialist(macro/country/bond/fx/signal) 병렬 → '
             'Synthesizer → Editor. 모듈화·확장성 ↑. API key 불필요, 즉시 동작.',
    )
    swarm_mode = gen_mode.startswith('🧩')

    # ─── Swarm용 context builder ───
    def _build_swarm_ctx(kind, data):
        """deterministic 함수가 사용하는 facts를 swarm용 CommentaryContext로 변환."""
        from commentary_swarm import CommentaryContext
        # 보유 국가 ETF의 return만 수집
        country_returns = {}
        country_map = [('캐나다', 'ewc'), ('일본', 'ewj'), ('중국', 'mchi'),
                       ('영국', 'ewu'), ('스위스', 'ewl'), ('한국', 'ewy')]
        for ctry, key in country_map:
            if ctry in _held_countries:
                country_returns[ctry] = round(data.get(key, 0), 2)
        # 보유 채권 ETF의 return만 수집
        bond_returns = {}
        bond_map = [('HYG', 'hyg'), ('LQD', 'lqd'), ('IEF', 'ief'),
                    ('EMBD', 'embd'), ('PHYL', 'phyl')]
        for label, key in bond_map:
            tk_form = key + ' us equity'
            if tk_form in _held_bonds:
                bond_returns[label] = round(data.get(key, 0), 2)
        if 'a451540' in _held_bonds or '451540 ks equity' in _held_bonds:
            bond_returns['TIGER 종합채권'] = 0.0  # KR proxy
        if 'a357870' in _held_bonds or '357870 ks equity' in _held_bonds:
            bond_returns['TIGER CD금리'] = 0.0
        # 시그널 최근 12개 (Excel 시그널 추이 시트에서 추출 시도)
        signal_history = []
        try:
            wb_h = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
            if '시그널 추이' in wb_h.sheetnames:
                ws_h = wb_h['시그널 추이']
                rows_h = []
                for ri in range(2, ws_h.max_row + 1):
                    d_v = ws_h.cell(ri, 1).value
                    s_v = ws_h.cell(ri, 2).value
                    if d_v and s_v:
                        d_str = d_v.strftime('%Y-%m-%d') if hasattr(d_v, 'strftime') else str(d_v)
                        if d_str <= end_date_str:
                            rows_h.append(s_v)
                signal_history = rows_h[-12:]
        except Exception:
            pass
        # period 라벨
        if kind == 'weekly':
            period_label = f"{data['end'].year}년 {data['end'].month}월 {data['end'].day}일 주간"
            period_range = f"{data['start'].month}/{data['start'].day} ~ {data['end'].month}/{data['end'].day}"
        else:
            period_label = f"{data['end'].year}년 {data['end'].month}월"
            period_range = f"{data['start'].month}/{data['start'].day} ~ {data['end'].month}/{data['end'].day}"
        return CommentaryContext(
            kind=kind,
            period_label=period_label,
            period_range=period_range,
            signal=current_signal,
            signal_date=str(last[0]),
            agg_eq_pct=round(_agg_eq_pct, 2),
            neu_eq_pct=round(_neu_eq_pct, 2),
            market_returns={
                'SPY':  round(data['spy'], 2),
                'QQQ':  round(data['qqq'], 2),
                'ACWI': round(data['acwi'], 2),
                'EFA':  round(data['efa'], 2),
                'VWO':  round(data['vwo'], 2),
            },
            country_returns=country_returns,
            bond_returns=bond_returns,
            fx_return=round(data['usdkrw'], 2),
            signal_history=signal_history,
        )

    def _run_swarm(kind, data):
        """Stochastic swarm — 매 호출마다 다른 narrative (same facts)."""
        from commentary_swarm import generate_commentary_sync
        ctx = _build_swarm_ctx(kind, data)
        return generate_commentary_sync(ctx)

    def _swarm_state_key(kind, data):
        """Session state key — (kind, period)로 식별."""
        return f"commentary_swarm_{kind}_{data['start']}_{data['end']}"

    def _swarm_get_or_create(kind, data, regen: bool = False):
        """Session state 캐시 — data 변경 또는 regen=True 시만 재생성."""
        key = _swarm_state_key(kind, data)
        if regen or key not in st.session_state:
            st.session_state[key] = _run_swarm(kind, data)
        return st.session_state[key]

    tab_w, tab_m = st.tabs(['📅 Weekly', '📆 Monthly'])
    with tab_w:
        if swarm_mode:
            try:
                # 재생성 버튼 (우상단)
                rcol1, rcol2 = st.columns([5, 1])
                with rcol2:
                    regen_w = st.button('🔄 재생성', key='regen_weekly_btn',
                                         help='Swarm 재실행 → 새 narrative 변형 생성')
                result = _swarm_get_or_create('weekly', weekly_data, regen=regen_w)
                weekly_text = result['final']
                st.markdown(weekly_text)
                with st.expander('🔬 Swarm 내부 (specialist drafts + synthesizer 출력)'):
                    for role, draft in result['drafts'].items():
                        st.markdown(f'**[{role}]** {draft.get("narrative", "")}')
                    st.markdown('---')
                    st.markdown('**[synthesized (pre-editor)]**')
                    st.code(result['synthesized'][:1500], language='markdown')
                st.caption(f'※ 🧩 Swarm Agent 생성 (분량 {len(weekly_text)}자) — '
                            f'5 specialist 병렬 → Synthesizer → Editor.  '
                            f'🔄 재생성 클릭 시 같은 데이터에서 다른 표현 narrative.')
            except Exception as e:
                st.warning(f'⚠ Swarm 실패: {e}\n\n→ Rule-based fallback으로 전환합니다.')
                weekly_text = make_weekly_commentary(
                    weekly_data, current_signal, last[0],
                    m_agg['ITD'], m_neu['ITD'],
                    m_agg['WTD'], m_neu['WTD'],
                )
                st.markdown(weekly_text)
                st.caption(f'※ Rule-based fallback (분량: 약 {len(weekly_text)}자)')
        else:
            weekly_text = make_weekly_commentary(
                weekly_data, current_signal, last[0],
                m_agg['ITD'], m_neu['ITD'],
                m_agg['WTD'], m_neu['WTD'],
            )
            st.markdown(weekly_text)
            st.caption(f'※ 조회일 기준 직전 5영업일 시장 데이터 기반 자동 생성 (분량: 약 {len(weekly_text)}자)')

    with tab_m:
        if is_month_end:
            if swarm_mode:
                try:
                    rcol1m, rcol2m = st.columns([5, 1])
                    with rcol2m:
                        regen_m = st.button('🔄 재생성', key='regen_monthly_btn',
                                             help='Swarm 재실행 → 새 narrative 변형 생성')
                    result = _swarm_get_or_create('monthly', monthly_data, regen=regen_m)
                    monthly_text = result['final']
                    st.markdown(monthly_text)
                    with st.expander('🔬 Swarm 내부 (specialist drafts + synthesizer 출력)'):
                        for role, draft in result['drafts'].items():
                            st.markdown(f'**[{role}]** {draft.get("narrative", "")}')
                        st.markdown('---')
                        st.markdown('**[synthesized (pre-editor)]**')
                        st.code(result['synthesized'][:2000], language='markdown')
                    st.caption(f'※ 🧩 Swarm Agent 생성 (분량 {len(monthly_text)}자) — '
                                f'5 specialist 병렬 → Synthesizer → Editor.  '
                                f'🔄 재생성 클릭 시 같은 데이터에서 다른 표현 narrative.')
                except Exception as e:
                    st.warning(f'⚠ Swarm 실패: {e}\n\n→ Rule-based fallback으로 전환합니다.')
                    monthly_text = make_monthly_commentary(
                        monthly_data, current_signal, last[0],
                        _agg_eq_pct, _neu_eq_pct,
                        _held_countries, _held_bonds,
                    )
                    st.markdown(monthly_text)
                    st.caption(f'※ Rule-based fallback (분량: 약 {len(monthly_text)}자)')
            else:
                monthly_text = make_monthly_commentary(
                    monthly_data, current_signal, last[0],
                    _agg_eq_pct, _neu_eq_pct,
                    _held_countries, _held_bonds,
                )
                st.markdown(monthly_text)
                st.caption(f'※ 월말 영업일 MTD 시장 데이터 기반 자동 생성 (분량: 약 {len(monthly_text)}자)')
        else:
            st.info(
                f'📌 Monthly Commentary는 조회일이 **매월 마지막 영업일**일 때만 자동 생성됩니다.\n\n'
                f'현재 Report Date({report_date.strftime("%Y-%m-%d (%a)")})는 월말 마지막 영업일이 아닙니다.\n\n'
                f'※ Report Date를 해당 월의 마지막 영업일(예: 2026-06-30)로 변경 시 표시됩니다.'
            )

st.markdown('')  # small spacing

# ============== 1줄 Summary (always visible at top) ==============
summary_text = (
    f'⭐ 적극형 ITD **{m_agg["ITD"]:+.2f}%** (Sharpe {m_agg["Sharpe"]:.2f}) | '
    f'중립형 ITD **{m_neu["ITD"]:+.2f}%** (Sharpe {m_neu["Sharpe"]:.2f}) | '
    f'현재 시그널 **{sig_color} {current_signal}** | '
    f'Bull 적중률 **{hits/n_bull*100:.1f}%** ({hits}/{n_bull})'
)
st.success(summary_text)

# ============== Navigation Blocks ==============
if 'page' not in st.session_state:
    st.session_state.page = 'Portfolio'

# Custom CSS: turn nav buttons into tall card blocks
st.markdown('''
<style>
div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:has(button[data-testid="stBaseButton-primary"]),
div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:has(button[data-testid="stBaseButton-secondary"]) {
}
button[data-testid="stBaseButton-primary"],
button[data-testid="stBaseButton-secondary"] {
    min-height: 110px !important;
    border-radius: 14px !important;
    border-width: 2px !important;
    font-size: 16px !important;
    font-weight: 600 !important;
    white-space: pre-line !important;
    padding: 18px 12px !important;
    line-height: 1.45 !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease !important;
}
button[data-testid="stBaseButton-primary"] {
    background-color: #1F3A68 !important;
    border-color: #1F3A68 !important;
    color: white !important;
    box-shadow: 0 4px 10px rgba(31,58,104,0.25) !important;
}
button[data-testid="stBaseButton-secondary"] {
    background-color: white !important;
    border-color: #C7D2E0 !important;
    color: #1F3A68 !important;
}
button[data-testid="stBaseButton-secondary"]:hover {
    border-color: #1F3A68 !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 4px 10px rgba(31,58,104,0.15) !important;
}
</style>
''', unsafe_allow_html=True)

nav_cols = st.columns(5)
nav_items = [
    ('Portfolio',         '📊 Portfolio\n자산배분 · 종목별 성과 · 보유 비중'),
    ('Performance',       '📈 Performance\nWTD/MTD/YTD/ITD · 누적성과 · 위험 지표'),
    ('Signal Validation', '🎯 Signal Validation\n시그널 적중률 · 통계 검정 · 이벤트'),
    ('Universe',          '🌐 Universe\n투자 유니버스 종목 리스트'),
    ('월별 포트폴리오',     '📑 월별 포트폴리오\n월별 비중·성과 + 시그널/ACWI 필터'),
]
for col, (page_key, label) in zip(nav_cols, nav_items):
    with col:
        is_active = st.session_state.page == page_key
        if st.button(label,
                     key=f'nav_{page_key}', use_container_width=True,
                     type='primary' if is_active else 'secondary'):
            st.session_state.page = page_key
            st.rerun()

st.markdown('---')

# ============================================================
# Page: Performance
# ============================================================
if st.session_state.page == 'Performance':
    # KPI Cards
    def kpi_row(label, m, ref=None):
        cols = st.columns(5)
        cols[0].markdown(f'### {label}')
        for col, key, name in zip(cols[1:], ['WTD','MTD','YTD','ITD'], ['WTD','MTD','YTD','ITD']):
            val = m[key]
            delta_str = None
            if ref is not None:
                delta = val - ref[key]
                delta_str = f'{delta:+.2f}%p vs 실제'
            col.metric(name, f'{val:+.2f}%', delta=delta_str)

    if '적극형' in profiles:
        st.markdown('#### 📈 적극형')
        kpi_row('실제 운용', m_agg)
        if swap_scenario != 0:
            kpi_row(f'시나리오 ({swap_scenario}%)', m_agg_sc, ref=m_agg)

    if '중립형' in profiles:
        st.markdown('#### 📉 중립형')
        kpi_row('실제 운용', m_neu)
        if swap_scenario != 0:
            kpi_row(f'시나리오 ({swap_scenario}%)', m_neu_sc, ref=m_neu)

    # ─── 누적성과 추이 (MTD/QTD/YTD/ITD + 1M/3M/6M/1Y/3Y/a/5Y/a) ───
    st.markdown('---')
    st.subheader('📊 누적성과 추이 (캘린더 · 롤링 윈도우)')

    def _period_start(end_ts, period):
        """period start 계산.
          캘린더: MTD/QTD/YTD/ITD
          롤링: 1M, 3M, 6M, 1Y, 3Y, 5Y (end_ts에서 정확한 시간 차감)
        """
        if period == 'MTD':
            return pd.Timestamp(end_ts.year, end_ts.month, 1)
        if period == 'QTD':
            q_start_month = ((end_ts.month - 1) // 3) * 3 + 1
            return pd.Timestamp(end_ts.year, q_start_month, 1)
        if period == 'YTD':
            return pd.Timestamp(end_ts.year, 1, 1)
        if period == 'ITD':
            return pd.Timestamp(agg_rebals[0][0])
        # 롤링 윈도우 (월/년 단위)
        if period == '1M': return end_ts - pd.DateOffset(months=1)
        if period == '3M': return end_ts - pd.DateOffset(months=3)
        if period == '6M': return end_ts - pd.DateOffset(months=6)
        if period == '1Y': return end_ts - pd.DateOffset(years=1)
        if period == '3Y': return end_ts - pd.DateOffset(years=3)
        if period == '5Y': return end_ts - pd.DateOffset(years=5)
        return pd.Timestamp(agg_rebals[0][0])

    ANNUALIZED_PERIODS = {'3Y', '5Y'}  # CAGR display

    def _build_period_chart(period, end_ts):
        start_ts = _period_start(end_ts, period)
        fig = go.Figure()
        end_values = {}  # label -> cumulative return %

        def _slice_cum(series, label):
            s = series[(series.index >= start_ts) & (series.index <= end_ts)]
            if len(s) == 0:
                return None
            cum = (1 + s).cumprod()
            end_values[label] = (cum.iloc[-1] - 1) * 100
            return cum

        if '적극형' in profiles:
            c = _slice_cum(agg_actual, '적극형')
            if c is not None:
                fig.add_trace(go.Scatter(
                    x=c.index, y=(c - 1) * 100,
                    name='적극형 (실제)', line=dict(color='#DC2626', width=3),
                    hovertemplate='%{x|%Y-%m-%d}<br>%{y:.2f}%<extra></extra>',
                ))
            if swap_scenario != 0:
                c_sc = _slice_cum(agg_scenario, f'적극형 {swap_scenario}%')
                if c_sc is not None:
                    fig.add_trace(go.Scatter(
                        x=c_sc.index, y=(c_sc - 1) * 100,
                        name=f'적극형 ({swap_scenario}% 시나리오)',
                        line=dict(color='#DC2626', width=2, dash='dash'),
                        hovertemplate='%{x|%Y-%m-%d}<br>%{y:.2f}%<extra></extra>',
                    ))
        if '중립형' in profiles:
            c_n = _slice_cum(neu_actual, '중립형')
            if c_n is not None:
                fig.add_trace(go.Scatter(
                    x=c_n.index, y=(c_n - 1) * 100,
                    name='중립형 (실제)', line=dict(color='#2563EB', width=3),
                    hovertemplate='%{x|%Y-%m-%d}<br>%{y:.2f}%<extra></extra>',
                ))
            if swap_scenario != 0:
                c_n_sc = _slice_cum(neu_scenario, f'중립형 {swap_scenario}%')
                if c_n_sc is not None:
                    fig.add_trace(go.Scatter(
                        x=c_n_sc.index, y=(c_n_sc - 1) * 100,
                        name=f'중립형 ({swap_scenario}% 시나리오)',
                        line=dict(color='#2563EB', width=2, dash='dash'),
                        hovertemplate='%{x|%Y-%m-%d}<br>%{y:.2f}%<extra></extra>',
                    ))
        if show_benchmark:
            c_spy = _slice_cum(spy_daily, 'SPY')
            c_acwi = _slice_cum(acwi_daily, 'ACWI')
            if c_spy is not None:
                fig.add_trace(go.Scatter(
                    x=c_spy.index, y=(c_spy - 1) * 100,
                    name='SPY', line=dict(color='#9CA3AF', width=2),
                    hovertemplate='%{x|%Y-%m-%d}<br>%{y:.2f}%<extra></extra>',
                ))
            if c_acwi is not None:
                fig.add_trace(go.Scatter(
                    x=c_acwi.index, y=(c_acwi - 1) * 100,
                    name='ACWI', line=dict(color='#6B7280', width=2, dash='dot'),
                    hovertemplate='%{x|%Y-%m-%d}<br>%{y:.2f}%<extra></extra>',
                ))

        # Bull 시그널 음영 (period 윈도우 내로 클립)
        for i in range(len(agg_rebals)):
            if agg_rebals[i][1] == 'Bull':
                s = pd.Timestamp(agg_rebals[i][0])
                e = pd.Timestamp(agg_rebals[i+1][0]) if i < len(agg_rebals)-1 else end_ts
                s_clip = max(s, start_ts); e_clip = min(e, end_ts)
                if s_clip < e_clip:
                    fig.add_vrect(x0=s_clip, x1=e_clip,
                                  fillcolor='#FEE2E2', opacity=0.3, line_width=0)

        fig.update_layout(
            height=480 if period == 'ITD' else 420, hovermode='x unified',
            yaxis_title='누적수익률 (%)', xaxis_title='날짜',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='left', x=0),
            margin=dict(l=40, r=20, t=20, b=40),
        )
        fig.add_hline(y=0, line_dash='dash', line_color='black', opacity=0.5)
        return fig, start_ts, end_values

    end_ts_period = pd.Timestamp(end_date_str)
    PERIOD_DEFS = [
        ('📅 MTD',  'MTD'),  ('🔄 1M',   '1M'),
        ('📊 QTD',  'QTD'),  ('🔄 3M',   '3M'),  ('🔄 6M',   '6M'),
        ('📈 YTD',  'YTD'),  ('🔄 1Y',   '1Y'),
        ('📐 3Y/a', '3Y'),   ('📐 5Y/a', '5Y'),
        ('🏁 ITD',  'ITD'),
    ]
    period_tabs = st.tabs([lbl for lbl, _ in PERIOD_DEFS])
    for tab, (label, period) in zip(period_tabs, PERIOD_DEFS):
        with tab:
            fig_p, start_p, end_vals = _build_period_chart(period, end_ts_period)
            days = (end_ts_period - start_p).days
            years = days / 365.25
            base_caption = (
                f'**기간:** {start_p.strftime("%Y-%m-%d")} → {end_ts_period.strftime("%Y-%m-%d")}'
                f'  ({days}일{" / 약 " + f"{years:.2f}년" if years >= 0.5 else ""})'
            )
            if period == 'ITD':
                base_caption += '  · 운용시작 → 조회일'
            base_caption += '  |  ※ 분홍 음영 = Bull 시그널 발효 구간 (해당 기간 내)'
            st.caption(base_caption)

            # 연환산 CAGR 표시 (3Y/5Y/ITD)
            if period in ANNUALIZED_PERIODS or period == 'ITD':
                if end_vals and years > 0:
                    cagr_parts = []
                    for lbl, val in end_vals.items():
                        try:
                            cagr = ((1 + val / 100) ** (1 / years) - 1) * 100
                            cagr_parts.append(f'{lbl} **{cagr:+.2f}%**')
                        except Exception:
                            pass
                    if cagr_parts:
                        st.caption(f'**📐 연환산(CAGR):** ' + ' / '.join(cagr_parts))

            st.plotly_chart(fig_p, use_container_width=True, key=f'cum_{period.lower()}')

            # 운용기간 부족 경고 (3Y/5Y에서 portfolio 데이터 누락)
            if period in ANNUALIZED_PERIODS:
                fund_inception = pd.Timestamp(agg_rebals[0][0])
                if start_p < fund_inception:
                    short_years = (end_ts_period - fund_inception).days / 365.25
                    st.warning(
                        f'⚠ 운용기간({short_years:.1f}년)이 {period} 윈도우보다 짧아 적극·중립 시리즈는 '
                        f'운용 이후 구간({fund_inception.strftime("%Y-%m-%d")}부터)만 표시됩니다. '
                        f'벤치마크(SPY/ACWI)만 전 기간 표시.'
                    )

    # ─── 배분별 성과 추이 (국가/섹터/팩터, 같은 기간 탭 사용) ───
    st.markdown('---')
    st.subheader('🧭 배분별 성과 추이 (국가 · 섹터 · 팩터)')
    st.caption(
        '국가별 / 섹터별 / 팩터별 **look-through 누적수익률 추이**.  '
        '각 카테고리 일별 수익률 = ∑_ETF (보유 비중 × ETF 카테고리 share × ETF 일별 return) / ∑(가중합)  '
        '— 슬리브 자체의 카테고리 노출 가중평균. **상단 누적성과 추이와 동일한 기간 윈도우** 사용.'
    )

    # ─── 🔥 유니버스 비중 분포 Heatmap ───
    st.markdown('##### 🔥 유니버스 비중 분포 Heatmap (ETF × 카테고리)')
    st.caption(
        '각 ETF가 국가/섹터/팩터에 노출된 **look-through 비중(%)** 매트릭스.  '
        '🟢 **초록 = 낮은 비중**, 🔴 **빨강 = 높은 비중**. 셀 내 숫자 = 비중 %(5% 이상만 표시).'
    )

    BROAD_ETF_ORDER = ['spy us equity', 'qqq us equity', 'iwm us equity',
                       'acwi us equity', 'efa us equity', 'vwo us equity']

    def _build_lookthrough_heatmap(lookthrough_dict, category_order=None,
                                    zmax=100, min_text=5, dim_label='카테고리'):
        # ETF 순서: 광역 ETFs 먼저, 그 다음 단일국가 ETFs 알파벳 순
        broad_in = [tk for tk in BROAD_ETF_ORDER if tk in lookthrough_dict]
        others = sorted([tk for tk in lookthrough_dict if tk not in broad_in])
        tk_list = broad_in + others
        ticker_labels = [TICKER_MAP.get(tk, tk).upper() if TICKER_MAP.get(tk, tk).upper() != tk.upper()
                         else tk.split(' ')[0].upper() for tk in tk_list]

        # 카테고리 collect
        all_cats = set()
        for tk in tk_list:
            all_cats |= set(lookthrough_dict[tk].keys())
        if category_order:
            cats = [c for c in category_order if c in all_cats]
            for c in sorted(all_cats):
                if c not in cats:
                    cats.append(c)
        else:
            totals = {c: sum(lookthrough_dict[tk].get(c, 0) for tk in tk_list) for c in all_cats}
            cats = sorted(all_cats, key=lambda c: totals[c], reverse=True)

        # z matrix
        z = [[lookthrough_dict[tk].get(c, 0) for tk in tk_list] for c in cats]
        text = [[f'{v:.0f}' if v >= min_text else '' for v in row] for row in z]

        fig = go.Figure(data=go.Heatmap(
            z=z, x=ticker_labels, y=cats,
            colorscale='RdYlGn_r',  # 초록(낮음) → 노랑 → 빨강(높음)
            zmin=0, zmax=zmax,
            text=text, texttemplate='%{text}',
            textfont={'size': 9, 'color': 'black'},
            hovertemplate='ETF: %{x}<br>' + dim_label + ': %{y}<br>비중: %{z:.1f}%<extra></extra>',
            colorbar=dict(title='비중(%)', thickness=12, len=0.9),
        ))
        height = max(280, len(cats) * 22 + 120)
        fig.update_layout(
            height=height,
            xaxis=dict(side='top', tickangle=-45, tickfont=dict(size=10),
                       title=dict(text='ETF Ticker', standoff=10)),
            yaxis=dict(tickfont=dict(size=10), autorange='reversed',
                       title=dict(text=dim_label, standoff=10)),
            margin=dict(l=120, r=20, t=90, b=30),
        )
        return fig

    def _build_universe_table(lookthrough_dict):
        """ETF Ticker / Full Name 매핑 테이블 (heatmap과 동일 순서)."""
        broad_in = [tk for tk in BROAD_ETF_ORDER if tk in lookthrough_dict]
        others = sorted([tk for tk in lookthrough_dict if tk not in broad_in])
        rows = []
        for tk in broad_in + others:
            sym_raw = TICKER_MAP.get(tk, tk)
            sym = sym_raw.upper() if sym_raw.upper() != tk.upper() else tk.split(' ')[0].upper()
            rows.append({'Ticker': sym, 'Full Name': TICKER_NAMES.get(tk, tk)})
        return pd.DataFrame(rows)

    hm_tabs = st.tabs(['🌍 국가 Heatmap', '🏭 섹터 Heatmap', '🎯 팩터 Heatmap (MSCI)'])

    with hm_tabs[0]:
        col_hm, col_tbl = st.columns([4, 1])
        with col_hm:
            fig_hm_c = _build_lookthrough_heatmap(LOOKTHROUGH, dim_label='국가', zmax=100)
            st.plotly_chart(fig_hm_c, use_container_width=True, key='heatmap_country')
        with col_tbl:
            st.markdown('**🏷️ ETF Universe**')
            df_u = _build_universe_table(LOOKTHROUGH)
            st.dataframe(df_u, use_container_width=True, hide_index=True,
                         height=int(fig_hm_c.layout.height) - 30)
        st.caption(
            ':grey[**※ 해석:** 컬럼별로 ETF 1개의 국가 분포가 위에서 아래로 표시.  '
            'SPY/QQQ/IWM은 US 100% 단일색, ACWI/EFA/VWO는 다국가 분산. '
            '단일국가 ETF(EWJ/MCHI 등)는 해당 국가에서만 100% 빨강.]'
        )

    with hm_tabs[1]:
        SECTOR_ORDER = ['정보기술', '금융', '헬스케어', '임의소비재', '통신서비스',
                        '산업재', '필수소비재', '에너지', '소재', '유틸리티', '부동산']
        col_hm, col_tbl = st.columns([4, 1])
        with col_hm:
            fig_hm_s = _build_lookthrough_heatmap(SECTOR_LOOKTHROUGH,
                                                    category_order=SECTOR_ORDER,
                                                    dim_label='섹터', zmax=50, min_text=3)
            st.plotly_chart(fig_hm_s, use_container_width=True, key='heatmap_sector')
        with col_tbl:
            st.markdown('**🏷️ ETF Universe**')
            df_u = _build_universe_table(SECTOR_LOOKTHROUGH)
            st.dataframe(df_u, use_container_width=True, hide_index=True,
                         height=int(fig_hm_s.layout.height) - 30)
        st.caption(
            ':grey[**※ 해석:** QQQ는 정보기술 50%로 가장 빨간 셀.  '
            'EFA/EWJ는 금융·산업재 비중이 상대적으로 큰 가치형 분포.  '
            'VWO/MCHI는 임의소비재·통신서비스 비중 큼 (EM 특성).]'
        )

    with hm_tabs[2]:
        col_hm, col_tbl = st.columns([4, 1])
        with col_hm:
            fig_hm_f = _build_lookthrough_heatmap(FACTOR_LOOKTHROUGH_MSCI,
                                                    category_order=FACTOR_ORDER,
                                                    dim_label='팩터', zmax=50, min_text=5)
            st.plotly_chart(fig_hm_f, use_container_width=True, key='heatmap_factor')
        with col_tbl:
            st.markdown('**🏷️ ETF Universe**')
            df_u = _build_universe_table(FACTOR_LOOKTHROUGH_MSCI)
            st.dataframe(df_u, use_container_width=True, hide_index=True,
                         height=int(fig_hm_f.layout.height) - 30)
        st.caption(
            ':grey[**※ 해석:** QQQ Growth 42% / Quality 23%로 성장·품질 편향.  '
            'EFA·EWJ·EWU Value 28~32%로 가치 편향.  '
            'EWL(스위스) Quality 26% / EWZ(브라질) Value 32% 등 지역적 팩터 색깔이 한눈에 가시화.  '
            'Fama-French 회귀 베타와는 별개의 정적 매핑 결과.]'
        )

    st.markdown('---')

    PERF_COLORS = [
        '#1F77B4', '#FF7F0E', '#2CA02C', '#D62728', '#9467BD', '#8C564B',
        '#E377C2', '#7F7F7F', '#BCBD22', '#17BECF', '#1F3A68', '#C48D43',
    ]

    # ─── 방법론 A: 순수 절대 수익률 (비중 무관) ───
    def _category_absolute_return_series(lookthrough_dict, start_ts, end_ts):
        """카테고리별 순수 절대 수익률 (%).

        포트폴리오 비중 미사용 — look-through share만 사용해 ETF들을 가중평균.
        카테고리 c 일별 수익률 = ∑_ETF (share_etf_c × r_etf_daily) / ∑_ETF (share_etf_c)
        Universe: lookthrough_dict의 모든 ETF (가격 데이터 있는 것).
        """
        if start_ts >= end_ts:
            return {}
        et_data = {}
        for tk in lookthrough_dict:
            sym = TICKER_MAP.get(tk)
            if sym is None or sym not in px.columns:
                continue
            prices_sliced = px[sym][(px[sym].index >= start_ts) & (px[sym].index <= end_ts)]
            if len(prices_sliced) < 2:
                continue
            et_data[tk] = prices_sliced.pct_change().fillna(0)
        if not et_data:
            return {}
        all_cats = set()
        for tk in et_data:
            all_cats |= set(lookthrough_dict[tk].keys())
        result = {}
        for cat in all_cats:
            weighted_ret = None
            total_share = 0.0
            for tk, rets in et_data.items():
                share = lookthrough_dict[tk].get(cat, 0) / 100
                if share <= 0:
                    continue
                contribution = share * rets
                if weighted_ret is None:
                    weighted_ret = contribution.copy()
                else:
                    weighted_ret = weighted_ret.add(contribution, fill_value=0)
                total_share += share
            if total_share <= 0 or weighted_ret is None:
                continue
            daily_avg = weighted_ret / total_share
            result[cat] = ((1 + daily_avg).cumprod() - 1) * 100
        return result

    # ─── 방법론 B: Historical 포트폴리오 기여 (bps) ───
    def _category_historical_contribution_series(lookthrough_dict, rebals_in,
                                                   start_ts, end_ts):
        """카테고리별 historical 포트폴리오 기여 시계열 (bps).

        각 리밸 구간 [rebal_i, rebal_{i+1}]에서 그 시점 실제 비중 사용:
          일별 카테고리 c 기여(bps) = ∑_ETF (w_i_etf% × share_etf_c% × r_etf_daily_frac)
        과거 시점별 비중 변화 모두 반영. 누적 bps over time.
        """
        if not rebals_in or start_ts >= end_ts:
            return {}
        px_slice = px[(px.index >= start_ts) & (px.index <= end_ts)]
        if len(px_slice) < 2:
            return {}
        all_cats = set()
        for tk in lookthrough_dict:
            all_cats |= set(lookthrough_dict[tk].keys())
        # 일별 contribution accumulator
        contrib = {cat: pd.Series(0.0, index=px_slice.index) for cat in all_cats}
        for i in range(len(rebals_in)):
            rb_date = pd.Timestamp(rebals_in[i][0])
            next_date = (pd.Timestamp(rebals_in[i+1][0]) if i + 1 < len(rebals_in)
                         else end_ts + pd.Timedelta(days=1))
            weights = rebals_in[i][2]
            period_idx = px_slice.index[(px_slice.index >= rb_date) &
                                          (px_slice.index < next_date)]
            if len(period_idx) < 1:
                continue
            for tk, w in weights.items():
                if tk not in lookthrough_dict or w <= 0:
                    continue
                sym = TICKER_MAP.get(tk)
                if sym is None or sym not in px.columns:
                    continue
                try:
                    etf_ret = px[sym].loc[period_idx].pct_change().fillna(0)
                except Exception:
                    continue
                for cat, share in lookthrough_dict[tk].items():
                    # w(%) × share(%) × r(fraction) = bps (portfolio level)
                    contrib_bps = w * share * etf_ret
                    contrib[cat].loc[period_idx] = contrib[cat].loc[period_idx].add(
                        contrib_bps, fill_value=0)
        # 누적 bps
        result = {cat: ser.cumsum() for cat, ser in contrib.items() if ser.abs().sum() > 0}
        return result

    # ─── FF 팩터 방법론 A: long-short proxy 누적수익률 (%) ───
    def _ff_factor_returns_cum(start_ts, end_ts):
        factor_daily = compute_ff_factor_returns_daily(px)
        if factor_daily is None:
            return {}
        sliced = factor_daily[(factor_daily.index >= start_ts) & (factor_daily.index <= end_ts)]
        if len(sliced) == 0:
            return {}
        result = {}
        for f in FF_FACTOR_ORDER:
            if f in sliced.columns:
                result[f] = ((1 + sliced[f]).cumprod() - 1) * 100
        return result

    # ─── FF 팩터 방법론 B: Historical 포트폴리오 기여 (bps) ───
    def _ff_factor_historical_contribution(rebals_in, betas_map, start_ts, end_ts):
        """각 리밸 시점 weights × ETF betas → portfolio net beta → × 일별 factor return
        → 누적 bps.
        """
        if not rebals_in or not betas_map or start_ts >= end_ts:
            return {}
        factor_daily = compute_ff_factor_returns_daily(px)
        if factor_daily is None:
            return {}
        px_slice = factor_daily[(factor_daily.index >= start_ts) &
                                 (factor_daily.index <= end_ts)]
        if len(px_slice) < 2:
            return {}
        contrib = {f: pd.Series(0.0, index=px_slice.index) for f in FF_FACTOR_ORDER}
        for i in range(len(rebals_in)):
            rb_date = pd.Timestamp(rebals_in[i][0])
            next_date = (pd.Timestamp(rebals_in[i+1][0]) if i + 1 < len(rebals_in)
                         else end_ts + pd.Timedelta(days=1))
            weights = rebals_in[i][2]
            period_idx = px_slice.index[(px_slice.index >= rb_date) &
                                          (px_slice.index < next_date)]
            if len(period_idx) < 1:
                continue
            # Portfolio net beta = ∑ over ETFs (w_portfolio_fraction × etf_beta)
            net_betas = {f: 0.0 for f in FF_FACTOR_ORDER}
            for tk, w in weights.items():
                if tk not in FACTOR_LOOKTHROUGH_MSCI or w <= 0:
                    continue
                sym = TICKER_MAP.get(tk)
                if sym not in betas_map:
                    continue
                w_frac = w / 100  # portfolio weight as fraction
                for f in FF_FACTOR_ORDER:
                    net_betas[f] += w_frac * betas_map[sym].get(f, 0)
            # 일별 contribution (bps) = net_beta × factor_return_daily × 10000
            for f in FF_FACTOR_ORDER:
                if f not in px_slice.columns:
                    continue
                daily_contrib_bps = net_betas[f] * px_slice[f].loc[period_idx] * 10000
                contrib[f].loc[period_idx] = contrib[f].loc[period_idx].add(
                    daily_contrib_bps, fill_value=0)
        return {f: ser.cumsum() for f, ser in contrib.items() if ser.abs().sum() > 0}

    def _build_perf_chart(series_dict, top_n=None, height=440,
                            yaxis_title='누적수익률 (%)', value_fmt='.2f'):
        if not series_dict:
            return None
        items = list(series_dict.items())
        if top_n is not None:
            items.sort(key=lambda kv: kv[1].iloc[-1] if len(kv[1]) > 0 else 0,
                       reverse=True)
            items = items[:top_n]
        fig = go.Figure()
        for i, (cat, ser) in enumerate(items):
            fig.add_trace(go.Scatter(
                x=ser.index, y=ser.values,
                name=cat, mode='lines',
                line=dict(color=PERF_COLORS[i % len(PERF_COLORS)], width=2),
                hovertemplate=('%{x|%Y-%m-%d}<br>' + cat +
                               ': %{y:' + value_fmt + '}<extra></extra>'),
            ))
        fig.add_hline(y=0, line_dash='dash', line_color='black', opacity=0.5)
        fig.update_layout(
            height=height, hovermode='x unified',
            yaxis_title=yaxis_title, xaxis_title='날짜',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='left', x=0,
                        font=dict(size=9)),
            margin=dict(l=40, r=20, t=20, b=40),
        )
        return fig

    # 방어형 ETF 베타 매트릭스 사전 계산 (FF Historical 기여 탭용, 캐시됨)
    _etf_syms_all = sorted({TICKER_MAP[tk] for tk in FACTOR_LOOKTHROUGH_MSCI
                             if tk in TICKER_MAP and TICKER_MAP[tk] in px.columns})

    # ─── 🔘 방법론 토글 ───
    st.markdown('##### 🔘 방법론 선택')
    methodology = st.radio(
        '',
        ['💯 절대 수익률 (비중 무관)',
         '💰 포트폴리오 기여 (historical 비중)'],
        horizontal=True,
        key='alloc_perf_methodology',
        label_visibility='collapsed',
    )
    is_absolute = methodology.startswith('💯')
    if is_absolute:
        st.caption(
            ':grey[**💯 절대 수익률 모드:** 포트폴리오 비중 미사용. '
            '카테고리 c 일별 수익률 = ∑_ETF (share_etf_c × r_etf_daily) / ∑(share). '
            'Universe = look-through dict 모든 ETF. Y축 = **누적수익률(%)**.]'
        )
    else:
        st.caption(
            ':grey[**💰 historical 기여 모드:** 각 리밸 시점 실제 portfolio 비중 사용. '
            '일별 카테고리 기여 = ∑_ETF (w_시점_etf% × share_etf_c% × r_etf_daily_fraction). '
            '과거 시점별 비중 변화 모두 반영. Y축 = **누적 기여(bps, portfolio level)**.]'
        )

    dim_tabs = st.tabs(['🌍 국가별', '🏭 섹터별', '🎯 팩터별'])

    def _get_series_for_dim(lookthrough_dict, start_ts, end_ts):
        if is_absolute:
            return _category_absolute_return_series(lookthrough_dict, start_ts, end_ts)
        return _category_historical_contribution_series(
            lookthrough_dict, agg_rebals, start_ts, end_ts)

    _yaxis_lbl = '누적수익률 (%)' if is_absolute else '누적 기여 (bps)'
    _val_fmt   = '.2f' if is_absolute else '.1f'

    # ─── 🌍 국가별 ───
    with dim_tabs[0]:
        st.caption('보유 ETF의 국가 구성 look-through 기준. Top 10 국가 표시.')
        c_period_tabs = st.tabs([lbl for lbl, _ in PERIOD_DEFS])
        for tab, (label, period) in zip(c_period_tabs, PERIOD_DEFS):
            with tab:
                start_p = _period_start(end_ts_period, period)
                days = (end_ts_period - start_p).days
                st.caption(
                    f'**기간:** {start_p.strftime("%Y-%m-%d")} → {end_ts_period.strftime("%Y-%m-%d")} ({days}일)'
                )
                series = _get_series_for_dim(LOOKTHROUGH, start_p, end_ts_period)
                fig = _build_perf_chart(series, top_n=10,
                                          yaxis_title=_yaxis_lbl, value_fmt=_val_fmt)
                if fig is None:
                    st.info('데이터 부족')
                else:
                    st.plotly_chart(fig, use_container_width=True,
                                    key=f'perf_country_{period.lower()}')

    # ─── 🏭 섹터별 ───
    with dim_tabs[1]:
        st.caption('보유 ETF의 GICS 11 섹터 look-through 기준.')
        s_period_tabs = st.tabs([lbl for lbl, _ in PERIOD_DEFS])
        for tab, (label, period) in zip(s_period_tabs, PERIOD_DEFS):
            with tab:
                start_p = _period_start(end_ts_period, period)
                days = (end_ts_period - start_p).days
                st.caption(
                    f'**기간:** {start_p.strftime("%Y-%m-%d")} → {end_ts_period.strftime("%Y-%m-%d")} ({days}일)'
                )
                series = _get_series_for_dim(SECTOR_LOOKTHROUGH, start_p, end_ts_period)
                fig = _build_perf_chart(series, top_n=11,
                                          yaxis_title=_yaxis_lbl, value_fmt=_val_fmt)
                if fig is None:
                    st.info('데이터 부족')
                else:
                    st.plotly_chart(fig, use_container_width=True,
                                    key=f'perf_sector_{period.lower()}')

    # ─── 🎯 팩터별 (MSCI / FF) ───
    with dim_tabs[2]:
        st.caption('6개 표준 팩터 — MSCI 큐레이션 look-through vs FF long-short proxy.')
        f_method_tabs = st.tabs(['🎨 MSCI Style', '📐 Fama-French'])

        # MSCI
        with f_method_tabs[0]:
            st.caption(':orange[정적 매핑 추정치 — Style Box 분석 기반.]')
            fm_period_tabs = st.tabs([lbl for lbl, _ in PERIOD_DEFS])
            for tab, (label, period) in zip(fm_period_tabs, PERIOD_DEFS):
                with tab:
                    start_p = _period_start(end_ts_period, period)
                    days = (end_ts_period - start_p).days
                    st.caption(
                        f'**기간:** {start_p.strftime("%Y-%m-%d")} → {end_ts_period.strftime("%Y-%m-%d")} ({days}일)'
                    )
                    series = _get_series_for_dim(FACTOR_LOOKTHROUGH_MSCI,
                                                   start_p, end_ts_period)
                    fig = _build_perf_chart(series, top_n=6,
                                              yaxis_title=_yaxis_lbl, value_fmt=_val_fmt)
                    if fig is None:
                        st.info('데이터 부족')
                    else:
                        st.plotly_chart(fig, use_container_width=True,
                                        key=f'perf_factor_msci_{period.lower()}')

        # FF
        with f_method_tabs[1]:
            if is_absolute:
                st.caption(
                    ':blue[long-short proxy 절대 누적수익률 (비중 무관) — '
                    'Value=VLUE-SPY · Size=IWM-SPY · Quality=QUAL-SPY · '
                    'Momentum=MTUM-SPY · Low Vol=USMV-SPY · MKT=SPY.]'
                )
            else:
                st.caption(
                    ':blue[Historical portfolio 팩터 기여 (bps) — '
                    '각 리밸 시점 portfolio Net Beta × 일별 factor return × 10000 누적.]'
                )

            # FF historical 기여 모드용 betas
            ff_betas = None
            if not is_absolute:
                with st.spinner('FF 베타 계산 중...'):
                    ff_betas = compute_ff_betas_for_etfs(
                        str(end_date_str), tuple(_etf_syms_all), px)

            ff_period_tabs = st.tabs([lbl for lbl, _ in PERIOD_DEFS])
            for tab, (label, period) in zip(ff_period_tabs, PERIOD_DEFS):
                with tab:
                    start_p = _period_start(end_ts_period, period)
                    days = (end_ts_period - start_p).days
                    st.caption(
                        f'**기간:** {start_p.strftime("%Y-%m-%d")} → {end_ts_period.strftime("%Y-%m-%d")} ({days}일)'
                    )
                    if is_absolute:
                        series = _ff_factor_returns_cum(start_p, end_ts_period)
                    else:
                        series = _ff_factor_historical_contribution(
                            agg_rebals, ff_betas or {}, start_p, end_ts_period)
                    fig = _build_perf_chart(series, top_n=None,
                                              yaxis_title=_yaxis_lbl, value_fmt=_val_fmt)
                    if fig is None:
                        st.info('데이터 부족 (Update 클릭으로 재시도)')
                    else:
                        st.plotly_chart(fig, use_container_width=True,
                                        key=f'perf_factor_ff_{period.lower()}')

    # Risk Metrics Table
    st.markdown('---')
    st.subheader('📐 위험 지표 비교')

    risk_df = pd.DataFrame({
        '지표': ['연환산 CAGR', '연환산 변동성', 'Sharpe Ratio', 'Sortino Ratio', 'MaxDD'],
        '적극형': [f'{m_agg["CAGR"]:.2f}%', f'{m_agg["Vol"]:.2f}%',
                  f'{m_agg["Sharpe"]:.2f}', f'{m_agg["Sortino"]:.2f}', f'{m_agg["MaxDD"]:.2f}%'],
        f'적극형 ({swap_scenario}%)': [f'{m_agg_sc["CAGR"]:.2f}%', f'{m_agg_sc["Vol"]:.2f}%',
                                       f'{m_agg_sc["Sharpe"]:.2f}', f'{m_agg_sc["Sortino"]:.2f}',
                                       f'{m_agg_sc["MaxDD"]:.2f}%'],
        '중립형': [f'{m_neu["CAGR"]:.2f}%', f'{m_neu["Vol"]:.2f}%',
                  f'{m_neu["Sharpe"]:.2f}', f'{m_neu["Sortino"]:.2f}', f'{m_neu["MaxDD"]:.2f}%'],
        'SPY': [f'{m_spy["CAGR"]:.2f}%', f'{m_spy["Vol"]:.2f}%',
                f'{m_spy["Sharpe"]:.2f}', f'{m_spy["Sortino"]:.2f}', f'{m_spy["MaxDD"]:.2f}%'],
        'ACWI': [f'{m_acwi["CAGR"]:.2f}%', f'{m_acwi["Vol"]:.2f}%',
                 f'{m_acwi["Sharpe"]:.2f}', f'{m_acwi["Sortino"]:.2f}', f'{m_acwi["MaxDD"]:.2f}%'],
    })
    st.dataframe(risk_df, use_container_width=True, hide_index=True)

    # Alerts (performance-related)
    st.markdown('---')
    st.subheader('🚨 주간 Alerts & Watchlist')
    latest = agg_rebals[-1][2]
    alerts = []
    if current_signal == 'Bull' and rule_applied:
        alerts.append(('🟢 Info', 'Bull 시그널 활성, 50% SPY→QQQ 스왑 룰 적용 중'))
    elif current_signal == 'Base':
        alerts.append(('🟢 Info', 'Base 시그널 — 룰 미적용, 원본 비중 유지'))
    elif current_signal == 'Bear':
        alerts.append(('🔴 Critical', 'Bear 시그널 발효 — 위험자산 축소 확인 필요'))
    if hits/n_bull < 0.55:
        alerts.append(('🟡 Watch', f'Bull 적중률 {hits/n_bull*100:.1f}% — 6개월 이동평균 점검 필요'))
    else:
        alerts.append(('🟢 Info', f'Bull 적중률 {hits/n_bull*100:.1f}% — 정상 범위 (>55%)'))
    if m_agg['WTD'] < -3:
        alerts.append(('🟠 Warn', f'적극형 WTD {m_agg["WTD"]:+.2f}% — 3% 하락 임계 초과'))
    if m_agg['MTD'] < -5:
        alerts.append(('🟠 Warn', f'적극형 MTD {m_agg["MTD"]:+.2f}% — 5% 하락 한계 초과'))
    spy_qqq_sum = latest.get('spy us equity', 0) + latest.get('qqq us equity', 0)
    if spy_qqq_sum > 50:
        alerts.append(('🟡 Watch', f'SPY+QQQ 합산 {spy_qqq_sum:.1f}% — 미국 집중도 점검 권고'))
    alpha_vs_spy = m_agg['ITD'] - m_spy['ITD']
    if alpha_vs_spy < -5:
        alerts.append(('🟠 Warn', f'적극형 vs SPY ITD alpha {alpha_vs_spy:+.2f}%p — 벤치 underperform'))
    else:
        alerts.append(('🟢 Info', f'적극형 vs SPY ITD alpha {alpha_vs_spy:+.2f}%p'))
    alpha_vs_acwi = m_agg['ITD'] - m_acwi['ITD']
    if alpha_vs_acwi >= 0:
        alerts.append(('🟢 Info', f'적극형 vs ACWI ITD alpha {alpha_vs_acwi:+.2f}%p — 글로벌 BM 대비 양호'))
    else:
        alerts.append(('🟡 Watch', f'적극형 vs ACWI ITD alpha {alpha_vs_acwi:+.2f}%p'))
    for level, msg in alerts:
        if 'Info' in level: st.success(f'{level}  {msg}')
        elif 'Watch' in level or 'Warn' in level: st.warning(f'{level}  {msg}')
        elif 'Critical' in level: st.error(f'{level}  {msg}')

    # Download metrics
    st.markdown('---')
    metrics_df = pd.DataFrame({
        '지표': ['WTD','MTD','YTD','ITD','CAGR','Vol','Sharpe','Sortino','MaxDD'],
        '적극형': [m_agg[k] for k in ['WTD','MTD','YTD','ITD','CAGR','Vol','Sharpe','Sortino','MaxDD']],
        '중립형': [m_neu[k] for k in ['WTD','MTD','YTD','ITD','CAGR','Vol','Sharpe','Sortino','MaxDD']],
        'SPY': [m_spy[k] for k in ['WTD','MTD','YTD','ITD','CAGR','Vol','Sharpe','Sortino','MaxDD']],
        'ACWI': [m_acwi[k] for k in ['WTD','MTD','YTD','ITD','CAGR','Vol','Sharpe','Sortino','MaxDD']],
    })
    csv = metrics_df.to_csv(index=False).encode('utf-8-sig')
    st.download_button(
        '📥 Performance 지표 CSV 다운로드',
        data=csv,
        file_name=f'aimvp_metrics_{report_date.strftime("%Y%m%d")}.csv',
        mime='text/csv', use_container_width=True,
    )

# ============================================================
# Page: Portfolio
# ============================================================
elif st.session_state.page == 'Portfolio':
    # Compute 중립형 weights (signal is shared with 적극)
    neu_last = neu_rebals[-1]
    neu_spy_w = neu_last[2].get('spy us equity', 0)
    neu_qqq_w = neu_last[2].get('qqq us equity', 0)

    # ─── Signal status (shared) ───
    sig_cols = st.columns([1, 1.4, 1.4])
    with sig_cols[0]:
        st.markdown('### 🎯 현재 시그널')
        st.markdown(f'## {sig_color} {current_signal}')
        st.markdown(f'**발효:** {last[0]}')
        if rule_applied:
            st.error('🔴 50% 스왑 룰 적용 중')
        else:
            st.info('⚪ 룰 미적용 (Base/Bear)')
    with sig_cols[1]:
        st.markdown('### 📈 적극형')
        st.metric('SPY', f'{spy_w:.2f}%')
        st.metric('QQQ', f'{qqq_w:.2f}%')
        st.caption(f'주식합 {sum(v for k,v in last[2].items() if k in {"spy us equity","qqq us equity","acwi us equity","efa us equity","vwo us equity","iwm us equity","ewa us equity","ewc us equity","enzl us equity","ewg us equity","norw us equity","ewo us equity","ewh us equity","ews us equity","ewi us equity","ewn us equity","ewj us equity","mchi us equity","ewk us equity","ewq us equity","ewu us equity","ewl us equity","ewz us equity","inda us equity","ewy us equity","eden us equity","eis us equity","ewd us equity","ewp us equity","pdbc us equity"}):.2f}%')
    with sig_cols[2]:
        st.markdown('### 📉 중립형')
        st.metric('SPY', f'{neu_spy_w:.2f}%')
        st.metric('QQQ', f'{neu_qqq_w:.2f}%')
        st.caption(f'주식합 {sum(v for k,v in neu_last[2].items() if k in {"spy us equity","qqq us equity","acwi us equity","efa us equity","vwo us equity","iwm us equity","ewa us equity","ewc us equity","enzl us equity","ewg us equity","norw us equity","ewo us equity","ewh us equity","ews us equity","ewi us equity","ewn us equity","ewj us equity","mchi us equity","ewk us equity","ewq us equity","ewu us equity","ewl us equity","ewz us equity","inda us equity","ewy us equity","eden us equity","eis us equity","ewd us equity","ewp us equity","pdbc us equity"}):.2f}%')

    # ─── 📋 시그널별 자산 배분 정책 테이블 ───
    st.markdown('##### 📋 시그널별 자산 배분 정책 (%)')
    alloc_table_cols = st.columns(2)
    alloc_policy = {
        '적극형': {
            'Bull': (90, 5,  5),
            'Base': (75, 10, 15),
            'Bear': (60, 25, 15),
        },
        '중립형': {
            'Bull': (60, 35, 5),
            'Base': (45, 40, 15),
            'Bear': (30, 55, 15),
        },
    }
    for col, (profile_lbl, alloc) in zip(alloc_table_cols, alloc_policy.items()):
        with col:
            st.markdown(f'**{profile_lbl}**')
            df_alloc = pd.DataFrame({
                '시그널': ['🟢 Bull', '🟡 Base', '🔴 Bear'],
                '주식(%)':  [alloc['Bull'][0], alloc['Base'][0], alloc['Bear'][0]],
                '채권(%)':  [alloc['Bull'][1], alloc['Base'][1], alloc['Bear'][1]],
                '현금(%)':  [alloc['Bull'][2], alloc['Base'][2], alloc['Bear'][2]],
            })
            styled_alloc = df_alloc.style.background_gradient(
                subset=['주식(%)'], cmap='Reds', vmin=0, vmax=100,
            ).background_gradient(
                subset=['채권(%)'], cmap='Blues', vmin=0, vmax=100,
            ).background_gradient(
                subset=['현금(%)'], cmap='Greys', vmin=0, vmax=30,
            ).format({'주식(%)': '{:.0f}', '채권(%)': '{:.0f}', '현금(%)': '{:.0f}'})
            st.dataframe(styled_alloc, use_container_width=True, hide_index=True)
    st.caption(
        ':grey[**자산 매핑:** 주식 = MSCI ACWI / 채권 = BNDW (Bloomberg Global Aggregate proxy) / 현금 = **한국 CD(91일물) 금리** (proxy 연 3.5%).  '
        '**Base** = profile-specific 정적 벤치마크와 동일 (알파 산출 기준).  '
        '**시그널별 합:** 모두 100% (Bull 위험자산 ↑ / Bear 방어자산 ↑).]'
    )

    # ─── 🎯 프로파일별 벤치마크(BM) 정의 ───
    st.markdown('##### 🎯 프로파일별 벤치마크 (BM) 정의')
    bm_def = {
        '적극형': {'ACWI (주식)': 75, 'BNDW (채권)': 10, 'CD(91일) (현금)': 15},
        '중립형': {'ACWI (주식)': 45, 'BNDW (채권)': 40, 'CD(91일) (현금)': 15},
    }
    bm_cols = st.columns(2)
    for col, (profile_lbl, comps) in zip(bm_cols, bm_def.items()):
        with col:
            st.markdown(f'**{profile_lbl} BM**')
            df_bm = pd.DataFrame({
                '구성 자산': list(comps.keys()),
                '비중(%)': list(comps.values()),
            })
            styled_bm = df_bm.style.background_gradient(
                subset=['비중(%)'], cmap='Greens', vmin=0, vmax=100,
            ).format({'비중(%)': '{:.0f}'})
            st.dataframe(styled_bm, use_container_width=True, hide_index=True)
            formula = ' + '.join(f'{w}% {k.split(" ")[0]}' for k, w in comps.items())
            st.caption(f':grey[**= {formula}**]')
    st.caption(
        ':grey[**벤치마크(BM) 정의:** profile별 **정적(static) 벤치마크**로, 각 시그널의 **Base 배분과 동일**합니다  '
        '(주식 = MSCI ACWI / 채권 = BNDW = Bloomberg Global Aggregate proxy / 현금 = **한국 CD(91일물) 금리** proxy 연 3.5%).  '
        '대시보드 전반의 **상승·하락 캡처(전체 범위)**, **배분별 성과** 평가의 기준선으로 사용.  '
        '비중 합 = 100%. 주식·채권은 Yahoo Finance auto-adjusted, 현금은 CD 91일 상수 accrual '
        '(357870.KS는 2023-06 이후만 존재해 전체 기간 커버 위해 상수 proxy).]'
    )

    # ─── Allocation pies side-by-side ───
    st.markdown('---')

    def build_allocation_pie(weights, title):
        group_w = {'미국 주식': 0, '해외 주식': 0, '채권': 0, '현금/MMF': 0}
        for tk, w in weights.items():
            if tk in ('spy us equity','qqq us equity','acwi us equity','iwm us equity'):
                group_w['미국 주식'] += w
            elif tk in ('hyg us equity','lqd us equity','bkln us equity','ief us equity',
                        'embd us equity','gto us equity','phyl us equity',
                        'a357870','a114820','a451540',
                        '357870 ks equity','114820 ks equity'):
                group_w['채권'] += w
            elif tk in ('usdkrw curncy','cash_krw'):
                group_w['현금/MMF'] += w
            else:
                group_w['해외 주식'] += w
        total = sum(group_w.values())
        if total > 0:
            group_w = {k: v/total*100 for k, v in group_w.items()}
        fig = go.Figure(go.Pie(
            labels=list(group_w.keys()),
            values=list(group_w.values()),
            marker_colors=['#DC2626', '#F59E0B', '#2563EB', '#9CA3AF'],
            textinfo='label+percent',
            textfont_size=11,
            hole=0.4,
        ))
        fig.update_layout(
            height=320, margin=dict(l=10, r=10, t=40, b=10),
            showlegend=False,
            title=dict(text=title, font=dict(size=14), x=0.5),
        )
        return fig

    # ─── 🎯 시그널별 기간별 누적성과 추이 (Bull / Base / Bear) ───
    st.markdown('### 🎯 시그널별 누적성과 추이 (Bull · Base · Bear)')
    st.caption(
        '주식 = **MSCI ACWI** / 채권 = **BNDW (Bloomberg Global Aggregate proxy)** / 현금 = **한국 CD(91일물) 금리**(proxy 연 3.5%) 기준 합성 시나리오.  '
        '일별 수익률 = w_주식×r_ACWI + w_채권×r_BNDW + w_현금×r_CD91를 compound. (Base 라인 = profile 정적 BM)'
    )

    # 시그널별 자산 배분 (profile 별 차등)
    SIGNAL_ALLOC = {
        '적극형': {
            'Bull': {'주식': 90, '채권': 5,  '현금': 5},
            'Base': {'주식': 75, '채권': 10, '현금': 15},
            'Bear': {'주식': 60, '채권': 25, '현금': 15},
        },
        '중립형': {
            'Bull': {'주식': 60, '채권': 35, '현금': 5},
            'Base': {'주식': 45, '채권': 40, '현금': 15},
            'Bear': {'주식': 30, '채권': 55, '현금': 15},
        },
    }
    SIGNAL_LINE_COLORS = {'Bull': '#DC2626', 'Base': '#F59E0B', 'Bear': '#10B981'}

    # 기간 정의 (Performance 페이지와 동일)
    SIG_PERIOD_DEFS = [
        ('📅 MTD',  'MTD'),  ('🔄 1M',   '1M'),
        ('📊 QTD',  'QTD'),  ('🔄 3M',   '3M'),  ('🔄 6M',   '6M'),
        ('📈 YTD',  'YTD'),  ('🔄 1Y',   '1Y'),
        ('📐 3Y/a', '3Y'),   ('📐 5Y/a', '5Y'),
        ('🏁 ITD',  'ITD'),
    ]
    ANN_SIG = {'3Y', '5Y'}

    def _sig_period_start(end_ts_p, period):
        if period == 'MTD': return pd.Timestamp(end_ts_p.year, end_ts_p.month, 1)
        if period == 'QTD':
            q = ((end_ts_p.month - 1) // 3) * 3 + 1
            return pd.Timestamp(end_ts_p.year, q, 1)
        if period == 'YTD': return pd.Timestamp(end_ts_p.year, 1, 1)
        if period == 'ITD': return pd.Timestamp(agg_rebals[0][0])
        if period == '1M':  return end_ts_p - pd.DateOffset(months=1)
        if period == '3M':  return end_ts_p - pd.DateOffset(months=3)
        if period == '6M':  return end_ts_p - pd.DateOffset(months=6)
        if period == '1Y':  return end_ts_p - pd.DateOffset(years=1)
        if period == '3Y':  return end_ts_p - pd.DateOffset(years=3)
        if period == '5Y':  return end_ts_p - pd.DateOffset(years=5)
        return pd.Timestamp(agg_rebals[0][0])

    end_ts_sig = pd.Timestamp(end_date_str)
    px_sig_all = px[px.index <= end_ts_sig]
    today_sig = px_sig_all.index[-1] if len(px_sig_all) >= 2 else None

    # 실제 portfolio 일별 수익률 매핑
    _ACTUAL_RETS = {'적극형': agg_actual, '중립형': neu_actual}
    _ACTUAL_COLOR = {'적극형': '#1F3A68', '중립형': '#C48D43'}

    def _build_signal_period_chart(profile_name, period, key_suffix):
        if today_sig is None or 'ACWI' not in px.columns:
            return None, None, None
        start_ts = _sig_period_start(today_sig, period)
        idx = px_sig_all.index[(px_sig_all.index >= start_ts) &
                                (px_sig_all.index <= today_sig)]
        if len(idx) < 2:
            return None, start_ts, idx
        acwi_ret = px['ACWI'].reindex(idx).pct_change().fillna(0)
        bndw_ret = (px['BNDW'].reindex(idx).pct_change().fillna(0)
                     if 'BNDW' in px.columns else pd.Series(0.0, index=idx))
        # 현금 = 한국 CD(91일물) 금리 일별 누적 (BM과 동일, 첫날 0 앵커)
        cd_ret = pd.Series((1 + CD91_ANNUAL) ** (1 / 252) - 1, index=idx)
        if len(cd_ret) > 0:
            cd_ret.iloc[0] = 0.0

        fig = go.Figure()
        end_vals = {}
        # 3 signal scenarios
        for sig, alloc in SIGNAL_ALLOC[profile_name].items():
            w_eq = alloc['주식'] / 100
            w_bd = alloc['채권'] / 100
            w_ca = alloc['현금'] / 100
            daily = w_eq * acwi_ret + w_bd * bndw_ret + w_ca * cd_ret
            cum = ((1 + daily).cumprod() - 1) * 100
            end_vals[sig] = cum.iloc[-1] if len(cum) > 0 else 0
            fig.add_trace(go.Scatter(
                x=cum.index, y=cum.values,
                name=f'{sig} (주식 {alloc["주식"]}% / 채권 {alloc["채권"]}% / 현금 {alloc["현금"]}%)',
                mode='lines+markers',
                line=dict(color=SIGNAL_LINE_COLORS[sig], width=2.0, dash='dot'),
                marker=dict(size=3),
                hovertemplate='%{x|%Y-%m-%d}<br>' + sig + ': %{y:+.2f}%<extra></extra>',
            ))

        # 실제 portfolio cumulative — historical 월별 리밸 시점 비중 적용
        # (simulate_portfolio가 각 리밸 i의 weights를 다음 리밸까지 사용해 일별 return 계산)
        actual_series = _ACTUAL_RETS.get(profile_name)
        if actual_series is not None:
            actual_sliced = actual_series[(actual_series.index >= start_ts) &
                                            (actual_series.index <= today_sig)].copy()
            if len(actual_sliced) >= 2:
                # 기간 시작일 0% 앵커 (시나리오 라인의 fillna(0)와 동일 규약 — 기간 진입
                # 직전 overnight 수익률이 누적에 섞이지 않도록 첫 수익률을 0으로)
                actual_sliced.iloc[0] = 0.0
                actual_cum = ((1 + actual_sliced).cumprod() - 1) * 100
                end_vals['실제'] = actual_cum.iloc[-1] if len(actual_cum) > 0 else 0
                fig.add_trace(go.Scatter(
                    x=actual_cum.index, y=actual_cum.values,
                    name=f'⭐ {profile_name} 실제 (historical 월별 비중)',
                    mode='lines',
                    line=dict(color=_ACTUAL_COLOR[profile_name], width=3.5),
                    hovertemplate='%{x|%Y-%m-%d}<br>' + profile_name +
                                  ' 실제 (historical): %{y:+.2f}%<extra></extra>',
                ))

        fig.add_hline(y=0, line_dash='dot', line_color='gray', line_width=1)
        fig.update_layout(
            height=380, hovermode='x unified',
            xaxis_title='날짜', yaxis_title='누적수익률 (%)',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='left', x=0,
                        font=dict(size=9)),
            margin=dict(t=30, b=40, l=40, r=20),
            title=dict(text=f'<b>🥧 {profile_name}</b>', x=0.02, font=dict(size=14)),
        )
        return fig, start_ts, idx

    # 기간 탭 전체 너비로 표시
    sig_period_tabs = st.tabs([lbl for lbl, _ in SIG_PERIOD_DEFS])
    for tab, (lbl, period) in zip(sig_period_tabs, SIG_PERIOD_DEFS):
        with tab:
            # 적극형
            fig_a, start_a, idx_a = _build_signal_period_chart('적극형', period, 'agg')
            if fig_a is not None:
                st.plotly_chart(fig_a, use_container_width=True,
                                 key=f'sig_perf_agg_{period.lower()}')
            else:
                st.info(f'적극형 {period} 데이터 부족')
            # 중립형
            fig_n, start_n, idx_n = _build_signal_period_chart('중립형', period, 'neu')
            if fig_n is not None:
                st.plotly_chart(fig_n, use_container_width=True,
                                 key=f'sig_perf_neu_{period.lower()}')
            else:
                st.info(f'중립형 {period} 데이터 부족')

            # 기간 caption + 연환산 (3Y/5Y/ITD)
            if start_a is not None and idx_a is not None and len(idx_a) >= 2:
                days = (today_sig - start_a).days
                years = days / 365.25
                cap = (f'**🗓️ 기간:** {start_a.strftime("%Y-%m-%d")} → '
                        f'{today_sig.strftime("%Y-%m-%d")} ({days}일')
                if years >= 0.5:
                    cap += f' / 약 {years:.2f}년'
                cap += f', {len(idx_a)}영업일).'
                # CAGR
                if period in ANN_SIG or period == 'ITD':
                    if years > 0:
                        # Recompute end values for CAGR display (적극형 기준)
                        _, _, _ = None, None, None
                        cagr_parts = []
                        # 적극형 end values 다시 산출
                        for sig, alloc in SIGNAL_ALLOC['적극형'].items():
                            w_eq, w_bd, w_ca = alloc['주식']/100, alloc['채권']/100, alloc['현금']/100
                            acwi_ret = px['ACWI'].reindex(idx_a).pct_change().fillna(0)
                            bndw_ret = (px['BNDW'].reindex(idx_a).pct_change().fillna(0)
                                         if 'BNDW' in px.columns else pd.Series(0.0, index=idx_a))
                            cd_ret = pd.Series((1 + CD91_ANNUAL) ** (1 / 252) - 1, index=idx_a)
                            if len(cd_ret) > 0:
                                cd_ret.iloc[0] = 0.0
                            daily = w_eq * acwi_ret + w_bd * bndw_ret + w_ca * cd_ret
                            cum = ((1 + daily).cumprod() - 1) * 100
                            if len(cum) > 0 and years > 0:
                                try:
                                    cagr = ((1 + cum.iloc[-1] / 100) ** (1 / years) - 1) * 100
                                    cagr_parts.append(f'{sig} **{cagr:+.2f}%**')
                                except Exception:
                                    pass
                        if cagr_parts:
                            cap += f'  📐 **적극형 연환산(CAGR):** ' + ' / '.join(cagr_parts)
                st.caption(':grey[' + cap + ']')

    st.caption(
        ':grey[**📊 시나리오 시그널 배분 (정적):**  '
        '적극형 — Bull(90/5/5) · Base(75/10/15) · Bear(60/25/15) — 주식/채권/현금 %.  '
        '중립형 — Bull(60/35/5) · Base(45/40/15) · Bear(30/55/15).  '
        '시나리오는 주식=MSCI ACWI / 채권=BNDW / 현금=한국 CD(91일물) 금리 비중을 **고정** 유지하여 시뮬레이션.  '
        '\n\n**⭐ 실제 portfolio (historical 월별 비중):**  '
        '각 리밸 시점의 **실제 보유 비중**을 다음 리밸까지 적용해 일별 수익률을 compound — '
        '즉, 매월 변하는 실제 자산 배분 의사결정이 반영된 진짜 운용 성과 시계열.  '
        '`simulate_portfolio(rebals, px)`가 rebals[i][2]를 [rebals[i][0], rebals[i+1][0]) 구간에 적용.  '
        '\n\n**해석:** ⭐ 라인이 시그널 시나리오들 사이/위/아래 어디에 있는지 보면 '
        '실제 운용이 어떤 시그널 가정 대비 우위/열위였는지 즉시 비교 가능.  '
        '데이터 출처: Yahoo Finance auto-adjusted (Total Return).]'
    )

    # ─── 📊 상승 캡처 vs 하락 캡처 (Upside / Downside Capture) ───
    st.markdown('---')
    st.markdown('### 📊 상승 캡처 vs 하락 캡처 (Upside / Downside Capture)')
    st.caption(
        '이상적인 방어형은 **상승 캡처는 높고**(상방을 잘 따라감) **하락 캡처는 낮아야**(하방을 덜 떨어짐) 합니다.  '
        '두 막대 높이가 같으면 **비대칭 = 0** — 상·하방을 같은 비율로 누른다는 뜻.  '
        '시그널별(Bull · Base · **Bear**)과 전략 전체를 나눠 산출합니다.'
    )

    cap_t1, cap_t2 = st.columns(2)
    with cap_t1:
        cap_profile = st.radio('프로파일', ['적극형', '중립형'],
                                horizontal=True, key='cap_asym_profile')
    with cap_t2:
        cap_scope = st.radio('분석 범위 (벤치마크 기준)',
                              ['전체 포트폴리오 (실제 BM)', '주식 only (ACWI 100%)'],
                              horizontal=True, key='cap_asym_scope')
    cap_equity_only = cap_scope.startswith('주식')
    cap_rebals = agg_rebals if cap_profile == '적극형' else neu_rebals

    def _cap_monthly_rows(rebals, profile_name, equity_only, end_override=None):
        """월별 (signal, 포트 월수익%, BM 월수익%) 리스트.
        equity_only=True  → 포트=주식 슬리브(광역+국가주식, 100% 정규화), BM=ACWI 100%.
        equity_only=False → 포트=전체 실제 비중,                      BM=profile 정적 BM.
        end_override: 지정 시 해당 일자(as-of)까지로 컷오프 (특정일 캡처 비교용). None이면 Report Date."""
        rows = []
        if 'ACWI' not in px.columns:
            return rows
        end_ts = pd.Timestamp(end_override) if end_override else pd.Timestamp(end_date_str)
        px_up = px[px.index <= end_ts]
        if len(px_up) < 2:
            return rows
        for i in range(len(rebals)):
            rb_date = pd.Timestamp(rebals[i][0])
            if rb_date > end_ts:
                break
            nxt = (min(pd.Timestamp(rebals[i + 1][0]), end_ts)
                   if i + 1 < len(rebals) else end_ts)
            idx = px_up.index[(px_up.index >= rb_date) & (px_up.index <= nxt)]
            if len(idx) < 2:
                continue
            signal = rebals[i][1]
            weights = rebals[i][2]
            if equity_only:
                sel = {tk: w for tk, w in weights.items()
                       if TICKER_CATEGORY.get(tk) in EQUITY_SLEEVE_CATS}
            else:
                sel = dict(weights)
            tot = sum(sel.values())
            if tot <= 0:
                continue
            # 포트 월간 수익률 = 일별 weighted return compound
            port_daily = pd.Series(0.0, index=idx)
            used = False
            for tk, w in sel.items():
                sym = TICKER_MAP.get(tk)
                if sym is None or sym not in px.columns:
                    continue
                r = px[sym].reindex(idx).pct_change().fillna(0)
                port_daily = port_daily + (w / tot) * r
                used = True
            if not used:
                continue
            port_ret = ((1 + port_daily).prod() - 1) * 100
            # 벤치마크 월간 수익률
            if equity_only:
                bm_daily = px['ACWI'].reindex(idx).pct_change().fillna(0)
            else:
                bm_daily = compute_benchmark_daily_return(profile_name, px, idx)
            bm_ret = ((1 + bm_daily).prod() - 1) * 100
            rows.append((signal, port_ret, bm_ret))
        return rows

    def _cap_ratios(rows):
        """(signal, port%, bm%) → {group: {up, dn, asym, n_up, n_dn}}.
        합산법: 상승 캡처 = Σ포트(BM↑월) / ΣBM(BM↑월), 하락 캡처 = Σ포트(BM↓월) / ΣBM(BM↓월)."""
        def _calc(subset):
            up = [(p, b) for (p, b) in subset if b > 0]
            dn = [(p, b) for (p, b) in subset if b < 0]
            sb_up = sum(b for _, b in up)
            sb_dn = sum(b for _, b in dn)
            cap_up = (sum(p for p, _ in up) / sb_up) if sb_up > 0 else None
            cap_dn = (sum(p for p, _ in dn) / sb_dn) if sb_dn < 0 else None
            asym = (cap_up - cap_dn) if (cap_up is not None and cap_dn is not None) else None
            return {'up': cap_up, 'dn': cap_dn, 'asym': asym,
                    'n_up': len(up), 'n_dn': len(dn)}
        g = {}
        for sig in ['Bull', 'Base', 'Bear']:
            g[sig] = _calc([(p, b) for (s, p, b) in rows if s == sig])
        g['전체(전략)'] = _calc([(p, b) for (s, p, b) in rows])
        return g

    cap_rows = _cap_monthly_rows(cap_rebals, cap_profile, cap_equity_only)
    if not cap_rows:
        st.info('캡처 분석에 필요한 데이터가 부족합니다.')
    else:
        cap_groups = _cap_ratios(cap_rows)
        cap_bm_label = 'ACWI' if cap_equity_only else f'{cap_profile} BM'
        CAP_CATS = ['Bull', 'Base', 'Bear', '전체(전략)']
        CAP_HUE = {
            'Bull':      ('#B9842F', '#E6CFA0'),
            'Base':      ('#2F6E76', '#A9CDD1'),
            'Bear':      ('#9C3D34', '#DFACA4'),
            '전체(전략)': ('#3E5066', '#AAB7C9'),
        }
        up_vals = [cap_groups[c]['up'] for c in CAP_CATS]
        dn_vals = [cap_groups[c]['dn'] for c in CAP_CATS]

        fig_cap = go.Figure()
        fig_cap.add_trace(go.Bar(
            x=CAP_CATS, y=up_vals, name='상승 캡처',
            marker_color=[CAP_HUE[c][0] for c in CAP_CATS],
            text=[f'{v:.2f}' if v is not None else '' for v in up_vals],
            textposition='outside', textfont=dict(size=12), showlegend=False,
            hovertemplate='%{x}<br>상승 캡처: %{y:.2f}<extra></extra>',
        ))
        fig_cap.add_trace(go.Bar(
            x=CAP_CATS, y=dn_vals, name='하락 캡처',
            marker_color=[CAP_HUE[c][1] for c in CAP_CATS],
            text=[f'{v:.2f}' if v is not None else '' for v in dn_vals],
            textposition='outside', textfont=dict(size=12), showlegend=False,
            hovertemplate='%{x}<br>하락 캡처: %{y:.2f}<extra></extra>',
        ))
        # 범례 프록시 (진함 = 상승 / 연함 = 하락)
        fig_cap.add_trace(go.Scatter(
            x=[None], y=[None], mode='markers',
            marker=dict(size=13, color='#3E5066', symbol='square'),
            name='상승 캡처 (진함)'))
        fig_cap.add_trace(go.Scatter(
            x=[None], y=[None], mode='markers',
            marker=dict(size=13, color='#AAB7C9', symbol='square'),
            name='하락 캡처 (연함)'))
        # 1.00 = BM 동일 기준선
        fig_cap.add_hline(
            y=1.0, line_dash='dash', line_color='#888', line_width=1.2,
            annotation_text=f'1.00 = {cap_bm_label}와 동일',
            annotation_position='top right',
            annotation_font_size=11, annotation_font_color='#888')
        # x 라벨에 비대칭 부기
        ticktext = []
        for c in CAP_CATS:
            a = cap_groups[c]['asym']
            a_str = f'{a:+.2f}' if a is not None else '—'
            ticktext.append(
                f'<b>{c}</b><br><span style="font-size:11px;color:#666">비대칭 {a_str}</span>')
        all_vals = [v for v in up_vals + dn_vals if v is not None]
        cap_ymax = max(1.10, (max(all_vals) if all_vals else 1.0) * 1.18)
        cap_scope_label = '주식 only' if cap_equity_only else '전체 포트폴리오'
        fig_cap.update_layout(
            barmode='group', height=460, bargap=0.32, bargroupgap=0.08,
            title=dict(text=f'<b>{cap_profile} · {cap_scope_label} — BM: {cap_bm_label}</b>',
                       x=0.02, font=dict(size=14)),
            xaxis=dict(tickmode='array', tickvals=CAP_CATS, ticktext=ticktext),
            yaxis=dict(title='캡처 비율 (포트 / BM)', range=[0, cap_ymax],
                       gridcolor='#EEEEEE'),
            legend=dict(orientation='h', yanchor='bottom', y=1.02,
                        xanchor='right', x=1, font=dict(size=11)),
            margin=dict(t=60, b=60, l=60, r=30), plot_bgcolor='white',
        )
        st.plotly_chart(fig_cap, use_container_width=True, key='capture_asym_chart')

        # 요약 표
        cap_tbl = pd.DataFrame([{
            '구분': c,
            '상승 캡처': cap_groups[c]['up'],
            '하락 캡처': cap_groups[c]['dn'],
            '비대칭(상−하)': cap_groups[c]['asym'],
            '상승월(n)': cap_groups[c]['n_up'],
            '하락월(n)': cap_groups[c]['n_dn'],
        } for c in CAP_CATS])
        for _col in ['상승 캡처', '하락 캡처', '비대칭(상−하)']:
            cap_tbl[_col] = pd.to_numeric(cap_tbl[_col], errors='coerce')
        cap_styled = cap_tbl.style.format({
            '상승 캡처': '{:.3f}', '하락 캡처': '{:.3f}', '비대칭(상−하)': '{:+.3f}',
        }, na_rep='—').background_gradient(
            subset=['비대칭(상−하)'], cmap='RdYlGn', vmin=-0.3, vmax=0.3)
        st.dataframe(cap_styled, use_container_width=True, hide_index=True)

        # 전략 전체 인사이트
        g_all = cap_groups['전체(전략)']
        if g_all['asym'] is not None:
            if g_all['asym'] > 0.02:
                verdict = '상방을 더 따라가고 하방을 덜 떨어진 **방어 우위** 패턴 ✅'
            elif g_all['asym'] < -0.02:
                verdict = '하방을 상방보다 더 떨어진 **비대칭 열위** 패턴 ⚠️'
            else:
                verdict = '상·하방을 거의 같은 비율로 누르는 **중립 패턴** (비대칭 ≈ 0)'
            st.caption(
                f':grey[**전략 전체:** 상승 캡처 {g_all["up"]:.2f} / 하락 캡처 {g_all["dn"]:.2f} '
                f'→ 비대칭 {g_all["asym"]:+.2f} — {verdict}]')

        st.caption(
            ':grey[**📐 산식 (합산법):** 상승 캡처 = Σ(포트 월수익, BM↑월) / Σ(BM 월수익, BM↑월).  '
            '하락 캡처 = Σ(포트 월수익, BM↓월) / Σ(BM 월수익, BM↓월) — 둘 다 음수라 비율은 양수.  '
            '0 근처 BM월의 비율 폭주를 줄이기 위해 월평균이 아닌 **합산법** 사용.  '
            '**비대칭 = 상승 캡처 − 하락 캡처** (>0 = 이상적 방어형).  '
            '\n\n**주식 only:** 포트 = 주식 슬리브(광역+국가주식, 100% 정규화), BM = MSCI ACWI 100%.  '
            '**전체 포트폴리오:** 포트 = 실제 전체 비중, '
            'BM = profile 정적 벤치마크(적극형 ACWI 75 / BNDW 10 / CD91 15, '
            '중립형 ACWI 45 / BNDW 40 / CD91 15 · 현금=한국 CD 91일물 금리).  '
            'BM 월수익이 정확히 0인 달은 양측 모두에서 제외.  '
            '월 구간 = 각 리밸 일자 → 다음 리밸 일자(Report Date 이전 보유 1개월).]')

    # ─── 🎛️ 시그널 배분 직접 조정 → 시나리오 캡처 ───
    with st.expander('🎛️ 시그널 배분 직접 조정 → 시나리오 캡처'):
        _scope_lbl = ('주식 only (BM = ACWI 100%)' if cap_equity_only
                      else f'전체 포트폴리오 (BM = {cap_profile} 정적 BM)')
        st.caption(
            f'현재 스코프: **{_scope_lbl}** — 위 분석범위 토글과 연동.  '
            '시그널별 **주식/채권/현금 비중**(기본값 = **자산배분정책**)을 조정하면, 그 비중을 '
            '**실제 보유 슬리브 수익률**(주식·채권·현금)에 적용해 상승·하락 캡처를 재산출합니다.  '
            '**두 스코프 모두 조정 가능**. 합이 100이 아니면 자동 정규화.  '
            '(실제 포트 캡처는 위 메인 차트 참조.)')

        _RISK_CATS = {'광역 주식', '국가 주식', '원자재', '헤지펀드'}

        def _bucket(tk):
            _c = TICKER_CATEGORY.get(tk)
            return ('주식' if _c in _RISK_CATS else
                    '채권' if _c == '채권' else '현금' if _c == '현금' else None)

        _ets0 = pd.Timestamp(end_date_str)

        def _defw(s):
            _p = SIGNAL_ALLOC.get(cap_profile, {}).get(s)
            if _p:
                return (int(_p['주식']), int(_p['채권']), int(_p['현금']))
            return (90, 5, 5)

        _sig_emoji = {'Bull': '🟢', 'Base': '🟡', 'Bear': '🔴'}
        _sw = {}
        for _sig in ['Bull', 'Base', 'Bear']:
            _d = _defw(_sig)
            _c0, _c1, _c2, _c3, _c4 = st.columns([1.0, 1, 1, 1, 0.9])
            _c0.markdown(f'**{_sig_emoji[_sig]} {_sig}**')
            _we = _c1.number_input('주식%', 0, 100, _d[0], 1, key=f'scen_{cap_profile}_{_sig}_eq')
            _wb = _c2.number_input('채권%', 0, 100, _d[1], 1, key=f'scen_{cap_profile}_{_sig}_bd')
            _wc = _c3.number_input('현금%', 0, 100, _d[2], 1, key=f'scen_{cap_profile}_{_sig}_ca')
            _sm = _we + _wb + _wc
            _c4.markdown(f'합 **{_sm}**' + (' · 조정' if (_we, _wb, _wc) != _d else ' · 정책'))
            _sw[_sig] = (_we, _wb, _wc)

        def _sleeve_daily(weights, idx, which):
            _sel = {tk: v for tk, v in weights.items() if _bucket(tk) == which}
            _t = sum(_sel.values())
            _dd = pd.Series(0.0, index=idx)
            if _t <= 0:
                return _dd
            for _tk, _v in _sel.items():
                _sym = TICKER_MAP.get(_tk)
                if _sym is not None and _sym in px.columns:
                    _dd = _dd + (_v / _t) * px[_sym].reindex(idx).pct_change().fillna(0)
            return _dd

        def _scen_rows():
            rows = []
            if 'ACWI' not in px.columns:
                return rows
            _pu = px[px.index <= _ets0]
            if len(_pu) < 2:
                return rows
            for i in range(len(cap_rebals)):
                _rb = pd.Timestamp(cap_rebals[i][0])
                if _rb > _ets0:
                    break
                _nx = (min(pd.Timestamp(cap_rebals[i + 1][0]), _ets0)
                       if i + 1 < len(cap_rebals) else _ets0)
                _ix = _pu.index[(_pu.index >= _rb) & (_pu.index <= _nx)]
                if len(_ix) < 2:
                    continue
                _sg = cap_rebals[i][1]
                _w = cap_rebals[i][2]
                _e, _b, _c = _sw.get(_sg, _defw(_sg))
                _tt = _e + _b + _c
                if _tt <= 0:
                    continue
                # 시나리오 포트 = 입력(정책/조정) 비중 × 실제 슬리브 수익률 (두 스코프 공통)
                _pd = ((_e / _tt) * _sleeve_daily(_w, _ix, '주식')
                       + (_b / _tt) * _sleeve_daily(_w, _ix, '채권')
                       + (_c / _tt) * _sleeve_daily(_w, _ix, '현금'))
                _port = ((1 + _pd).prod() - 1) * 100
                if cap_equity_only:
                    _bm = ((1 + px['ACWI'].reindex(_ix).pct_change().fillna(0)).prod() - 1) * 100
                else:
                    _bm = ((1 + compute_benchmark_daily_return(cap_profile, px, _ix)).prod() - 1) * 100
                rows.append((_sg, _port, _bm))
            return rows

        def _cap_bar_fig(groups, bm_label):
            _cats = ['Bull', 'Base', 'Bear', '전체(전략)']
            _hue = {'Bull': ('#B9842F', '#E6CFA0'), 'Base': ('#2F6E76', '#A9CDD1'),
                    'Bear': ('#9C3D34', '#DFACA4'), '전체(전략)': ('#3E5066', '#AAB7C9')}
            _uv = [groups[c]['up'] for c in _cats]
            _dv = [groups[c]['dn'] for c in _cats]
            _f = go.Figure()
            _f.add_trace(go.Bar(x=_cats, y=_uv, marker_color=[_hue[c][0] for c in _cats],
                showlegend=False, text=[f'{v:.2f}' if v is not None else '' for v in _uv],
                textposition='outside', textfont=dict(size=12)))
            _f.add_trace(go.Bar(x=_cats, y=_dv, marker_color=[_hue[c][1] for c in _cats],
                showlegend=False, text=[f'{v:.2f}' if v is not None else '' for v in _dv],
                textposition='outside', textfont=dict(size=12)))
            _f.add_trace(go.Scatter(x=[None], y=[None], mode='markers',
                marker=dict(size=13, color='#3E5066', symbol='square'), name='상승 캡처 (진함)'))
            _f.add_trace(go.Scatter(x=[None], y=[None], mode='markers',
                marker=dict(size=13, color='#AAB7C9', symbol='square'), name='하락 캡처 (연함)'))
            _f.add_hline(y=1.0, line_dash='dash', line_color='#888', line_width=1.2,
                annotation_text=f'1.00 = {bm_label}와 동일', annotation_position='top right',
                annotation_font_size=11, annotation_font_color='#888')
            _tt = []
            for c in _cats:
                _a = groups[c]['asym']
                _astr = f'{_a:+.2f}' if _a is not None else '—'
                _tt.append(f'<b>{c}</b><br><span style="font-size:11px;color:#666">비대칭 {_astr}</span>')
            _allv = [v for v in _uv + _dv if v is not None]
            _ymax = max(1.10, (max(_allv) if _allv else 1.0) * 1.18)
            _f.update_layout(barmode='group', height=440, bargap=0.32,
                xaxis=dict(tickmode='array', tickvals=_cats, ticktext=_tt),
                yaxis=dict(title='캡처 비율 (포트 / BM)', range=[0, _ymax], gridcolor='#EEEEEE'),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1,
                            font=dict(size=11)),
                margin=dict(t=60, b=60, l=60, r=30), plot_bgcolor='white')
            return _f

        _srows = _scen_rows()
        if not _srows:
            st.info('시나리오 캡처 산출에 필요한 데이터가 부족합니다.')
        else:
            _sg2 = _cap_ratios(_srows)
            _bml = 'ACWI' if cap_equity_only else f'{cap_profile} BM'
            st.plotly_chart(_cap_bar_fig(_sg2, _bml),
                            use_container_width=True, key='capture_scenario_chart')
            _sc_tbl = pd.DataFrame([{
                '구분': c,
                '상승 캡처': _sg2[c]['up'],
                '하락 캡처': _sg2[c]['dn'],
                '비대칭(상−하)': _sg2[c]['asym'],
                '상승월(n)': _sg2[c]['n_up'],
                '하락월(n)': _sg2[c]['n_dn'],
            } for c in ['Bull', 'Base', 'Bear', '전체(전략)']])
            for _col in ['상승 캡처', '하락 캡처', '비대칭(상−하)']:
                _sc_tbl[_col] = pd.to_numeric(_sc_tbl[_col], errors='coerce')
            st.dataframe(_sc_tbl.style.format({
                '상승 캡처': '{:.3f}', '하락 캡처': '{:.3f}', '비대칭(상−하)': '{:+.3f}'},
                na_rep='—').background_gradient(
                subset=['비대칭(상−하)'], cmap='RdYlGn', vmin=-0.3, vmax=0.3),
                use_container_width=True, hide_index=True)
            _chg = ', '.join(s for s in ['Bull', 'Base', 'Bear'] if _sw.get(s) != _defw(s))
            _ctxt = (f'**조정 적용: {_chg}.** ' if _chg
                     else '**기본값 = 자산배분정책** (적극 90/5/5·75/10/15·60/25/15, 중립 60/35/5·45/40/15·30/55/15). ')
            st.caption(
                f':grey[{_ctxt}시나리오 포트 = w_주식·실제 주식슬리브 + w_채권·실제 채권슬리브 + w_현금·실제 현금 '
                f'(일별 compound).  BM = {_bml}.  주식↑ → 캡처↑, 채권·현금↑ → 하락 캡처↓(방어).  합산법.]')

    # ─── ❓ 원인 분석 버튼 (왜 Base·주식only에서 하락 캡처 > 상승 캡처인가) ───
    with st.expander('❓ 왜 Base·주식only에서 하락 캡처 > 상승 캡처인가 — 원인 분석 보기'):
        # 현재 표본 기준 Base·주식only 캡처 (라이브 계산 — 기존 헬퍼 재사용)
        def _cap_fmt(v):
            return f'{v:.3f}' if v is not None else '—'
        _base_cap = {}
        for _pf, _rb in [('적극형', agg_rebals), ('중립형', neu_rebals)]:
            try:
                _base_cap[_pf] = _cap_ratios(_cap_monthly_rows(_rb, _pf, True))['Base']
            except Exception:
                _base_cap[_pf] = {'up': None, 'dn': None, 'asym': None}
        _ba, _bn = _base_cap['적극형'], _base_cap['중립형']
        st.markdown(
            '**현재 표본 기준 (Base · 주식only, BM = ACWI 100%)**\n\n'
            '| 프로파일 | 상승 캡처 | 하락 캡처 | 비대칭(상−하) |\n'
            '|---|---|---|---|\n'
            f'| 적극형 | {_cap_fmt(_ba["up"])} | {_cap_fmt(_ba["dn"])} | {_cap_fmt(_ba["asym"])} |\n'
            f'| 중립형 | {_cap_fmt(_bn["up"])} | {_cap_fmt(_bn["dn"])} | {_cap_fmt(_bn["asym"])} |\n\n'
            '➡️ 두 프로파일 모두 **하락 캡처 > 상승 캡처** (음의 비대칭).'
        )
        st.markdown(_CAPTURE_RCA_MD)
        st.caption(
            ':grey[※ 본문의 β·α·ETF 기여 등 세부 수치는 **전체 표본(2022-03~2026-05, Report Date 2026-05-22)** '
            '기준 측정값입니다. **방향(하락>상승)은 표본·레짐에 견고**하나 **수치 크기는 표본에 민감**합니다.  '
            '검증: 10-에이전트 워크플로우(산술 재현·leave-one-out·베타 대칭성 3중 독립 확인).]'
        )

    # ─── 📉 위험·회복 지표 (적극형 vs 중립형) ───
    st.markdown('---')
    st.markdown('### 📉 위험·회복 지표 — MDD · 회복기간 · 롤링 연환산수익')
    st.caption(
        '연속체인 정정 엔진 기준 **실제(스왑 0%)** 일별 수익률로 산출.  '
        '**MDD** = 최대 낙폭, **회복기간** = MDD 저점 → 직전 고점 복귀까지, '
        '**롤링 연환산** = 모든 트레일링 12/36개월 창의 연환산 수익률 평균.')

    def _risk_metrics(daily):
        if daily is None or len(daily) < 30:
            return None
        cum = (1 + daily).cumprod()
        dd = cum / cum.cummax() - 1
        mdd = float(dd.min()) * 100
        trough = dd.idxmin()
        peak = cum.loc[:trough].idxmax()
        peak_val = cum.loc[peak]
        after = cum.loc[trough:]
        rec = after[after >= peak_val]
        if len(rec) > 0:
            rec_date = rec.index[0]
            rec_days = int((rec_date - trough).days)
            recovered = True
        else:
            rec_date = None
            rec_days = int((cum.index[-1] - trough).days)
            recovered = False

        def _roll(months):
            win = int(round(months / 12 * 252))
            if len(cum) <= win:
                return None
            ratio = (cum / cum.shift(win)).dropna()
            if len(ratio) == 0:
                return None
            return float((ratio ** (12.0 / months) - 1.0).mean()) * 100

        return {'mdd': mdd, 'peak': peak, 'trough': trough, 'rec_date': rec_date,
                'rec_days': rec_days, 'recovered': recovered,
                'roll12': _roll(12), 'roll36': _roll(36)}

    _rm_agg = _risk_metrics(agg_actual)
    _rm_neu = _risk_metrics(neu_actual)
    if _rm_agg is None or _rm_neu is None:
        st.info('위험 지표 산출에 필요한 데이터가 부족합니다.')
    else:
        _RM_CATS = ['MDD (%)', '12M 롤링 연환산 (%)', '36M 롤링 연환산 (%)']

        def _rm_vals(m):
            return [round(m['mdd'], 2),
                    round(m['roll12'], 2) if m['roll12'] is not None else None,
                    round(m['roll36'], 2) if m['roll36'] is not None else None]

        _agg_v, _neu_v = _rm_vals(_rm_agg), _rm_vals(_rm_neu)
        fig_rm = go.Figure()
        fig_rm.add_trace(go.Bar(
            x=_RM_CATS, y=_agg_v, name='적극형', marker_color='#1F3A68',
            text=[f'{v:+.2f}' if v is not None else '' for v in _agg_v],
            textposition='outside', textfont=dict(size=12)))
        fig_rm.add_trace(go.Bar(
            x=_RM_CATS, y=_neu_v, name='중립형', marker_color='#C48D43',
            text=[f'{v:+.2f}' if v is not None else '' for v in _neu_v],
            textposition='outside', textfont=dict(size=12)))
        fig_rm.add_hline(y=0, line_dash='dash', line_color='#888', line_width=1)
        fig_rm.update_layout(
            barmode='group', height=420, plot_bgcolor='white',
            yaxis=dict(title='%', gridcolor='#EEEEEE'),
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1,
                        font=dict(size=11)),
            margin=dict(t=56, b=40, l=50, r=20),
            title=dict(text='<b>적극형 vs 중립형 — 위험·롤링수익</b>', x=0.02, font=dict(size=14)))
        st.plotly_chart(fig_rm, use_container_width=True, key='risk_metrics_chart')

        def _rec_str(m):
            mo = m['rec_days'] / 30.44
            if m['recovered']:
                return f"{mo:.1f}개월 ({m['rec_days']}일)"
            return f"미회복 (저점 후 {mo:.1f}개월 경과)"

        _rm_tbl = pd.DataFrame({
            '지표': ['MDD (%)', 'MDD 회복기간 (저점→복귀)',
                   '12M 롤링 연환산 (평균, %)', '36M 롤링 연환산 (평균, %)'],
            '적극형': [f"{_rm_agg['mdd']:.2f}", _rec_str(_rm_agg),
                     f"{_rm_agg['roll12']:+.2f}" if _rm_agg['roll12'] is not None else '—',
                     f"{_rm_agg['roll36']:+.2f}" if _rm_agg['roll36'] is not None else '—'],
            '중립형': [f"{_rm_neu['mdd']:.2f}", _rec_str(_rm_neu),
                     f"{_rm_neu['roll12']:+.2f}" if _rm_neu['roll12'] is not None else '—',
                     f"{_rm_neu['roll36']:+.2f}" if _rm_neu['roll36'] is not None else '—'],
        })
        st.dataframe(_rm_tbl, use_container_width=True, hide_index=True)
        st.caption(
            f':grey[**적극형 MDD** {_rm_agg["mdd"]:.2f}% '
            f'(직전고점 {_rm_agg["peak"].date()} → 저점 {_rm_agg["trough"].date()}) · '
            f'**중립형 MDD** {_rm_neu["mdd"]:.2f}% (저점 {_rm_neu["trough"].date()}).  '
            'MDD 회복기간 = 저점 → 직전 고점 복귀까지 달력일.  '
            '롤링 연환산 = 일별 트레일링 12/36M(≈252/756영업일) 창 수익률을 연환산 후 단순평균.  '
            '실제(스왑 0%) · 연속체인 엔진 기준.]')

    # ─── 🎛️ 시그널 배분 직접 조정 → 시나리오 위험·회복 지표 ───
    with st.expander('🎛️ 시그널 배분 직접 조정 → 시나리오 위험·회복 지표'):
        st.caption(
            '시그널별 주식/채권/현금 비중(**기본값 = 자산배분정책**)을 조정하면, 그 비중을 **실제 보유 슬리브 구성에 '
            '비례 적용**해 연속체인 재시뮬레이션 후 MDD·회복기간·롤링 연환산을 재산출합니다.  '
            '(실제 포트 위험지표는 위 메인 표 참조.)')
        _rscn_prof = st.radio('프로파일', ['적극형', '중립형'], horizontal=True, key='rscn_profile')
        _rscn_reb = [r for r in (agg_rebals if _rscn_prof == '적극형' else neu_rebals)
                     if pd.Timestamp(r[0]) <= pd.Timestamp(end_date_str)]
        _rscn_base = agg_actual if _rscn_prof == '적극형' else neu_actual

        _RISK_CATS_R = {'광역 주식', '국가 주식', '원자재', '헤지펀드'}

        def _bktR(tk):
            _c = TICKER_CATEGORY.get(tk)
            return ('주식' if _c in _RISK_CATS_R else
                    '채권' if _c == '채권' else '현금' if _c == '현금' else None)

        def _dwR(s):
            _p = SIGNAL_ALLOC.get(_rscn_prof, {}).get(s)
            if _p:
                return (int(_p['주식']), int(_p['채권']), int(_p['현금']))
            return (90, 5, 5)

        _emjR = {'Bull': '🟢', 'Base': '🟡', 'Bear': '🔴'}
        _rovr = {}
        for _sig in ['Bull', 'Base', 'Bear']:
            _d = _dwR(_sig)
            _r0, _r1, _r2, _r3, _r4 = st.columns([1.0, 1, 1, 1, 0.9])
            _r0.markdown(f'**{_emjR[_sig]} {_sig}**')
            _e = _r1.number_input('주식%', 0, 100, _d[0], 1, key=f'rscn_{_rscn_prof}_{_sig}_eq')
            _b = _r2.number_input('채권%', 0, 100, _d[1], 1, key=f'rscn_{_rscn_prof}_{_sig}_bd')
            _c = _r3.number_input('현금%', 0, 100, _d[2], 1, key=f'rscn_{_rscn_prof}_{_sig}_ca')
            _r4.markdown(f'합 **{_e + _b + _c}**' + (' · 조정' if (_e, _b, _c) != _d else ' · 정책'))
            _rovr[_sig] = (_e, _b, _c)

        def _apply_ovr(rebals, ov):
            out = []
            for _d, _s, _w in rebals:
                o = ov.get(_s)
                if o is None:
                    out.append((_d, _s, _w)); continue
                _e, _b, _c = o; _t = _e + _b + _c
                if _t <= 0:
                    out.append((_d, _s, _w)); continue
                _eq = {tk: v for tk, v in _w.items() if _bktR(tk) == '주식'}
                _bd = {tk: v for tk, v in _w.items() if _bktR(tk) == '채권'}
                _ca = {tk: v for tk, v in _w.items() if _bktR(tk) == '현금'}
                _se, _sb, _sc = sum(_eq.values()), sum(_bd.values()), sum(_ca.values())
                _nw = {}
                if _se > 0:
                    for tk, v in _eq.items():
                        _nw[tk] = (_e / _t) * (v / _se) * 100
                if _sb > 0:
                    for tk, v in _bd.items():
                        _nw[tk] = (_b / _t) * (v / _sb) * 100
                if _sc > 0:
                    for tk, v in _ca.items():
                        _nw[tk] = (_c / _t) * (v / _sc) * 100
                out.append((_d, _s, _nw if _nw else _w))
            return out

        _any_r = any(_rovr.get(s) != _dwR(s) for s in ['Bull', 'Base', 'Bear'])
        _mod_reb = _apply_ovr(_rscn_reb, _rovr)
        _scen_daily = simulate_portfolio(_mod_reb, px, 0.0, end_date=end_date_str)
        _rm_base = _risk_metrics(_rscn_base)
        _rm_scen = _risk_metrics(_scen_daily)
        if _rm_base is None or _rm_scen is None:
            st.info('시나리오 위험 지표 산출에 필요한 데이터가 부족합니다.')
        else:
            _catsR = ['MDD (%)', '12M 롤링 연환산 (%)', '36M 롤링 연환산 (%)']

            def _vvR(m):
                return [round(m['mdd'], 2),
                        round(m['roll12'], 2) if m['roll12'] is not None else None,
                        round(m['roll36'], 2) if m['roll36'] is not None else None]

            _sv = _vvR(_rm_scen)
            _pcR = '#1F3A68' if _rscn_prof == '적극형' else '#C48D43'
            _figR = go.Figure()
            _figR.add_trace(go.Bar(x=_catsR, y=_sv, name='시나리오', marker_color=_pcR,
                text=[f'{v:+.2f}' if v is not None else '' for v in _sv],
                textposition='outside', textfont=dict(size=12)))
            _figR.add_hline(y=0, line_dash='dash', line_color='#888', line_width=1)
            _figR.update_layout(barmode='group', height=400, plot_bgcolor='white', showlegend=False,
                yaxis=dict(title='%', gridcolor='#EEEEEE'),
                margin=dict(t=50, b=40, l=50, r=20),
                title=dict(text=f'<b>{_rscn_prof} — 시나리오 위험·회복</b>', x=0.02, font=dict(size=13)))
            st.plotly_chart(_figR, use_container_width=True, key='risk_scenario_chart')

            def _rsR(m):
                _mo = m['rec_days'] / 30.44
                return f"{_mo:.1f}개월" if m['recovered'] else f"미회복 ({_mo:.1f}개월+)"

            _tR = pd.DataFrame({
                '지표': ['MDD (%)', 'MDD 회복기간', '12M 롤링 연환산 (%)', '36M 롤링 연환산 (%)'],
                '시나리오': [f"{_rm_scen['mdd']:.2f}", _rsR(_rm_scen),
                         f"{_rm_scen['roll12']:+.2f}", f"{_rm_scen['roll36']:+.2f}"],
            })
            st.dataframe(_tR, use_container_width=True, hide_index=True)
            if not _any_r:
                st.caption(':grey[**기본값 = 자산배분정책** (적극 90/5/5·75/10/15·60/25/15, 중립 60/35/5·45/40/15·30/55/15). '
                           '비중을 바꾸면 그 시그널 월이 재계산됩니다. 실제 포트 위험지표는 위 메인 표 참조.]')
            else:
                _chR = ', '.join(s for s in ['Bull', 'Base', 'Bear'] if _rovr.get(s) != _dwR(s))
                st.caption(
                    f':grey[**조정 적용: {_chR}** (기본=자산배분정책) — 시그널 월별 비중(주식/채권/현금)을 '
                    '실제 슬리브 종목 구성에 비례 적용 → 연속체인 재시뮬레이션 후 MDD·회복·롤링 재산출.]')

    # ─── 📈 월별 신호 × 주식·채권 방향성 추이 ───
    st.markdown('---')
    st.markdown('### 📈 월별 신호 × 주식·채권 방향성 추이')
    st.caption(
        '각 리밸 월의 **시그널(Bull·Base·Bear)** 과 **주식(ACWI)·채권(BNDW) 월수익률**을 하나의 추이 차트에 표시.  '
        '막대 방향으로 주식·채권 국면을, 상단 ■ 색으로 시그널을 봅니다.  아래 슬라이더로 기간 선택.')

    _dir_prof = st.radio('프로파일 (시그널 기준)', ['적극형', '중립형'],
                         horizontal=True, key='dir_prof')
    _dir_reb = [r for r in (agg_rebals if _dir_prof == '적극형' else neu_rebals)
                if pd.Timestamp(r[0]) <= pd.Timestamp(end_date_str)]
    _ets_d = pd.Timestamp(end_date_str)
    _pud = px[px.index <= _ets_d]
    _dir_rows = []
    for i in range(len(_dir_reb)):
        _rb = pd.Timestamp(_dir_reb[i][0])
        if _rb > _ets_d:
            break
        _nx = (min(pd.Timestamp(_dir_reb[i + 1][0]), _ets_d)
               if i + 1 < len(_dir_reb) else _ets_d)
        _ix = _pud.index[(_pud.index >= _rb) & (_pud.index <= _nx)]
        if len(_ix) < 2:
            continue
        _ra = (((1 + px['ACWI'].reindex(_ix).pct_change().fillna(0)).prod() - 1) * 100
               if 'ACWI' in px.columns else 0.0)
        _rbd = (((1 + px['BNDW'].reindex(_ix).pct_change().fillna(0)).prod() - 1) * 100
                if 'BNDW' in px.columns else 0.0)
        _dir_rows.append((_dir_reb[i][0], _dir_reb[i][1], round(_ra, 2), round(_rbd, 2)))

    if len(_dir_rows) < 2:
        st.info('추이 산출에 필요한 데이터가 부족합니다.')
    else:
        _months = [r[0] for r in _dir_rows]
        _rng = st.select_slider('기간 선택', options=_months,
                                value=(_months[0], _months[-1]), key='dir_range')
        _i0, _i1 = _months.index(_rng[0]), _months.index(_rng[1])
        if _i0 > _i1:
            _i0, _i1 = _i1, _i0
        _sub = _dir_rows[_i0:_i1 + 1]
        _mx = [r[0] for r in _sub]
        _eq = [r[2] for r in _sub]
        _bd = [r[3] for r in _sub]
        _sgs = [r[1] for r in _sub]
        _SCLR = {'Bull': '#10B981', 'Base': '#F59E0B', 'Bear': '#DC2626'}
        _figD = go.Figure()
        _figD.add_trace(go.Bar(x=_mx, y=_eq, name='주식 (ACWI) 월수익', marker_color='#1F3A68'))
        _figD.add_trace(go.Bar(x=_mx, y=_bd, name='채권 (BNDW) 월수익', marker_color='#C48D43'))
        _figD.add_hline(y=0, line_dash='dash', line_color='#888', line_width=1)
        _allv = _eq + _bd
        _base = max((max(_allv) if _allv else 0.0), 0.0)
        _ytop = _base + 2.4    # 실제 신호 (위 행)
        _ytop2 = _base + 1.0   # 정답 신호 (아래 행)
        # 정답 신호: ACWI 월수익 > BNDW 월수익 → Bull, 아니면 Bear
        _ans = ['Bull' if e > b else 'Bear' for e, b in zip(_eq, _bd)]
        # 실제 신호 (■, 상단)
        for _s in ['Bull', 'Base', 'Bear']:
            _xs = [m for m, sg in zip(_mx, _sgs) if sg == _s]
            if not _xs:
                continue
            _figD.add_trace(go.Scatter(
                x=_xs, y=[_ytop] * len(_xs), mode='markers', name=f'실제 {_s}',
                marker=dict(color=_SCLR[_s], size=11, symbol='square',
                            line=dict(color='white', width=1)),
                hovertemplate='%{x}<br>실제 신호: ' + _s + '<extra></extra>'))
        # 정답 신호 (●, 하단) — Bull/Bear만
        for _s in ['Bull', 'Bear']:
            _xs = [m for m, a in zip(_mx, _ans) if a == _s]
            if not _xs:
                continue
            _figD.add_trace(go.Scatter(
                x=_xs, y=[_ytop2] * len(_xs), mode='markers', name=f'정답 {_s}',
                marker=dict(color=_SCLR[_s], size=10, symbol='circle',
                            line=dict(color='white', width=1)),
                hovertemplate='%{x}<br>정답 신호: ' + _s + '<extra></extra>'))
        _figD.add_annotation(xref='paper', x=0, y=_ytop, yref='y', text='실제 ',
                             showarrow=False, xanchor='right', font=dict(size=10, color='#888'))
        _figD.add_annotation(xref='paper', x=0, y=_ytop2, yref='y', text='정답 ',
                             showarrow=False, xanchor='right', font=dict(size=10, color='#888'))
        _figD.update_layout(
            barmode='group', height=460, plot_bgcolor='white',
            xaxis_title='리밸 월', yaxis_title='월수익률 (%)',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0, font=dict(size=10)),
            margin=dict(t=50, b=70, l=50, r=20))
        st.plotly_chart(_figD, use_container_width=True, key='dir_trend_chart')
        st.caption(
            f':grey[**기간: {_rng[0]} ~ {_rng[1]} ({len(_sub)}개월).**  '
            '**실제 신호(■, 위)** = 🟢 Bull / 🟡 Base / 🔴 Bear.  '
            '**정답 신호(●, 아래)** = ACWI 월수익 **>** BNDW 월수익 → 🟢 Bull, **≤** → 🔴 Bear.  '
            '막대 = 주식(ACWI)·채권(BNDW) 월수익률.  '
            '주식↑·채권↑ = 위험선호, 주식↓·채권↑ = 안전선호(방어), 둘 다↓ = 동반 약세.  '
            '실제 신호 시퀀스는 선택 프로파일 기준(Bull/Base 공통, Bear만 상이).]')

        # ─── 📊 신호 적중 통계 (실제 vs 정답) ───
        st.markdown('##### 📊 신호 적중 통계 (실제 vs 정답)')
        # 방향 가치 = 신호 방향대로 ACWI−BNDW 롱숏을 취했을 때의 월간 손익
        #   Bull(위험선호) → +(ACWI−BNDW),  방어(Base·Bear) → −(ACWI−BNDW)
        _recs = []
        for _sg, _an, _ea, _eb in zip(_sgs, _ans, _eq, _bd):
            _adir = 'Bull' if _sg == 'Bull' else 'Bear'
            _sprd = _ea - _eb
            _val = _sprd if _adir == 'Bull' else -_sprd
            _recs.append((_sg, _an, _adir == _an, _val))
        _n = len(_recs)
        _nhit = sum(1 for r in _recs if r[2])
        _nmiss = _n - _nhit
        _hr = _nhit / _n * 100 if _n else 0.0
        _hit_r = [r[3] for r in _recs if r[2]]
        _miss_r = [r[3] for r in _recs if not r[2]]
        _avh = float(np.mean(_hit_r)) if _hit_r else 0.0
        _avm = float(np.mean(_miss_r)) if _miss_r else 0.0
        _gap = _avh - _avm
        _lost = _gap * _nmiss

        _k1, _k2, _k3, _k4 = st.columns(4)
        _k1.metric('평균 정답률', f'{_hr:.1f}%', f'{_nhit}/{_n}월')
        _k2.metric('적중 시 평균 방향가치', f'{_avh:+.2f}%p')
        _k3.metric('미적중 시 평균 방향가치', f'{_avm:+.2f}%p')
        _k4.metric('미적중 누적 비용', f'{_lost:+.2f}%p', f'격차 {_gap:+.2f}%p × {_nmiss}월',
                   delta_color='off')

        _sc1, _sc2 = st.columns([1.1, 1])
        with _sc1:
            st.caption('혼동 행렬 — 실제 신호 × 정답 (월 수)')
            _cm = {s: {'Bull': 0, 'Bear': 0} for s in ['Bull', 'Base', 'Bear']}
            for _sg, _an, _h, _v in _recs:
                _cm[_sg][_an] += 1
            _cm_df = pd.DataFrame({
                '실제 신호': ['🟢 Bull', '🟡 Base', '🔴 Bear'],
                '정답=Bull': [_cm[s]['Bull'] for s in ['Bull', 'Base', 'Bear']],
                '정답=Bear': [_cm[s]['Bear'] for s in ['Bull', 'Base', 'Bear']],
            })
            st.dataframe(_cm_df, use_container_width=True, hide_index=True)
        with _sc2:
            _fb = go.Figure()
            _fb.add_trace(go.Bar(x=['적중', '미적중'], y=[round(_avh, 2), round(_avm, 2)],
                marker_color=['#10B981', '#DC2626'],
                text=[f'{_avh:+.2f}%p', f'{_avm:+.2f}%p'], textposition='outside',
                textfont=dict(size=12)))
            _fb.add_hline(y=0, line_dash='dot', line_color='#888', line_width=1)
            _fb.update_layout(height=250, plot_bgcolor='white', showlegend=False,
                yaxis=dict(title='평균 방향가치 (%p)', gridcolor='#EEEEEE'),
                margin=dict(t=10, b=20, l=45, r=10))
            st.plotly_chart(_fb, use_container_width=True, key='dir_hit_bar')

        _stat_df = pd.DataFrame({
            '통계 항목': ['표본 (월)', '적중 (월)', '미적중 (월)', '평균 정답률 (%)',
                       '적중 평균 방향가치 (%p)', '미적중 평균 방향가치 (%p)',
                       '적중−미적중 격차 (%p)', '미적중 누적 비용 (%p)'],
            '값': [_n, _nhit, _nmiss, round(_hr, 1), round(_avh, 2), round(_avm, 2),
                  round(_gap, 2), round(_lost, 2)],
        })
        st.dataframe(_stat_df, use_container_width=True, hide_index=True)
        st.caption(
            ':grey[**정의** — 정답: ACWI 월수익 > BNDW → Bull, ≤ → Bear.  '
            '실제 이진방향: Bull → 위험선호(Bull), Base·Bear → 방어(Bear).  **적중** = 실제 방향 == 정답.  '
            '**방향 가치** = 신호 방향대로 주식−채권 롱숏을 취했을 때의 월간 손익 '
            '(Bull → +(ACWI−BNDW), 방어 → −(ACWI−BNDW)) → 적중이면 +, 미적중이면 −.  '
            '**미적중 누적 비용** = (적중 평균 − 미적중 평균) × 미적중 월수 = 미적중에서 적중 대비 놓친 누적 방향가치.  '
            '기간 슬라이더 선택 구간 기준.]')

    st.markdown('---')
    pie_cols = st.columns(2)
    with pie_cols[0]:
        st.markdown(f'### 🥧 적극형 자산배분 ({agg_rebals[-1][0]})')
        st.plotly_chart(build_allocation_pie(agg_rebals[-1][2], ''),
                        use_container_width=True, key='alloc_pie_agg')
    with pie_cols[1]:
        st.markdown(f'### 🥧 중립형 자산배분 ({neu_rebals[-1][0]})')
        st.plotly_chart(build_allocation_pie(neu_rebals[-1][2], ''),
                        use_container_width=True, key='alloc_pie_neu')

    # ─── 📊 통합 배분 비교 (종목·국가·섹터·팩터 × 다기간) ───
    st.markdown('---')
    st.subheader('📊 포트폴리오 배분 비교 (Look-Through · 다기간)')
    st.caption(
        '종목·국가·섹터·팩터를 **하나의 토글**로 전환하며, 현재 대비 **1·3·6·12개월 전** 포트폴리오와 비교합니다.  '
        '국가·섹터·팩터는 **주식 슬리브 100% 정규화**(적극·중립 동일), 종목별은 **전체 포트폴리오 비중**(프로파일별 상이).'
    )

    _ac1, _ac2, _ac3 = st.columns([1, 1.5, 1.5])
    with _ac1:
        alloc_profile = st.radio('프로파일', ['적극형', '중립형'],
                                 horizontal=True, key='alloc_cmp_profile')
    with _ac2:
        alloc_dim = st.radio('분류 기준', ['종목별', '국가별', '섹터별', '팩터별'],
                             horizontal=True, key='alloc_cmp_dim')
    with _ac3:
        alloc_lb = st.radio('비교 시점',
                            ['1개월 전', '3개월 전', '6개월 전', '1년 전', '특정시점'],
                            horizontal=True, key='alloc_cmp_lookback')
    alloc_rebals = agg_rebals if alloc_profile == '적극형' else neu_rebals
    _alloc_asof = None
    if alloc_lb == '특정시점':
        _fs_d = pd.Timestamp(alloc_rebals[0][0]).date() if alloc_rebals else report_date
        _def_d = max(_fs_d, report_date - timedelta(days=180))
        _alloc_asof = st.date_input(
            '📅 비교 기준일 선택 (as-of)', value=_def_d,
            min_value=_fs_d, max_value=report_date, key='alloc_cmp_asof',
            help='이 날짜에 유효했던 포트폴리오(해당 일자 이하 가장 가까운 과거 리밸)와 현재를 비교합니다.')
        _alloc_months = None
    else:
        _alloc_months = {'1개월 전': 1, '3개월 전': 3, '6개월 전': 6, '1년 전': 12}[alloc_lb]

    def _rebal_as_of(rebals, months_back):
        """현재 리밸 기준 months_back개월 전 시점에 유효했던 리밸 반환 (가장 가까운 과거)."""
        if not rebals:
            return None
        if months_back <= 0 or len(rebals) < 2:
            return rebals[-1]
        target = pd.Timestamp(rebals[-1][0]) - pd.DateOffset(months=months_back)
        cands = [r for r in rebals if pd.Timestamp(r[0]) <= target]
        return cands[-1] if cands else rebals[0]

    def _rebal_as_of_date(rebals, the_date):
        """지정 일자에 유효했던(해당 일자 이하 가장 가까운 과거) 리밸 반환."""
        if not rebals:
            return None
        t = pd.Timestamp(the_date)
        cands = [r for r in rebals if pd.Timestamp(r[0]) <= t]
        return cands[-1] if cands else rebals[0]

    def _alloc_breakdown(weights, dim):
        """선택 분류 기준의 {라벨: 비중%}."""
        if dim == '국가별':
            return dict(compute_country_breakdown(weights, top_n=999)[0])
        if dim == '섹터별':
            return dict(compute_sector_breakdown(weights, top_n=999)[0])
        if dim == '팩터별':
            return compute_factor_breakdown_msci(weights)[0]
        tot = sum(weights.values())
        if tot <= 0:
            return {}
        return {TICKER_NAMES.get(tk, tk): w / tot * 100 for tk, w in weights.items()}

    _cur_reb = alloc_rebals[-1] if alloc_rebals else None
    if alloc_lb == '특정시점' and _alloc_asof is not None:
        _past_reb = _rebal_as_of_date(alloc_rebals, _alloc_asof)
    else:
        _past_reb = _rebal_as_of(alloc_rebals, _alloc_months)
    _dim_unit = '종목' if alloc_dim == '종목별' else alloc_dim.replace('별', '')
    _prof_color = '#1F3A68' if alloc_profile == '적극형' else '#C48D43'

    if _cur_reb is None or _past_reb is None:
        st.info('비교에 필요한 리밸런싱 데이터가 부족합니다.')
    else:
        # ===== 비중 비교 (TIME ETF 스타일 2단: 현재 | 과거, 전체 종목·증감·신규) =====
        if _past_reb[0] == _cur_reb[0]:
            st.warning(f'선택한 비교 시점({alloc_lb} → {_past_reb[0]})이 현재 리밸({_cur_reb[0]})과 동일합니다. '
                       '다른 시점(또는 더 이른 특정일)을 선택하면 비교됩니다.')
        _cur_bd = _alloc_breakdown(_cur_reb[2], alloc_dim)
        _past_bd = _alloc_breakdown(_past_reb[2], alloc_dim)

        _ALLOC_TBL_CSS = """
<style>
.atbl{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:4px;}
.atbl th{text-align:left;padding:6px 8px;color:#66605C;font-weight:600;font-size:12px;
 border-bottom:1px solid rgba(51,48,46,0.22);}
.atbl th.ar,.atbl td.awt,.atbl td.achg{text-align:right;}
.atbl td{padding:6px 8px;border-bottom:1px solid rgba(51,48,46,0.08);}
.atbl tbody tr:hover{background:rgba(15,84,153,0.05);}
.atbl .ark{color:#0F5499;font-weight:700;width:30px;}
.atbl .anm{color:#33302E;}
.atbl .awt{font-weight:600;font-variant-numeric:tabular-nums;white-space:nowrap;}
.atbl .achg{font-variant-numeric:tabular-nums;font-size:12px;white-space:nowrap;}
.atbl .apos{color:#C0392B;}
.atbl .aneg{color:#0F5499;}
.atbl .aflat{color:#9A938C;}
.atbl .anew{background:#0D7680;color:#fff;border-radius:3px;padding:1px 7px;font-size:11px;font-weight:600;}
.atbl-title{font-family:Georgia,'Source Serif Pro',serif;font-weight:700;font-size:15px;
 color:#33302E;margin:4px 0 6px 2px;}
</style>
"""

        def _esc(s):
            return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        def _alloc_panel(bd, title, other_bd=None, show_chg=False):
            items = sorted([(L, w) for L, w in bd.items() if w > 1e-9],
                           key=lambda x: x[1], reverse=True)
            head_chg = '<th class="ar">증감</th>' if show_chg else ''
            body = []
            for i, (L, w) in enumerate(items, 1):
                chg = ''
                if show_chg:
                    ob = other_bd or {}
                    if L not in ob:
                        chg = '<td class="achg"><span class="anew">신규</span></td>'
                    else:
                        d = w - ob.get(L, 0.0)
                        cls = 'apos' if d > 0.005 else ('aneg' if d < -0.005 else 'aflat')
                        chg = f'<td class="achg {cls}">{d:+.2f}%p</td>'
                body.append(f'<tr><td class="ark">{i}</td><td class="anm">{_esc(L)}</td>'
                            f'<td class="awt">{w:.2f}%</td>{chg}</tr>')
            return (f'<div class="atbl-title">{_esc(title)}</div>'
                    f'<table class="atbl"><thead><tr><th>순위</th><th>{_dim_unit}</th>'
                    f'<th class="ar">비중(%)</th>{head_chg}</tr></thead>'
                    f'<tbody>{"".join(body)}</tbody></table>')

        st.markdown(_ALLOC_TBL_CSS, unsafe_allow_html=True)
        _pcol1, _pcol2 = st.columns(2)
        with _pcol1:
            st.markdown(_alloc_panel(_cur_bd, f'현재 ({_cur_reb[0]})',
                                     other_bd=_past_bd, show_chg=True),
                        unsafe_allow_html=True)
        with _pcol2:
            st.markdown(_alloc_panel(_past_bd, f'{alloc_lb} ({_past_reb[0]})'),
                        unsafe_allow_html=True)
        st.caption(
            f':grey[**증감 컬럼** = 좌측(현재) − 우측({alloc_lb}) 비중 변화(%p) — 🔴(+) / 🔵(−) / 신규.  '
            '국가·섹터·팩터 = ETF 룩스루 × 주식 슬리브 100% 정규화(적극·중립 동일), 종목별 = 전체 포트 비중.  '
            '팩터(MSCI Style)는 정적 추정치. **전체 종목 표시**, 각 패널 비중 내림차순.]')

    # ─── Bond Duration + YTM breakdown (unified) ───
    st.markdown('---')
    st.subheader('💵 채권 듀레이션 & YTM 분석')
    st.caption(
        '※ 적극형·중립형은 동일 채권 ETF 구성을 비율만 달리하므로, **채권 슬리브 기준 정규화 시 듀레이션·YTM·버킷 배분은 동일**합니다 (채권 슬리브 총 비중만 분리 표기).  '
        '보유 채권 ETF의 effective duration과 YTM(만기수익률)을 가중평균한 분석.  '
        '듀레이션 버킷: 단기(0~3년) / 중기(3~7년) / 장기(7년+).  '
        'YTM 버킷: 저금리(0~4%) / 중금리(4~6%) / 고금리(6~8%) / 초고금리(8%+).'
    )

    m_agg_bond = compute_bond_metrics(agg_rebals[-1][2]) if len(agg_rebals) > 0 else None
    m_neu_bond = compute_bond_metrics(neu_rebals[-1][2]) if len(neu_rebals) > 0 else None

    # Use whichever is non-None for unified shared metrics (they are identical when normalized)
    m_shared = m_agg_bond if m_agg_bond is not None else m_neu_bond

    if m_shared is None:
        st.warning('채권 보유 없음')
    else:
        # KPI cards — bond sleeve % differs per profile, shared metrics common
        agg_bond_pct = m_agg_bond['bond_total'] if m_agg_bond else 0
        neu_bond_pct = m_neu_bond['bond_total'] if m_neu_bond else 0
        k1, k2, k3, k4 = st.columns(4)
        k1.metric('📈 적극형 채권 슬리브', f'{agg_bond_pct:.2f}%')
        k2.metric('📉 중립형 채권 슬리브', f'{neu_bond_pct:.2f}%')
        k3.metric('가중평균 듀레이션 (공통)', f'{m_shared["wavg_dur"]:.2f}년')
        k4.metric('가중평균 YTM (공통)',     f'{m_shared["wavg_ytm"]:.2f}%')

        # Duration + YTM pie charts side by side
        pc1, pc2 = st.columns(2)
        with pc1:
            st.markdown('**📊 듀레이션 버킷 배분**')
            dur_labels = list(m_shared['dur_buckets'].keys())
            dur_vals   = list(m_shared['dur_buckets'].values())
            fig_dur = go.Figure(go.Pie(
                labels=dur_labels, values=dur_vals,
                marker_colors=['#10B981', '#F59E0B', '#DC2626'],
                textinfo='label+percent',
                textfont_size=11, hole=0.4,
            ))
            fig_dur.update_layout(height=300, margin=dict(l=10, r=10, t=10, b=10),
                                  showlegend=False)
            st.plotly_chart(fig_dur, use_container_width=True, key='bond_dur_unified')
        with pc2:
            st.markdown('**📊 YTM 버킷 배분**')
            ytm_labels = list(m_shared['ytm_buckets'].keys())
            ytm_vals   = list(m_shared['ytm_buckets'].values())
            fig_ytm = go.Figure(go.Pie(
                labels=ytm_labels, values=ytm_vals,
                marker_colors=['#3B82F6', '#10B981', '#F59E0B', '#DC2626'],
                textinfo='label+percent',
                textfont_size=11, hole=0.4,
            ))
            fig_ytm.update_layout(height=300, margin=dict(l=10, r=10, t=10, b=10),
                                  showlegend=False)
            st.plotly_chart(fig_ytm, use_container_width=True, key='bond_ytm_unified')

        # Unified bond detail table — show normalized 채권 슬리브 비중 + 양 프로필 절대 비중 참고
        st.markdown('**📋 보유 채권 ETF 상세 (채권 슬리브 정규화 + 프로필별 절대 비중)**')
        rows = []
        agg_details_map = {d['name']: d for d in m_agg_bond['details']} if m_agg_bond else {}
        neu_details_map = {d['name']: d for d in m_neu_bond['details']} if m_neu_bond else {}
        all_names = set(agg_details_map) | set(neu_details_map)
        for name in all_names:
            d = agg_details_map.get(name) or neu_details_map.get(name)
            agg_abs = agg_details_map[name]['weight'] if name in agg_details_map else 0
            neu_abs = neu_details_map[name]['weight'] if name in neu_details_map else 0
            sleeve_pct = (d['weight'] / m_shared['bond_total'] * 100) if m_shared['bond_total'] > 0 else 0
            rows.append({
                '채권 ETF':         d['name'],
                '채권 슬리브 비중(%)': round(sleeve_pct, 2),
                '적극형 절대(%)':    round(agg_abs, 2),
                '중립형 절대(%)':    round(neu_abs, 2),
                '듀레이션(년)':      round(d['duration'], 2),
                'YTM(%)':           round(d['ytm'], 2),
            })
        df_bond = pd.DataFrame(rows).sort_values('채권 슬리브 비중(%)', ascending=False).reset_index(drop=True)
        styled_bond = df_bond.style.background_gradient(
            subset=['채권 슬리브 비중(%)'], cmap='Greys', vmin=0, vmax=df_bond['채권 슬리브 비중(%)'].max() or 1,
        ).background_gradient(
            subset=['듀레이션(년)'], cmap='YlOrRd', vmin=0, vmax=10,
        ).background_gradient(
            subset=['YTM(%)'], cmap='YlGn', vmin=0, vmax=10,
        ).format({
            '채권 슬리브 비중(%)': '{:.2f}',
            '적극형 절대(%)':     '{:.2f}',
            '중립형 절대(%)':     '{:.2f}',
            '듀레이션(년)':       '{:.2f}',
            'YTM(%)':            '{:.2f}',
        })
        st.dataframe(styled_bond, use_container_width=True, hide_index=True)
        st.caption(
            ':grey[**📊 데이터 출처:** 듀레이션·YTM은 각 ETF의 최신 fact sheet 대표값 (정적 메타데이터).  '
            '**채권 슬리브 비중(%)** = 각 ETF의 채권 슬리브 내 정규화 비중 (공통).  '
            '**적극형/중립형 절대(%)** = 전체 펀드 NAV 대비 절대 비중 (프로필별 차이).  '
            'Phase 2에서 Refinitiv/Morningstar API 실시간 연동 권고.]'
        )

    # Holdings Performance
    st.markdown('---')
    st.subheader('📋 최신 종목별 성과 및 기여도')

    agg_perf_df, agg_rebal_date, agg_today = compute_holdings_perf(agg_rebals, px, end_date_str)
    neu_perf_df, neu_rebal_date, neu_today = compute_holdings_perf(neu_rebals, px, end_date_str)

    st.caption(
        f'기준 가중치: 최근 리밸런싱 ({agg_rebal_date})  |  '
        f'가격 기준일: {agg_today}  |  '
        f'※ Total Return 기준 (배당 재투자 반영, auto-adjusted) | '
        f'기여(bps) = 비중 × 수익률'
    )

    def format_perf_df(df):
        if df.empty:
            return df
        styled = df.style.background_gradient(
            subset=['1D Return(%)', '1W Return(%)', 'MTD Return(%)'],
            cmap='RdYlGn', vmin=-5, vmax=5,
        ).background_gradient(
            subset=['1D 기여(bps)', '1W 기여(bps)', 'MTD 기여(bps)'],
            cmap='RdYlGn', vmin=-50, vmax=50,
        ).format({
            '비중(%)': '{:.2f}',
            '1D Return(%)':  '{:+.2f}',
            '1W Return(%)':  '{:+.2f}',
            'MTD Return(%)': '{:+.2f}',
            '1D 기여(bps)':  '{:+.1f}',
            '1W 기여(bps)':  '{:+.1f}',
            'MTD 기여(bps)': '{:+.1f}',
        })
        return styled

    def render_comparison(rebals, profile_name):
        comp = compute_comparison(rebals, px, end_date_str)
        if not comp:
            st.info('전월 리밸런스 데이터 부족 — 비교 불가')
            return
        st.markdown(
            f'**현재 포트폴리오:** {comp["cur_rebal"]} ({comp["cur_signal"]})  |  '
            f'**전월 포트폴리오:** {comp["prev_rebal"]} ({comp["prev_signal"]})'
        )
        st.caption(
            f'※ 현재 MTD 기간: {comp["cur_period"]}  |  '
            f'전월 보유 기간(직전 리밸 → 현재 리밸): {comp["prev_period"]}'
        )
        comp_df = comp['df']
        # Header KPIs
        delta_w = (comp_df['현재 비중(%)'] - comp_df['전월 비중(%)']).abs().sum() / 2
        sum_cur_w  = comp_df['MTD 기여(bps, 현재비중)'].sum()
        sum_prev_w = comp_df['MTD 기여(bps, 전월비중)'].sum()
        total_rebal_impact = sum_cur_w - sum_prev_w
        kc1, kc2, kc3, kc4 = st.columns(4)
        kc1.metric('MTD 기여 합 (현재비중)',  f'{sum_cur_w:+.1f} bps')
        kc2.metric('MTD 기여 합 (전월비중)',  f'{sum_prev_w:+.1f} bps')
        kc3.metric('총 리밸 효과', f'{total_rebal_impact:+.1f} bps',
                   delta=('비중 조절 성공' if total_rebal_impact >= 0 else '비중 조절 실패'))
        kc4.metric('비중 turnover (½·∑|Δw|)', f'{delta_w:.2f}%p')

        st.caption(
            '※ **MTD(bps)** = 종목의 MTD total return (% × 100, 비중 무관)  |  '
            '**MTD 기여(bps, 현재비중)** = 현재 비중 × 현재 MTD return  |  '
            '**MTD 기여(bps, 전월비중)** = 전월 비중 × 현재 MTD return (counter-factual)  |  '
            '**리밸런싱 효과(bps)** = (현재 - 전월) × 현재 MTD return → 양수면 비중 조절 성공, 음수면 실패.'
        )

        styled = comp_df.style.background_gradient(
            subset=['비중 변화(%p)'], cmap='RdBu_r', vmin=-10, vmax=10,
        ).background_gradient(
            subset=['MTD(bps)'], cmap='RdYlGn', vmin=-500, vmax=500,
        ).background_gradient(
            subset=['MTD 기여(bps, 현재비중)', 'MTD 기여(bps, 전월비중)', '리밸런싱 효과(bps)'],
            cmap='RdYlGn', vmin=-50, vmax=50,
        ).format({
            '전월 비중(%)':              '{:.2f}',
            '현재 비중(%)':              '{:.2f}',
            '비중 변화(%p)':             '{:+.2f}',
            'MTD(bps)':                  '{:+.1f}',
            'MTD 기여(bps, 현재비중)':   '{:+.1f}',
            'MTD 기여(bps, 전월비중)':   '{:+.1f}',
            '리밸런싱 효과(bps)':         '{:+.1f}',
        })
        st.dataframe(styled, use_container_width=True, hide_index=True, height=600)

        # Data source / measurement period footnote
        st.caption(
            ':grey[**📊 데이터 출처:** Yahoo Finance — auto-adjusted (배당 재투자 반영, Total Return 기준).  '
            f'**🗓️ 성과 측정 기간 (Report Date {end_date_str} 기준):** MTD = 이번 달 1영업일 → 오늘.  '
            f'**현재 vs 전월 비교:** 현재 = {comp["cur_rebal"]} 리밸 weights / 전월 = {comp["prev_rebal"]} 리밸 weights, 동일한 현재 MTD return을 곱해 비중 효과만 분리.  '
            '**KR 채권·현금 (TIGER CD금리·국채·종합채권, KRW Cash)**은 가격 데이터 부재로 proxy 상수 수익률 적용.]'
        )

    def render_holdings(df, rebals, profile_name):
        if df.empty:
            st.warning(f'{profile_name}: 데이터 없음')
            return
        total_w = df['비중(%)'].sum()
        total_1d = df['1D 기여(bps)'].sum()
        total_1w = df['1W 기여(bps)'].sum()
        total_mtd = df['MTD 기여(bps)'].sum()

        c1, c2, c3, c4 = st.columns(4)
        c1.metric(f'{profile_name} 합계 비중', f'{total_w:.1f}%')
        c2.metric('포트폴리오 1D 기여', f'{total_1d:+.1f} bps')
        c3.metric('포트폴리오 1W 기여', f'{total_1w:+.1f} bps')
        c4.metric('포트폴리오 MTD 기여', f'{total_mtd:+.1f} bps')

        # ─── 구분별 성과 breakdown ───
        st.markdown('##### 📂 구분별 성과 (광역주식 / 국가주식 / 채권 / 현금)')
        cat_breakdown = []
        cat_breakdown.append({
            '구분': '🟦 포트폴리오 합계',
            '비중(%)': round(total_w, 2),
            '1D 기여(bps)':  round(total_1d, 1),
            '1W 기여(bps)':  round(total_1w, 1),
            'MTD 기여(bps)': round(total_mtd, 1),
        })
        for cat_label, cat_key in [('🌐 광역 주식', '광역 주식'),
                                   ('🏳️ 국가 주식', '국가 주식'),
                                   ('💵 채권',      '채권'),
                                   ('💴 현금',      '현금')]:
            sub = df[df['구분'] == cat_key]
            cat_breakdown.append({
                '구분': cat_label,
                '비중(%)':       round(sub['비중(%)'].sum(),       2) if not sub.empty else 0.0,
                '1D 기여(bps)':  round(sub['1D 기여(bps)'].sum(),  1) if not sub.empty else 0.0,
                '1W 기여(bps)':  round(sub['1W 기여(bps)'].sum(),  1) if not sub.empty else 0.0,
                'MTD 기여(bps)': round(sub['MTD 기여(bps)'].sum(), 1) if not sub.empty else 0.0,
            })
        cat_df = pd.DataFrame(cat_breakdown)

        cat_styled = cat_df.style.background_gradient(
            subset=['1D 기여(bps)', '1W 기여(bps)', 'MTD 기여(bps)'],
            cmap='RdYlGn', vmin=-100, vmax=200,
        ).format({
            '비중(%)':       '{:.2f}',
            '1D 기여(bps)':  '{:+.1f}',
            '1W 기여(bps)':  '{:+.1f}',
            'MTD 기여(bps)': '{:+.1f}',
        })
        st.dataframe(cat_styled, use_container_width=True, hide_index=True)

        with st.expander(f'📊 {profile_name} 전체 종목 상세 테이블 (열 헤더 클릭하여 정렬)'):
            sort_by = st.selectbox(
                '정렬 기준',
                ['구분 (광역주식 → 국가주식 → 채권 → 현금)',
                 '비중(%)', 'MTD 기여(bps)', '1W 기여(bps)', '1D 기여(bps)',
                 'MTD Return(%)', '1W Return(%)', '1D Return(%)'],
                key=f'sort_{profile_name}',
            )
            sort_asc = st.checkbox('오름차순', key=f'asc_{profile_name}', value=False)
            if sort_by.startswith('구분'):
                CAT_ORDER = ['광역 주식', '국가 주식', '원자재', '헤지펀드', '채권', '현금', '기타']
                df_sorted = df.copy()
                df_sorted['_cat'] = df_sorted['구분'].map({c: i for i, c in enumerate(CAT_ORDER)}).fillna(99)
                df_sorted = (df_sorted
                             .sort_values(['_cat', '비중(%)'], ascending=[True, False])
                             .drop(columns='_cat')
                             .reset_index(drop=True))
            else:
                df_sorted = df.sort_values(sort_by, ascending=sort_asc).reset_index(drop=True)
            st.dataframe(format_perf_df(df_sorted), use_container_width=True, hide_index=True, height=600)
            st.caption(
                ':grey[**📊 데이터 출처:** Yahoo Finance — auto-adjusted (배당 재투자 반영, Total Return 기준).  '
                f'**🗓️ 성과 측정 기간 (Report Date {end_date_str} 기준):**  '
                '**1D Return** = 직전 영업일 종가 → 오늘 종가.  '
                '**1W Return** = T-7영업일 종가 → 오늘 종가.  '
                '**MTD Return** = 이번 달 1영업일 → 오늘.  '
                '**기여(bps)** = 비중 × 수익률 (=종목별 portfolio 기여도).  '
                '**KR 채권·현금**은 가격 데이터 부재로 proxy 상수 수익률 적용 (CD/국채 연 3.5~4%, KRW Cash 0%, USD Cash는 USDKRW 환율 변동).]'
            )

        # ─── 현재 vs 전월 포트폴리오 비교 ───
        st.markdown('---')
        st.subheader(f'🔄 {profile_name} 현재 vs 전월 포트폴리오 비교 (비중 + MTD 기여)')
        render_comparison(rebals, profile_name)

    # ─── 🔬 심화 분석 표시 토글 ───
    st.markdown('---')
    _toggle_col1, _toggle_col2 = st.columns([4, 1])
    with _toggle_col1:
        st.markdown('### 🔬 심화 분석 섹션 (리밸 효과 추이 · 자산 배분 효과)')
    with _toggle_col2:
        show_deep_analytics = st.toggle(
            '표시',
            value=st.session_state.get('show_deep_analytics', True),
            key='show_deep_analytics',
            help='ON: 리밸런싱 효과 추이 + 자산 배분 효과 심층 분석 표시.  '
                 'OFF: 위 3개 섹션 숨김 (페이지 가벼움).',
        )
    if not show_deep_analytics:
        st.caption(':grey[💡 심화 분석 섹션이 **숨김 처리**되었습니다. 위 토글로 다시 표시.]')

    tab_agg, tab_neu = st.tabs(['📈 적극형', '📉 중립형'])
    with tab_agg:
        render_holdings(agg_perf_df, agg_rebals, '적극형')
    with tab_neu:
        render_holdings(neu_perf_df, neu_rebals, '중립형')

    if st.session_state.get('show_deep_analytics', True):
        # ─── 리밸런싱 효과 추이 (운용시작일 → 조회일) ───
        st.markdown('---')
        st.subheader('📈 리밸런싱 효과 추이 (운용시작일 → 조회일)')
        st.caption(
            '각 리밸런스 시점의 비중 조절 결정이 다음 리밸런스(또는 조회일)까지 발휘한 효과(bps)를 시계열로 표시.  '
            '계산식: ∑_t (현재비중 - 전월비중) × 보유기간 종목 return.  '
            '막대 = 해당 리밸 시점의 period 효과 / 선 = 운용시작 이후 누적 효과.'
        )

        agg_hist = compute_rebal_effect_history(agg_rebals, px, end_date_str)
        neu_hist = compute_rebal_effect_history(neu_rebals, px, end_date_str)

        if (agg_hist is None or agg_hist.empty) and (neu_hist is None or neu_hist.empty):
            st.info('리밸런스 히스토리 부족 — 추이 산출 불가')
        else:
            SIGNAL_COLORS = {'Bull': '#10B981', 'Base': '#F59E0B', 'Bear': '#DC2626'}
            SIGNAL_EMOJI  = {'Bull': '🟢', 'Base': '🟡', 'Bear': '🔴'}

            # ─── 🔍 뷰 모드 선택 ───
            view_mode = st.radio(
                '🔍 뷰 모드',
                ['📈 누적 시계열 (시그널 필터)',
                 '📅 달력 히트맵',
                 '🏆 시그널 성적표',
                 '🎨 시계열 막대 (시그널 색상)'],
                horizontal=True,
                key='rebal_view_mode',
                help='월 독립성을 존중하려면 달력/성적표/색상막대 권장. '
                     '시계열 누적은 시점간 의존성이 시각적으로 강조됨.',
            )

            # ─── 📅 달력 히트맵 ───
            if view_mode == '📅 달력 히트맵':
                st.caption(
                    ':grey[**각 셀** = 해당 월의 리밸 효과 (bps). '
                    '**색상** = 양수 초록 / 음수 빨강. **이모지** = 그 달의 시그널. '
                    '월별 완전 독립 — 시점간 누적 관계 시각화 안 함.]'
                )

                def _build_calendar_heatmap(hist, profile_name, key_suffix):
                    if hist is None or hist.empty:
                        return
                    df = hist.copy()
                    df['year'] = pd.to_datetime(df['date']).dt.year
                    df['month'] = pd.to_datetime(df['date']).dt.month
                    years = sorted(df['year'].unique())
                    months = list(range(1, 13))
                    z, text = [], []
                    for y in years:
                        z_row, t_row = [], []
                        for m in months:
                            sub = df[(df['year'] == y) & (df['month'] == m)]
                            if not sub.empty:
                                bps = sub['impact_bps'].iloc[0]
                                sig = sub['signal'].iloc[0]
                                z_row.append(bps)
                                t_row.append(f'{SIGNAL_EMOJI.get(sig, "")}<br>{bps:+.0f}')
                            else:
                                z_row.append(None)
                                t_row.append('')
                        z.append(z_row); text.append(t_row)
                    # zmax/zmin: 데이터 절댓값 기준 대칭
                    vals = [v for row in z for v in row if v is not None]
                    vrange = max(abs(min(vals)), abs(max(vals))) if vals else 200
                    fig = go.Figure(data=go.Heatmap(
                        z=z, x=[f'{m}월' for m in months], y=[str(y) for y in years],
                        text=text, texttemplate='%{text}', textfont={'size': 11},
                        colorscale='RdYlGn', zmid=0, zmin=-vrange, zmax=vrange,
                        hovertemplate='%{y}년 %{x}<br>%{z:+.1f} bps<extra></extra>',
                        colorbar=dict(title='bps', thickness=12),
                    ))
                    fig.update_layout(
                        height=max(220, len(years) * 65 + 100),
                        xaxis=dict(side='top', tickfont=dict(size=11)),
                        yaxis=dict(autorange='reversed', tickfont=dict(size=11)),
                        margin=dict(l=40, r=20, t=60, b=20),
                        title=dict(text=f'<b>{profile_name}</b>', x=0.02,
                                    font=dict(size=13)),
                    )
                    st.plotly_chart(fig, use_container_width=True,
                                    key=f'cal_heatmap_{key_suffix}')

                cal_tabs = st.tabs(['📈 적극형', '📉 중립형'])
                with cal_tabs[0]:
                    if agg_hist is not None and not agg_hist.empty:
                        _build_calendar_heatmap(agg_hist, '적극형', 'agg')
                    else:
                        st.info('적극형 데이터 없음')
                with cal_tabs[1]:
                    if neu_hist is not None and not neu_hist.empty:
                        _build_calendar_heatmap(neu_hist, '중립형', 'neu')
                    else:
                        st.info('중립형 데이터 없음')

            # ─── 🏆 시그널 성적표 ───
            elif view_mode == '🏆 시그널 성적표':
                st.caption(
                    ':grey[**시점 무관 단순 통계** — 각 리밸을 독립 sample로 취급해 시그널별 평균·적중률·최고·최저를 표시.  '
                    '시계열 의존성 없음 → 월 독립성 완전 존중.]'
                )

                def _scorecard(hist, profile_name):
                    if hist is None or hist.empty:
                        st.info(f'{profile_name} 데이터 없음')
                        return
                    st.markdown(f'#### {profile_name} 시그널 성적표')
                    cols = st.columns(3)
                    for col, sig in zip(cols, ['Bull', 'Base', 'Bear']):
                        sub = hist[hist['signal'] == sig]
                        n = len(sub)
                        emoji = SIGNAL_EMOJI[sig]
                        with col:
                            if n == 0:
                                st.markdown(f'### {emoji} {sig}')
                                st.markdown('**0회 발생**')
                                st.caption('해당 시그널 발효 이력 없음')
                                continue
                            avg = sub['impact_bps'].mean()
                            median = sub['impact_bps'].median()
                            win_rate = (sub['impact_bps'] > 0).sum() / n * 100
                            max_v = sub['impact_bps'].max()
                            min_v = sub['impact_bps'].min()
                            std_v = sub['impact_bps'].std() if n > 1 else 0
                            st.markdown(f'### {emoji} {sig}')
                            st.markdown(f'**{n}회 발생**')
                            st.metric('평균 효과', f'{avg:+.1f} bps')
                            st.metric('적중률', f'{win_rate:.1f}%')
                            c1, c2 = st.columns(2)
                            c1.metric('최고', f'{max_v:+.0f}')
                            c2.metric('최저', f'{min_v:+.0f}')
                            st.caption(f'중간값 {median:+.1f} / 변동성 {std_v:.1f}')

                _scorecard(agg_hist, '적극형')
                st.markdown('---')
                _scorecard(neu_hist, '중립형')

            # ─── 🎨 시계열 막대 (시그널 색상) ───
            elif view_mode == '🎨 시계열 막대 (시그널 색상)':
                st.caption(
                    ':grey[**누적선 제거** — 막대만으로 시점별 효과 표시. '
                    '막대 색상 = 시그널 (🟢 Bull / 🟡 Base / 🔴 Bear).  '
                    '시간 순서는 유지하되 누적 의존성을 시각적으로 강조하지 않음.]'
                )

                def _build_colored_bars(hist, profile_name, key_suffix):
                    if hist is None or hist.empty:
                        return
                    fig = go.Figure()
                    for sig in ['Bull', 'Base', 'Bear']:
                        sub = hist[hist['signal'] == sig]
                        if sub.empty:
                            continue
                        fig.add_trace(go.Bar(
                            x=sub['date'], y=sub['impact_bps'],
                            name=f'{SIGNAL_EMOJI[sig]} {sig} (n={len(sub)})',
                            marker_color=SIGNAL_COLORS[sig],
                            hovertemplate=('%{x|%Y-%m-%d}<br>' + sig +
                                           ': %{y:+.1f} bps<extra></extra>'),
                        ))
                    fig.add_hline(y=0, line_dash='dot', line_color='gray', line_width=1)
                    fig.update_layout(
                        title=dict(text=f'<b>{profile_name}</b>', x=0.02,
                                    font=dict(size=13)),
                        height=380, hovermode='x unified',
                        xaxis_title='리밸런스 일자', yaxis_title='period 효과 (bps)',
                        legend=dict(orientation='h', yanchor='bottom', y=1.02,
                                    xanchor='right', x=1, font=dict(size=10)),
                        margin=dict(t=50, b=40, l=40, r=20),
                    )
                    st.plotly_chart(fig, use_container_width=True,
                                    key=f'colored_bars_{key_suffix}')

                bar_tabs = st.tabs(['📈 적극형', '📉 중립형'])
                with bar_tabs[0]:
                    _build_colored_bars(agg_hist, '적극형', 'agg')
                with bar_tabs[1]:
                    _build_colored_bars(neu_hist, '중립형', 'neu')

            # ─── 📈 누적 시계열 (기존 코드) ───
            elif view_mode == '📈 누적 시계열 (시그널 필터)':
                # ─── 🎯 시그널 필터 ───
                signal_filter = st.radio(
                    '🎯 시그널 필터',
                    ['전체', 'Bull', 'Base', 'Bear', '📊 시그널 비교'],
                    horizontal=True,
                    key='rebal_signal_filter',
                    help='특정 시그널 구간만 분리해 누적 기여 흐름 확인. 비교 모드는 3 시그널 누적선 동시 표시.',
                )

                def _filter_recompute(hist, signal):
                    """signal=='all'이면 그대로, 아니면 해당 시그널만 필터링 후 누적 재계산."""
                    if hist is None or hist.empty:
                        return hist
                    if signal == 'all':
                        return hist
                    filtered = hist[hist['signal'] == signal].copy().sort_values('date').reset_index(drop=True)
                    if not filtered.empty:
                        filtered['cumulative_bps'] = filtered['impact_bps'].cumsum()
                    return filtered

                if signal_filter == '📊 시그널 비교':
                    # ─── 시그널 비교 모드: 3 시그널 누적선 × 2 profile ───
                    st.caption(
                        ':grey[**비교 모드:** 실선=적극형 / 점선=중립형. 색상 = 시그널(Bull 초록 / Base 앰버 / Bear 빨강).  '
                        '각 라인은 해당 시그널 이벤트만의 누적 기여 — Bull 라인이 우상향이면 50% 스왑 룰이 알파 소스로 작동.]'
                    )
                    fig_rebal = go.Figure()
                    for profile_name, hist, dash_style in [
                        ('적극형', agg_hist, 'solid'),
                        ('중립형', neu_hist, 'dot'),
                    ]:
                        if hist is None or hist.empty:
                            continue
                        for signal in ['Bull', 'Base', 'Bear']:
                            fh = _filter_recompute(hist, signal)
                            if fh is None or fh.empty:
                                continue
                            fig_rebal.add_trace(go.Scatter(
                                x=fh['date'], y=fh['cumulative_bps'],
                                name=f'{profile_name} · {signal} (n={len(fh)})',
                                mode='lines+markers',
                                line=dict(color=SIGNAL_COLORS[signal], width=2.4, dash=dash_style),
                                marker=dict(size=5),
                                hovertemplate=('%{x|%Y-%m-%d}<br>' +
                                               f'{profile_name} {signal}: ' + '%{y:+.1f} bps<extra></extra>'),
                            ))
                    fig_rebal.add_hline(y=0, line_dash='dot', line_color='gray', line_width=1)
                    fig_rebal.update_layout(
                        height=520, hovermode='x unified',
                        xaxis_title='리밸런스 일자',
                        yaxis_title='시그널별 누적 리밸 효과 (bps)',
                        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1,
                                    font=dict(size=9)),
                        margin=dict(t=50, b=40, l=40, r=20),
                    )
                    st.plotly_chart(fig_rebal, use_container_width=True,
                                    key='rebal_effect_trend_compare')
                else:
                    # ─── Single mode (전체/Bull/Base/Bear) ───
                    sig_key = 'all' if signal_filter == '전체' else signal_filter
                    agg_f = _filter_recompute(agg_hist, sig_key)
                    neu_f = _filter_recompute(neu_hist, sig_key)

                    # Bear small sample 경고
                    if sig_key == 'Bear':
                        a_n = 0 if agg_f is None else len(agg_f)
                        n_n = 0 if neu_f is None else len(neu_f)
                        st.warning(
                            f'⚠ Bear 시그널 sample size 매우 작음 (적극 n={a_n}, 중립 n={n_n}) — '
                            f'추이의 통계적 신뢰도 제한적. 참고용으로만 활용 권장.'
                        )

                    bar_alpha = 0.55 if sig_key != 'all' else 0.45

                    fig_rebal = make_subplots(specs=[[{'secondary_y': True}]])
                    if agg_f is not None and not agg_f.empty:
                        bar_color_agg = (f'rgba(16,185,129,{bar_alpha})'  if sig_key == 'Bull'
                                         else f'rgba(245,158,11,{bar_alpha})'  if sig_key == 'Base'
                                         else f'rgba(220,38,38,{bar_alpha})'   if sig_key == 'Bear'
                                         else 'rgba(31,58,104,0.45)')
                        fig_rebal.add_trace(go.Bar(
                            x=agg_f['date'], y=agg_f['impact_bps'],
                            name='적극형 period', marker_color=bar_color_agg,
                            hovertemplate='%{x|%Y-%m-%d}<br>적극형 period: %{y:+.1f} bps<extra></extra>',
                        ), secondary_y=False)
                        fig_rebal.add_trace(go.Scatter(
                            x=agg_f['date'], y=agg_f['cumulative_bps'],
                            name='적극형 누적', mode='lines+markers',
                            line=dict(color='#1F3A68', width=2.6),
                            marker=dict(size=6),
                            hovertemplate='%{x|%Y-%m-%d}<br>적극형 누적: %{y:+.1f} bps<extra></extra>',
                        ), secondary_y=True)
                    if neu_f is not None and not neu_f.empty:
                        bar_color_neu = (f'rgba(16,185,129,{bar_alpha*0.8})'  if sig_key == 'Bull'
                                         else f'rgba(245,158,11,{bar_alpha*0.8})'  if sig_key == 'Base'
                                         else f'rgba(220,38,38,{bar_alpha*0.8})'   if sig_key == 'Bear'
                                         else 'rgba(196,141,67,0.45)')
                        fig_rebal.add_trace(go.Bar(
                            x=neu_f['date'], y=neu_f['impact_bps'],
                            name='중립형 period', marker_color=bar_color_neu,
                            hovertemplate='%{x|%Y-%m-%d}<br>중립형 period: %{y:+.1f} bps<extra></extra>',
                        ), secondary_y=False)
                        fig_rebal.add_trace(go.Scatter(
                            x=neu_f['date'], y=neu_f['cumulative_bps'],
                            name='중립형 누적', mode='lines+markers',
                            line=dict(color='#C48D43', width=2.6),
                            marker=dict(size=6),
                            hovertemplate='%{x|%Y-%m-%d}<br>중립형 누적: %{y:+.1f} bps<extra></extra>',
                        ), secondary_y=True)
                    fig_rebal.add_hline(y=0, line_dash='dot', line_color='gray', line_width=1)
                    title_suffix = '' if sig_key == 'all' else f' · {signal_filter} 시그널만'
                    fig_rebal.update_layout(
                        barmode='group', height=500, hovermode='x unified',
                        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                        margin=dict(t=50, b=40, l=40, r=20),
                        xaxis_title='리밸런스 일자' + title_suffix,
                    )
                    fig_rebal.update_yaxes(title_text='Period 리밸 효과 (bps)', secondary_y=False)
                    fig_rebal.update_yaxes(title_text='누적 리밸 효과 (bps)', secondary_y=True)
                    st.plotly_chart(fig_rebal, use_container_width=True,
                                    key=f'rebal_effect_trend_{sig_key}')

                # ─── Summary KPIs (시그널 필터 반영) ───
                def _stats(hist):
                    if hist is None or hist.empty:
                        return 0.0, 0, 0, 0.0
                    total = hist['impact_bps'].sum()
                    n = len(hist)
                    pos = int((hist['impact_bps'] > 0).sum())
                    avg = hist['impact_bps'].mean()
                    return total, n, pos, avg

                if signal_filter == '📊 시그널 비교':
                    sig_rows = []
                    for profile_name, hist in [('적극형', agg_hist), ('중립형', neu_hist)]:
                        if hist is None or hist.empty:
                            continue
                        for signal in ['Bull', 'Base', 'Bear']:
                            fh = _filter_recompute(hist, signal)
                            total, n, pos, avg = _stats(fh)
                            sig_rows.append({
                                'Profile': profile_name,
                                '시그널': signal,
                                '이벤트(회)': n,
                                '누적 효과(bps)': round(total, 1),
                                '평균/리밸(bps)': round(avg, 1) if n else 0,
                                '양(+) 적중': pos,
                                '적중률(%)': round(pos / n * 100, 1) if n else 0,
                            })
                    if sig_rows:
                        sig_df = pd.DataFrame(sig_rows)
                        sig_styled = sig_df.style.background_gradient(
                            subset=['누적 효과(bps)', '평균/리밸(bps)'], cmap='RdYlGn', vmin=-300, vmax=300,
                        ).background_gradient(
                            subset=['적중률(%)'], cmap='RdYlGn', vmin=30, vmax=70,
                        ).format({
                            '누적 효과(bps)': '{:+.1f}',
                            '평균/리밸(bps)': '{:+.1f}',
                            '적중률(%)':       '{:.1f}',
                        })
                        st.dataframe(sig_styled, use_container_width=True, hide_index=True)
                else:
                    sig_key_f = 'all' if signal_filter == '전체' else signal_filter
                    agg_total, agg_n, agg_pos, agg_avg = _stats(_filter_recompute(agg_hist, sig_key_f))
                    neu_total, neu_n, neu_pos, neu_avg = _stats(_filter_recompute(neu_hist, sig_key_f))
                    rc1, rc2, rc3, rc4 = st.columns(4)
                    sig_lbl = '전체' if sig_key_f == 'all' else sig_key_f
                    rc1.metric(f'적극형 {sig_lbl} 누적', f'{agg_total:+.1f} bps',
                               delta=f'평균 {agg_avg:+.1f} bps/리밸' if agg_n else None)
                    rc2.metric(f'적극형 {sig_lbl} 양(+) 적중',
                               f'{agg_pos}/{agg_n}회',
                               delta=f'{(agg_pos/agg_n*100):.1f}%' if agg_n else None)
                    rc3.metric(f'중립형 {sig_lbl} 누적', f'{neu_total:+.1f} bps',
                               delta=f'평균 {neu_avg:+.1f} bps/리밸' if neu_n else None)
                    rc4.metric(f'중립형 {sig_lbl} 양(+) 적중',
                               f'{neu_pos}/{neu_n}회',
                               delta=f'{(neu_pos/neu_n*100):.1f}%' if neu_n else None)

            # Detail table (expander) — 정확한 리밸 일자 기준으로 outer-join (카르테시안 곱 방지)
            with st.expander('📊 리밸 시점별 효과 상세 (적극 vs 중립, 리밸 일자 매칭)'):
                def _prep(hist, prefix):
                    if hist is None or hist.empty:
                        return None
                    d = hist[['date', 'signal', 'impact_bps', 'cumulative_bps']].copy()
                    d['date'] = pd.to_datetime(d['date']).dt.strftime('%Y-%m-%d')
                    d = d.rename(columns={
                        'signal':         f'{prefix}_signal',
                        'impact_bps':     f'{prefix} period(bps)',
                        'cumulative_bps': f'{prefix} 누적(bps)',
                    })
                    return d

                a_df = _prep(agg_hist, '적극형')
                n_df = _prep(neu_hist, '중립형')

                if a_df is not None and n_df is not None:
                    merged = a_df.merge(n_df, on='date', how='outer').sort_values('date').reset_index(drop=True)
                elif a_df is not None:
                    merged = a_df
                else:
                    merged = n_df

                if merged is not None and not merged.empty:
                    # 시그널 통합 (동일하면 단일, 다르면 '적극 / 중립', 한쪽만이면 명시)
                    sig_a = merged.get('적극형_signal', pd.Series([None]*len(merged)))
                    sig_n = merged.get('중립형_signal', pd.Series([None]*len(merged)))
                    def _sig_row(a, n):
                        a_ok = pd.notna(a); n_ok = pd.notna(n)
                        if a_ok and n_ok: return a if a == n else f'{a} / {n}'
                        if a_ok: return f'{a} (적극만)'
                        if n_ok: return f'{n} (중립만)'
                        return '—'
                    merged['시그널'] = [_sig_row(a, n) for a, n in zip(sig_a, sig_n)]
                    merged = merged.rename(columns={'date': '리밸 일자'})

                    # 출력 컬럼 순서 정리
                    preferred = ['리밸 일자', '시그널',
                                 '적극형 period(bps)', '적극형 누적(bps)',
                                 '중립형 period(bps)', '중립형 누적(bps)']
                    cols = [c for c in preferred if c in merged.columns]
                    display = merged[cols].reset_index(drop=True)

                    num_cols = [c for c in display.columns if 'bps' in c]
                    styled = display.style.background_gradient(
                        subset=num_cols, cmap='RdYlGn', vmin=-300, vmax=300,
                    ).format({c: '{:+.1f}' for c in num_cols}, na_rep='—')
                    st.dataframe(styled, use_container_width=True, hide_index=True, height=420)

                    # 비대칭 리밸 진단 (한쪽 profile에만 발생한 리밸 이벤트 수)
                    only_agg = int(merged[('중립형 period(bps)' if '중립형 period(bps)' in merged.columns else '리밸 일자')].isna().sum()) \
                               if '중립형 period(bps)' in merged.columns else 0
                    only_neu = int(merged[('적극형 period(bps)' if '적극형 period(bps)' in merged.columns else '리밸 일자')].isna().sum()) \
                               if '적극형 period(bps)' in merged.columns else 0

                    st.caption(
                        f':grey[※ **리밸 일자(exact date)** 기준 outer-join. '
                        f'한쪽 profile에만 발생한 리밸 이벤트는 다른 쪽이 "—"로 표기.  '
                        f'**적극만 발생:** {only_agg}회 (예: Bear emergency rebal 등) / '
                        f'**중립만 발생:** {only_neu}회.  '
                        f'동일 월에 두 profile의 리밸 일자가 다르면 별도 row로 표시 (카르테시안 곱 방지).]'
                    )

            start_d = agg_rebals[0][0] if agg_rebals else (neu_rebals[0][0] if neu_rebals else '—')
            st.caption(
                ':grey[**📊 데이터 출처:** Yahoo Finance — auto-adjusted (배당 재투자 반영, Total Return 기준).  '
                f'**🗓️ 측정 구간:** {start_d} (운용시작) → {end_date_str} (조회일).  '
                '**Period 효과:** 해당 리밸 시점의 비중 조절이 다음 리밸까지(마지막 리밸은 조회일까지) 발휘한 영향.  '
                '**누적:** 운용시작 이후 모든 리밸 효과의 합 — 비중 조절 의사결정의 누적 알파 기여.  '
                '**KR 채권·현금**은 proxy 상수 수익률 적용.]'
            )

        # ============================================================
        # 🔬 자산 배분 효과 심층 분석 (3개 분석)
        # ============================================================
        def _wilcoxon_p(arr):
            """Wilcoxon signed-rank test p-value (two-sided). scipy 없으면 N/A 반환."""
            try:
                from scipy.stats import wilcoxon
                arr = arr[~np.isnan(arr)]
                if len(arr) < 6 or np.all(arr == 0):
                    return float('nan')
                return float(wilcoxon(arr).pvalue)
            except Exception:
                return float('nan')

        st.markdown('---')
        st.subheader('🔬 자산 배분 효과 심층 분석')
        st.caption(
            '시그널 독립성을 존중한 자산 배분 효과 3개 분석.  '
            '**1) 시그널×자산** — 각 시그널에서 자산별 평균 비중·기여 분석.  '
            '**2) 광역주식 vs ACWI 100%** — 광역 슬리브 내 종목 선택의 알파.  '
            '**3) 전체 주식 vs ACWI 100%** — 개별국가 배분 알파.'
        )

        analysis_tabs = st.tabs([
            '📊 분석 1: 시그널별 자산 효과',
            '🌐 분석 2: 광역주식 vs ACWI 100%',
            '🗺️ 분석 3: 전체 주식 vs ACWI 100%',
        ])

        # ─── 분석 1: 시그널별 자산 효과 ───
        with analysis_tabs[0]:
            st.caption(
                '각 시그널을 **독립 이벤트로 취급** — 시그널별로 자산별 평균 비중·평균 return·평균 기여(bps)를 산출.  '
                '시점 누적 무관 → 월 독립성 존중.  '
                'Y = 자산 (Top 15 비중) / X = 시그널 / 셀 = 평균 기여(bps).'
            )

            def _render_signal_asset(rebals, profile_name, key_suffix):
                df = compute_signal_asset_breakdown(rebals, px, end_date_str)
                if df is None or df.empty:
                    st.info(f'{profile_name} 데이터 부족')
                    return
                # Aggregate: signal × ticker → mean
                agg = df.groupby(['signal', 'sym']).agg(
                    avg_weight=('weight', 'mean'),
                    avg_return=('return_pct', 'mean'),
                    avg_contrib=('contrib_bps', 'mean'),
                    events=('weight', 'size'),
                ).reset_index()

                # Top 15 by average weight across all signals
                top_syms = (df.groupby('sym')['weight'].mean()
                              .sort_values(ascending=False).head(15).index.tolist())
                agg_top = agg[agg['sym'].isin(top_syms)]

                # Pivot for heatmap: rows=ticker, cols=signal
                pivot = agg_top.pivot(index='sym', columns='signal', values='avg_contrib').fillna(0)
                # Reorder rows by top sym order
                pivot = pivot.reindex(top_syms)
                # Reorder cols Bull→Base→Bear
                sig_cols = [s for s in ['Bull', 'Base', 'Bear'] if s in pivot.columns]
                pivot = pivot[sig_cols]

                st.markdown(f'#### 📊 {profile_name} — 시그널 × 자산 평균 기여(bps)')

                # Heatmap
                text_matrix = [[f'{v:+.0f}' for v in row] for row in pivot.values]
                fig_h = go.Figure(data=go.Heatmap(
                    z=pivot.values, x=pivot.columns.tolist(),
                    y=pivot.index.tolist(),
                    text=text_matrix, texttemplate='%{text}',
                    textfont={'size': 11},
                    colorscale='RdYlGn', zmid=0,
                    hovertemplate='%{x} 시그널<br>자산: %{y}<br>평균 기여: %{z:+.1f} bps<extra></extra>',
                    colorbar=dict(title='bps', thickness=12),
                ))
                fig_h.update_layout(
                    height=max(380, len(top_syms) * 25 + 100),
                    xaxis=dict(side='top', tickfont=dict(size=12)),
                    yaxis=dict(autorange='reversed', tickfont=dict(size=10)),
                    margin=dict(l=120, r=20, t=60, b=20),
                )
                st.plotly_chart(fig_h, use_container_width=True,
                                key=f'sig_asset_heatmap_{key_suffix}')

                # Top contributor per signal bar chart
                st.markdown(f'#### 🏆 시그널별 Top 5 기여 자산')
                top_tabs = st.tabs([f'🟢 Bull', f'🟡 Base', f'🔴 Bear'])
                for tab, sig in zip(top_tabs, ['Bull', 'Base', 'Bear']):
                    with tab:
                        sub = agg[agg['signal'] == sig].sort_values('avg_contrib', ascending=False)
                        if sub.empty:
                            st.info(f'{sig} 이벤트 없음')
                            continue
                        top5 = sub.head(5)
                        bot5 = sub.tail(5).sort_values('avg_contrib', ascending=True)
                        c1, c2 = st.columns(2)
                        with c1:
                            st.markdown(f'**🥇 Top 5 (+)**')
                            fig_p = go.Figure(go.Bar(
                                x=top5['avg_contrib'], y=top5['sym'],
                                orientation='h', marker_color='#10B981',
                                text=[f'{v:+.1f}' for v in top5['avg_contrib']],
                                textposition='outside',
                            ))
                            fig_p.update_layout(
                                height=260, yaxis=dict(autorange='reversed'),
                                margin=dict(l=10, r=60, t=10, b=20),
                                xaxis_title='평균 기여 (bps)',
                            )
                            st.plotly_chart(fig_p, use_container_width=True,
                                            key=f'sig_top_{sig}_{key_suffix}')
                        with c2:
                            st.markdown(f'**🥉 Bottom 5 (-)**')
                            fig_b = go.Figure(go.Bar(
                                x=bot5['avg_contrib'], y=bot5['sym'],
                                orientation='h', marker_color='#DC2626',
                                text=[f'{v:+.1f}' for v in bot5['avg_contrib']],
                                textposition='outside',
                            ))
                            fig_b.update_layout(
                                height=260, yaxis=dict(autorange='reversed'),
                                margin=dict(l=10, r=60, t=10, b=20),
                                xaxis_title='평균 기여 (bps)',
                            )
                            st.plotly_chart(fig_b, use_container_width=True,
                                            key=f'sig_bot_{sig}_{key_suffix}')

            a1_tabs = st.tabs(['📈 적극형', '📉 중립형'])
            with a1_tabs[0]:
                _render_signal_asset(agg_rebals, '적극형', 'agg')
            with a1_tabs[1]:
                _render_signal_asset(neu_rebals, '중립형', 'neu')

        # ─── 분석 2: 광역주식 vs ACWI 100% ───
        with analysis_tabs[1]:
            st.caption(
                '광역주식 슬리브(SPY/QQQ/IWM/ACWI/EFA/VWO)를 sleeve-normalize(100%)하여 '
                '**ACWI 단일 100% 시나리오와 누적 비교**.  '
                '알파 = 실제 광역 슬리브 return − ACWI return → 광역 ETF 선택 가치 측정.'
            )

            def _render_broad_vs_acwi(rebals, profile_name, key_suffix):
                df = compute_acwi_cf_alpha(rebals, px, end_date_str, sleeve='broad')
                if df is None or df.empty:
                    st.info(f'{profile_name} 데이터 부족')
                    return
                # KPI
                st.markdown(f'#### 🌐 {profile_name} — 광역주식 vs ACWI 100%')
                final_actual = df['cum_actual'].iloc[-1]
                final_acwi = df['cum_acwi'].iloc[-1]
                final_alpha = final_actual - final_acwi
                n = len(df)
                win_n = (df['alpha'] > 0).sum()
                k1, k2, k3, k4 = st.columns(4)
                k1.metric('광역 슬리브 누적', f'{final_actual:+.2f}%')
                k2.metric('ACWI 100% 누적', f'{final_acwi:+.2f}%')
                k3.metric('알파', f'{final_alpha:+.2f}%p',
                          delta='우위' if final_alpha > 0 else '열위')
                k4.metric('월별 승률', f'{win_n}/{n}회',
                          delta=f'{win_n/n*100:.1f}%' if n else None)

                # Cumulative chart
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=df['date'], y=df['cum_actual'],
                    name='실제 광역 슬리브', mode='lines+markers',
                    line=dict(color='#1F3A68', width=2.5),
                    hovertemplate='%{x|%Y-%m-%d}<br>실제: %{y:+.2f}%<extra></extra>',
                ))
                fig.add_trace(go.Scatter(
                    x=df['date'], y=df['cum_acwi'],
                    name='ACWI 100%', mode='lines+markers',
                    line=dict(color='#9CA3AF', width=2, dash='dash'),
                    hovertemplate='%{x|%Y-%m-%d}<br>ACWI: %{y:+.2f}%<extra></extra>',
                ))
                fig.add_trace(go.Scatter(
                    x=df['date'], y=df['cum_alpha'],
                    name='누적 알파(우축)', mode='lines',
                    line=dict(color='#DC2626', width=2),
                    yaxis='y2',
                    hovertemplate='%{x|%Y-%m-%d}<br>알파: %{y:+.2f}%p<extra></extra>',
                ))
                fig.add_hline(y=0, line_dash='dot', line_color='gray', line_width=1)
                fig.update_layout(
                    height=440, hovermode='x unified',
                    xaxis_title='리밸런스 일자',
                    yaxis=dict(title='누적 수익률 (%)'),
                    yaxis2=dict(title='누적 알파 (%p)', overlaying='y', side='right'),
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                    margin=dict(t=50, b=40, l=40, r=40),
                )
                st.plotly_chart(fig, use_container_width=True,
                                key=f'broad_acwi_{key_suffix}')

                # Per-ETF contribution heatmap
                st.markdown('##### 📊 광역 ETF별 기여 (sleeve 기준)')
                etf_df = compute_broad_etf_contributions(rebals, px, end_date_str)
                if etf_df is not None and not etf_df.empty:
                    agg_etf = etf_df.groupby('sym').agg(
                        avg_weight=('sleeve_weight_pct', 'mean'),
                        avg_return=('return_pct', 'mean'),
                        cum_contrib=('contrib_bps', 'sum'),
                        avg_contrib=('contrib_bps', 'mean'),
                        events=('sleeve_weight_pct', 'count'),
                    ).round(2).sort_values('cum_contrib', ascending=False)
                    agg_etf = agg_etf.rename(columns={
                        'avg_weight':  '평균 슬리브 비중(%)',
                        'avg_return':  '평균 보유기간 return(%)',
                        'cum_contrib': '누적 기여(bps, 슬리브)',
                        'avg_contrib': '평균 기여/리밸(bps)',
                        'events':      '이벤트 수',
                    })
                    styled = agg_etf.style.background_gradient(
                        subset=['누적 기여(bps, 슬리브)', '평균 기여/리밸(bps)'],
                        cmap='RdYlGn', vmin=-200, vmax=200,
                    ).format({
                        '평균 슬리브 비중(%)':       '{:.2f}',
                        '평균 보유기간 return(%)':   '{:+.2f}',
                        '누적 기여(bps, 슬리브)':     '{:+.1f}',
                        '평균 기여/리밸(bps)':        '{:+.1f}',
                    })
                    st.dataframe(styled, use_container_width=True)

                # Monthly alpha distribution
                with st.expander('📊 월별 알파 분포 + 통계 검정'):
                    alpha_arr = df['alpha'].values
                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric('평균 알파', f'{np.mean(alpha_arr):+.2f}%p')
                    k2.metric('중간값', f'{np.median(alpha_arr):+.2f}%p')
                    k3.metric('표준편차', f'{np.std(alpha_arr):.2f}%p')
                    k4.metric('Wilcoxon p', f'{_wilcoxon_p(alpha_arr):.3f}'
                              if len(alpha_arr) >= 6 else 'N/A')
                    fig_h = go.Figure(go.Histogram(
                        x=alpha_arr, nbinsx=25, marker_color='#1F3A68',
                    ))
                    fig_h.add_vline(x=0, line_dash='dash', line_color='red')
                    fig_h.update_layout(
                        height=280, xaxis_title='월별 알파 (%p)', yaxis_title='이벤트 수',
                        margin=dict(t=20, b=40, l=40, r=20),
                    )
                    st.plotly_chart(fig_h, use_container_width=True,
                                    key=f'broad_alpha_dist_{key_suffix}')

            a2_tabs = st.tabs(['📈 적극형', '📉 중립형'])
            with a2_tabs[0]:
                _render_broad_vs_acwi(agg_rebals, '적극형', 'agg')
            with a2_tabs[1]:
                _render_broad_vs_acwi(neu_rebals, '중립형', 'neu')

        # ─── 분석 3: 전체 주식 vs ACWI 100% ───
        with analysis_tabs[2]:
            st.caption(
                '**전체 주식 슬리브** (광역 + 개별국가)를 sleeve-normalize(100%)하여 '
                '**ACWI 단일 100% Counter-factual과 누적 비교**.  '
                '알파 = 전체 주식 슬리브 return − ACWI return → 개별국가 배분 가치 측정.'
            )

            def _render_full_vs_acwi(rebals, profile_name, key_suffix):
                df = compute_acwi_cf_alpha(rebals, px, end_date_str, sleeve='equity_all')
                if df is None or df.empty:
                    st.info(f'{profile_name} 데이터 부족')
                    return
                st.markdown(f'#### 🗺️ {profile_name} — 전체 주식 vs ACWI 100%')
                final_actual = df['cum_actual'].iloc[-1]
                final_acwi = df['cum_acwi'].iloc[-1]
                final_alpha = final_actual - final_acwi
                n = len(df)
                win_n = (df['alpha'] > 0).sum()
                k1, k2, k3, k4 = st.columns(4)
                k1.metric('전체 주식 슬리브 누적', f'{final_actual:+.2f}%')
                k2.metric('ACWI 100% 누적', f'{final_acwi:+.2f}%')
                k3.metric('알파', f'{final_alpha:+.2f}%p',
                          delta='우위' if final_alpha > 0 else '열위')
                k4.metric('월별 승률', f'{win_n}/{n}회',
                          delta=f'{win_n/n*100:.1f}%' if n else None)

                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=df['date'], y=df['cum_actual'],
                    name='실제 전체 주식', mode='lines+markers',
                    line=dict(color='#1F3A68', width=2.5),
                    hovertemplate='%{x|%Y-%m-%d}<br>실제: %{y:+.2f}%<extra></extra>',
                ))
                fig.add_trace(go.Scatter(
                    x=df['date'], y=df['cum_acwi'],
                    name='ACWI 100%', mode='lines+markers',
                    line=dict(color='#9CA3AF', width=2, dash='dash'),
                    hovertemplate='%{x|%Y-%m-%d}<br>ACWI: %{y:+.2f}%<extra></extra>',
                ))
                fig.add_trace(go.Scatter(
                    x=df['date'], y=df['cum_alpha'],
                    name='누적 알파(우축)', mode='lines',
                    line=dict(color='#DC2626', width=2),
                    yaxis='y2',
                    hovertemplate='%{x|%Y-%m-%d}<br>알파: %{y:+.2f}%p<extra></extra>',
                ))
                fig.add_hline(y=0, line_dash='dot', line_color='gray', line_width=1)
                fig.update_layout(
                    height=440, hovermode='x unified',
                    xaxis_title='리밸런스 일자',
                    yaxis=dict(title='누적 수익률 (%)'),
                    yaxis2=dict(title='누적 알파 (%p)', overlaying='y', side='right'),
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                    margin=dict(t=50, b=40, l=40, r=40),
                )
                st.plotly_chart(fig, use_container_width=True,
                                key=f'full_acwi_{key_suffix}')

                # Yearly alpha breakdown
                st.markdown('##### 📅 연도별 누적 알파 (%p)')
                df_y = df.copy()
                df_y['year'] = pd.to_datetime(df_y['date']).dt.year
                yearly = df_y.groupby('year').agg(
                    ev=('alpha', 'count'),
                    avg=('alpha', 'mean'),
                    cum_alpha_pp=('alpha', 'sum'),  # 합산은 simple approx
                ).reset_index()
                fig_y = go.Figure(go.Bar(
                    x=yearly['year'].astype(str), y=yearly['cum_alpha_pp'],
                    marker_color=['#10B981' if v > 0 else '#DC2626'
                                  for v in yearly['cum_alpha_pp']],
                    text=[f'{v:+.2f}%p' for v in yearly['cum_alpha_pp']],
                    textposition='outside',
                    hovertemplate='%{x}년<br>알파 합: %{y:+.2f}%p<extra></extra>',
                ))
                fig_y.add_hline(y=0, line_dash='dot', line_color='gray', line_width=1)
                fig_y.update_layout(
                    height=320, xaxis_title='연도', yaxis_title='연도 내 월별 알파 합 (%p)',
                    margin=dict(t=20, b=40, l=40, r=20),
                )
                st.plotly_chart(fig_y, use_container_width=True,
                                key=f'full_yearly_{key_suffix}')

                # Distribution + test
                with st.expander('📊 월별 알파 분포 + 통계 검정'):
                    alpha_arr = df['alpha'].values
                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric('평균 알파', f'{np.mean(alpha_arr):+.2f}%p')
                    k2.metric('중간값', f'{np.median(alpha_arr):+.2f}%p')
                    k3.metric('표준편차', f'{np.std(alpha_arr):.2f}%p')
                    k4.metric('Wilcoxon p', f'{_wilcoxon_p(alpha_arr):.3f}'
                              if len(alpha_arr) >= 6 else 'N/A')
                    fig_h = go.Figure(go.Histogram(
                        x=alpha_arr, nbinsx=25, marker_color='#1F3A68',
                    ))
                    fig_h.add_vline(x=0, line_dash='dash', line_color='red')
                    fig_h.update_layout(
                        height=280, xaxis_title='월별 알파 (%p)', yaxis_title='이벤트 수',
                        margin=dict(t=20, b=40, l=40, r=20),
                    )
                    st.plotly_chart(fig_h, use_container_width=True,
                                    key=f'full_alpha_dist_{key_suffix}')

            a3_tabs = st.tabs(['📈 적극형', '📉 중립형'])
            with a3_tabs[0]:
                _render_full_vs_acwi(agg_rebals, '적극형', 'agg')
            with a3_tabs[1]:
                _render_full_vs_acwi(neu_rebals, '중립형', 'neu')

    # Holdings CSV downloads
    st.markdown('---')
    dcol1, dcol2 = st.columns(2)
    with dcol1:
        if not agg_perf_df.empty:
            agg_csv = agg_perf_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button('📥 적극형 종목 성과 CSV', data=agg_csv,
                               file_name=f'aimvp_적극형_holdings_{report_date.strftime("%Y%m%d")}.csv',
                               mime='text/csv', use_container_width=True)
    with dcol2:
        if not neu_perf_df.empty:
            neu_csv = neu_perf_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button('📥 중립형 종목 성과 CSV', data=neu_csv,
                               file_name=f'aimvp_중립형_holdings_{report_date.strftime("%Y%m%d")}.csv',
                               mime='text/csv', use_container_width=True)

# ============================================================
# Page: Signal Validation
# ============================================================
elif st.session_state.page == 'Signal Validation':
    # Top KPI cards
    sc1, sc2, sc3, sc4 = st.columns(4)
    hr_pct = hits / n_bull * 100
    sc1.metric('Bull 이벤트 수', f'{n_bull}회')
    sc2.metric('적중 / 미적중', f'{hits} / {n_bull - hits}')
    sc3.metric('적중률', f'{hr_pct:.1f}%',
               delta=f'{hr_pct-50:+.1f}%p vs 코인플립')
    sc4.metric('평균 초과 (QQQ-SPY)', f'{avg_excess:+.2f}%p')

    # Signal status detail
    st.markdown('---')
    st.subheader('🎯 현재 시그널 상태')
    cs1, cs2 = st.columns([1, 2])
    with cs1:
        st.markdown(f'## {sig_color} {current_signal}')
        st.markdown(f'**발효일:** {last[0]}')
        st.markdown(f'**SPY:** {spy_w:.1f}% | **QQQ:** {qqq_w:.1f}%')
        if rule_applied:
            st.error('🔴 50% 스왑 룰 적용 중')
        else:
            st.info('⚪ 룰 미적용 (Base/Bear)')
    with cs2:
        # Compute Win/Loss stats
        excesses = []
        for d in hit_details:
            ex_str = d['초과'].replace('%p','').replace('+','')
            try:
                excesses.append(float(ex_str))
            except: pass
        if excesses:
            wins = [e for e in excesses if e > 0]
            losses = [e for e in excesses if e <= 0]
            avg_w = np.mean(wins) if wins else 0
            avg_l = np.mean(losses) if losses else 0
            wl_ratio = abs(avg_w / avg_l) if avg_l != 0 else 0

            st.markdown('### 📊 Win/Loss 통계 (Bull QQQ vs SPY)')
            mc1, mc2, mc3 = st.columns(3)
            mc1.metric('평균 적중 W', f'+{avg_w:.2f}%p')
            mc2.metric('평균 미적중 L', f'{avg_l:.2f}%p')
            mc3.metric('W/L Ratio', f'{wl_ratio:.2f}')
            st.caption('※ Bull 시그널 발효 시점에 QQQ가 SPY 대비 1m 초과수익 ↗ → 적중')

    # Statistical tests
    st.markdown('---')
    st.subheader('📐 통계적 검정 결과 (Bull QQQ-SPY)')
    stat_df = pd.DataFrame({
        '검정': ['이항분포 (적중률=50%)',
                'Wilcoxon signed-rank (단측)',
                'Permutation test (단측, 10000회)',
                'Block Bootstrap 95% CI',
                'Welch t-test (참조)'],
        '검정통계 / 값': [f'{hits}/{n_bull} = {hr_pct:.1f}%',
                       'W = 178',
                       '—',
                       '[+0.082%p, +1.273%p]',
                       't = 1.681'],
        'p-value': ['0.192',  '0.032', '0.048', '0 미포함', '0.108'],
        '5% 유의': ['✗ 미달', '✓ 유의', '✓ 유의', '✓ 유의', '✗ 미달 (정규성 가정)'],
    })
    st.dataframe(stat_df, use_container_width=True, hide_index=True)
    st.info('★ Wilcoxon/Permutation/Bootstrap 모두 5% 유의 통과 — '
            'Bull 시그널의 QQQ-SPY 스왑 룰이 통계적으로 유효함을 입증.')

    # Bull events detail
    st.markdown('---')
    st.subheader('📋 Bull 이벤트별 적중 상세')
    hit_df = pd.DataFrame(hit_details)
    if not hit_df.empty:
        st.dataframe(hit_df, use_container_width=True, hide_index=True, height=500)
    else:
        st.warning('Bull 이벤트 데이터 없음')

    # Download
    st.markdown('---')
    hit_csv = pd.DataFrame(hit_details).to_csv(index=False).encode('utf-8-sig')
    st.download_button(
        '📥 Bull 이벤트 적중 상세 CSV',
        data=hit_csv,
        file_name=f'aimvp_hit_details_{report_date.strftime("%Y%m%d")}.csv',
        mime='text/csv', use_container_width=True,
    )

    # ============================================================
    # 🧩 슬리브별 리밸런싱 정합 분석 (주식 슬리브 vs 채권 슬리브)
    # ============================================================
    st.markdown('---')
    st.subheader('🧩 슬리브별 리밸런싱 정합 분석 (주식 슬리브 vs 채권 슬리브)')
    st.caption(
        '전체 포트폴리오를 **주식 슬리브(광역+국가)** 와 **채권 슬리브**로 분해해 '
        '슬리브 내부 비중 조절이 얼마나 알파를 만들었는지 추이로 표시.  '
        '계산식: 슬리브 내 ∑_t (w_cur - w_prev) × r_{t, 보유기간} — '
        '슬리브 전체 비중 변화는 제외되고 **상대 비중 재배분 효과만 분리**됩니다.'
    )

    agg_eq = compute_sleeve_rebal_effect_history(agg_rebals, px, end_date_str, 'equity')
    agg_bd = compute_sleeve_rebal_effect_history(agg_rebals, px, end_date_str, 'bond')
    neu_eq = compute_sleeve_rebal_effect_history(neu_rebals, px, end_date_str, 'equity')
    neu_bd = compute_sleeve_rebal_effect_history(neu_rebals, px, end_date_str, 'bond')

    def _sleeve_stats(df):
        if df is None or df.empty:
            return {'total': 0.0, 'n': 0, 'pos': 0, 'hr': 0.0, 'avg': 0.0}
        total = df['impact_bps'].sum()
        n = len(df)
        pos = int((df['impact_bps'] > 0).sum())
        return {'total': total, 'n': n, 'pos': pos,
                'hr': pos / n * 100 if n else 0.0,
                'avg': total / n if n else 0.0}

    s_a_eq = _sleeve_stats(agg_eq)
    s_a_bd = _sleeve_stats(agg_bd)
    s_n_eq = _sleeve_stats(neu_eq)
    s_n_bd = _sleeve_stats(neu_bd)

    # ─── KPI row ───
    st.markdown('##### 📊 슬리브별 누적 알파 & 정합 적중률')
    sc1, sc2, sc3, sc4 = st.columns(4)
    sc1.metric('적극·주식 슬리브', f'{s_a_eq["total"]:+.1f} bps',
               delta=f'적중 {s_a_eq["hr"]:.1f}% ({s_a_eq["pos"]}/{s_a_eq["n"]}회)')
    sc2.metric('적극·채권 슬리브', f'{s_a_bd["total"]:+.1f} bps',
               delta=f'적중 {s_a_bd["hr"]:.1f}% ({s_a_bd["pos"]}/{s_a_bd["n"]}회)')
    sc3.metric('중립·주식 슬리브', f'{s_n_eq["total"]:+.1f} bps',
               delta=f'적중 {s_n_eq["hr"]:.1f}% ({s_n_eq["pos"]}/{s_n_eq["n"]}회)')
    sc4.metric('중립·채권 슬리브', f'{s_n_bd["total"]:+.1f} bps',
               delta=f'적중 {s_n_bd["hr"]:.1f}% ({s_n_bd["pos"]}/{s_n_bd["n"]}회)')

    SLEEVE_SERIES = [
        ('적극 · 주식 슬리브', '#1F3A68', agg_eq),
        ('적극 · 채권 슬리브', '#5B7BA7', agg_bd),
        ('중립 · 주식 슬리브', '#C48D43', neu_eq),
        ('중립 · 채권 슬리브', '#E0B370', neu_bd),
    ]

    # ─── Chart 1: cumulative rebal alpha by sleeve ───
    st.markdown('##### 📈 슬리브별 누적 리밸 알파 추이')
    st.caption('각 슬리브 내 비중 조절의 누적 algorithmic alpha — 운용시작 이후 누적합.')
    any_data = any(df is not None and not df.empty for _, _, df in SLEEVE_SERIES)
    if not any_data:
        st.info('슬리브 리밸 히스토리 부족 — 추이 산출 불가')
    else:
        fig_cum = go.Figure()
        for label, color, df in SLEEVE_SERIES:
            if df is None or df.empty:
                continue
            dash = 'dash' if '채권' in label else 'solid'
            fig_cum.add_trace(go.Scatter(
                x=df['date'], y=df['cumulative_bps'],
                name=label, mode='lines+markers',
                line=dict(color=color, width=2.4, dash=dash),
                marker=dict(size=5),
                hovertemplate='%{x|%Y-%m-%d}<br>' + label + ': %{y:+.1f} bps<extra></extra>',
            ))
        fig_cum.add_hline(y=0, line_dash='dot', line_color='gray', line_width=1)
        fig_cum.update_layout(
            height=460, hovermode='x unified',
            xaxis_title='리밸런스 일자', yaxis_title='누적 리밸 알파 (bps)',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
            margin=dict(t=50, b=40, l=40, r=20),
        )
        st.plotly_chart(fig_cum, use_container_width=True, key='sleeve_cumulative_alpha')

        # ─── Chart 2: rolling cumulative hit rate ───
        st.markdown('##### 🎯 슬리브별 정합 적중률 추이 (누적 %)')
        st.caption(
            '각 리밸 시점까지의 **누적 정합 적중률** = 양(+) period 건수 / 전체 건수.  '
            '50% = 코인플립 (의사결정이 무작위 수준).  '
            '50% 초과·우상향 = 의사결정의 일관된 유효성.'
        )
        fig_hr = go.Figure()
        for label, color, df in SLEEVE_SERIES:
            if df is None or df.empty:
                continue
            dash = 'dash' if '채권' in label else 'solid'
            fig_hr.add_trace(go.Scatter(
                x=df['date'], y=df['cum_hit_rate'],
                name=label, mode='lines',
                line=dict(color=color, width=2.4, dash=dash),
                hovertemplate='%{x|%Y-%m-%d}<br>' + label + ': %{y:.1f}%<extra></extra>',
            ))
        fig_hr.add_hline(y=50, line_dash='dash', line_color='red', line_width=1.2,
                         annotation_text='코인플립 50%', annotation_position='top right')
        fig_hr.update_layout(
            height=420, hovermode='x unified',
            xaxis_title='리밸런스 일자', yaxis_title='누적 정합 적중률 (%)',
            yaxis=dict(range=[0, 100]),
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
            margin=dict(t=50, b=40, l=40, r=20),
        )
        st.plotly_chart(fig_hr, use_container_width=True, key='sleeve_hit_rate_trend')

        # ─── Summary table ───
        st.markdown('##### 📋 슬리브별 정합 요약 (전체 기간)')
        summary_rows = [
            {'Profile': '적극형', '슬리브': '🌐 주식 (광역+국가)',
             '리밸 횟수': s_a_eq['n'], '양(+) 적중': s_a_eq['pos'],
             '적중률(%)': s_a_eq['hr'], '누적 알파(bps)': s_a_eq['total'],
             '평균/리밸(bps)': s_a_eq['avg']},
            {'Profile': '적극형', '슬리브': '💵 채권',
             '리밸 횟수': s_a_bd['n'], '양(+) 적중': s_a_bd['pos'],
             '적중률(%)': s_a_bd['hr'], '누적 알파(bps)': s_a_bd['total'],
             '평균/리밸(bps)': s_a_bd['avg']},
            {'Profile': '중립형', '슬리브': '🌐 주식 (광역+국가)',
             '리밸 횟수': s_n_eq['n'], '양(+) 적중': s_n_eq['pos'],
             '적중률(%)': s_n_eq['hr'], '누적 알파(bps)': s_n_eq['total'],
             '평균/리밸(bps)': s_n_eq['avg']},
            {'Profile': '중립형', '슬리브': '💵 채권',
             '리밸 횟수': s_n_bd['n'], '양(+) 적중': s_n_bd['pos'],
             '적중률(%)': s_n_bd['hr'], '누적 알파(bps)': s_n_bd['total'],
             '평균/리밸(bps)': s_n_bd['avg']},
        ]
        summary_df = pd.DataFrame(summary_rows)
        sum_styled = summary_df.style.background_gradient(
            subset=['적중률(%)'], cmap='RdYlGn', vmin=30, vmax=70,
        ).background_gradient(
            subset=['누적 알파(bps)', '평균/리밸(bps)'], cmap='RdYlGn', vmin=-200, vmax=200,
        ).format({
            '적중률(%)':      '{:.1f}',
            '누적 알파(bps)': '{:+.1f}',
            '평균/리밸(bps)': '{:+.1f}',
        })
        st.dataframe(sum_styled, use_container_width=True, hide_index=True)

        # ─── Sleeve detail expander ───
        with st.expander('📊 슬리브별 시점별 효과 상세 (모든 리밸)'):
            detail_rows = []
            for label, _, df in SLEEVE_SERIES:
                if df is None or df.empty:
                    continue
                t = df[['date', 'signal', 'impact_bps', 'cumulative_bps', 'cum_hit_rate']].copy()
                t.insert(0, '슬리브', label)
                t['date'] = pd.to_datetime(t['date']).dt.strftime('%Y-%m-%d')
                t = t.rename(columns={
                    'date': '리밸 일자',
                    'signal': '시그널',
                    'impact_bps': 'period(bps)',
                    'cumulative_bps': '누적(bps)',
                    'cum_hit_rate': '누적 적중률(%)',
                })
                detail_rows.append(t)
            if detail_rows:
                detail_df = pd.concat(detail_rows, ignore_index=True).sort_values(
                    ['리밸 일자', '슬리브']).reset_index(drop=True)
                detail_styled = detail_df.style.background_gradient(
                    subset=['period(bps)', '누적(bps)'], cmap='RdYlGn', vmin=-200, vmax=200,
                ).background_gradient(
                    subset=['누적 적중률(%)'], cmap='RdYlGn', vmin=30, vmax=70,
                ).format({
                    'period(bps)':     '{:+.1f}',
                    '누적(bps)':       '{:+.1f}',
                    '누적 적중률(%)': '{:.1f}',
                }, na_rep='—')
                st.dataframe(detail_styled, use_container_width=True, hide_index=True, height=480)

        # ─── Interpretation hint ───
        st.markdown('##### 💡 해석 가이드')
        eq_hr_avg = (s_a_eq['hr'] + s_n_eq['hr']) / 2 if (s_a_eq['n'] + s_n_eq['n']) > 0 else 0
        bd_hr_avg = (s_a_bd['hr'] + s_n_bd['hr']) / 2 if (s_a_bd['n'] + s_n_bd['n']) > 0 else 0
        eq_tot = s_a_eq['total'] + s_n_eq['total']
        bd_tot = s_a_bd['total'] + s_n_bd['total']
        st.info(
            f'**주식 슬리브:** 평균 적중률 {eq_hr_avg:.1f}%, 양 profile 합산 누적 {eq_tot:+.1f} bps — '
            f'{"의사결정 유효 ✓" if eq_hr_avg > 55 and eq_tot > 0 else ("코인플립 수준" if 45 <= eq_hr_avg <= 55 else "재검토 권고")}.  '
            f'**채권 슬리브:** 평균 적중률 {bd_hr_avg:.1f}%, 누적 {bd_tot:+.1f} bps — '
            f'{"의사결정 유효 ✓" if bd_hr_avg > 55 and bd_tot > 0 else ("코인플립 수준" if 45 <= bd_hr_avg <= 55 else "재검토 권고")}.'
        )

        st.caption(
            ':grey[**📊 데이터 출처:** Yahoo Finance — auto-adjusted (Total Return).  '
            f'**🗓️ 기간:** {agg_rebals[0][0] if agg_rebals else "—"} → {end_date_str}.  '
            '**주식 슬리브:** 광역 주식(SPY/QQQ/ACWI 등) + 국가 주식(EWJ/MCHI 등).  '
            '**채권 슬리브:** 채권 카테고리 (HYG, LQD, IEF, EMBD, PHYL, BKLN, GTO, KR채권 등).  '
            '**정합 적중률(Hit Rate):** 슬리브 내 비중 조절이 양(+) 알파를 만든 비율 (50% = 코인플립).  '
            '**누적 알파:** 슬리브 내부 비중 조절의 누적 기여 — 슬리브 전체 비중 변화·다른 슬리브 효과는 제외.]'
        )

        # ============================================================
        # 🛡️ 주식 하방 완충 분석 (방어 슬리브 = 채권 + 현금)
        # ============================================================
        st.markdown('---')
        st.markdown('#### 🛡️ 주식 하방 완충 분석 (채권+현금 슬리브 헷지 효과)')
        st.caption(
            '**주식 슬리브 contribution이 음(-)인 리밸 시점**에서 방어 슬리브(채권+현금)가 얼마나 손실을 상쇄했는지 측정.  '
            '**헷지 비율 = 방어 슬리브 기여 / |주식 슬리브 손실|** — ≥100% 완전 헷지, 0~100% 부분 완충, <0% 동반 손실.  '
            '※ 각 슬리브 기여 = ∑_t (현재 비중 w × 보유 기간 종목 return r) — **포트폴리오 레벨 기여(bps)**.'
        )

        agg_dh = compute_downside_hedge_history(agg_rebals, px, end_date_str)
        neu_dh = compute_downside_hedge_history(neu_rebals, px, end_date_str)

        def _dh_stats(df):
            if df is None or df.empty:
                return {'n_total': 0, 'n_down': 0, 'avg_eq': 0.0, 'avg_def': 0.0,
                        'avg_hr': 0.0, 'hedge_succ': 0, 'full_hedge': 0, 'avg_net': 0.0,
                        'sum_eq': 0.0, 'sum_def': 0.0}
            n_total = len(df)
            down = df[df['is_downside']]
            if down.empty:
                return {'n_total': n_total, 'n_down': 0, 'avg_eq': 0.0, 'avg_def': 0.0,
                        'avg_hr': 0.0, 'hedge_succ': 0, 'full_hedge': 0, 'avg_net': 0.0,
                        'sum_eq': 0.0, 'sum_def': 0.0}
            return {
                'n_total': n_total,
                'n_down': len(down),
                'avg_eq': down['equity_contrib_bps'].mean(),
                'avg_def': down['defensive_contrib_bps'].mean(),
                'avg_hr': down['hedge_ratio'].mean(),
                'hedge_succ': int((down['hedge_ratio'] > 0).sum()),
                'full_hedge': int((down['hedge_ratio'] >= 100).sum()),
                'avg_net': down['net_contrib_bps'].mean(),
                'sum_eq': down['equity_contrib_bps'].sum(),
                'sum_def': down['defensive_contrib_bps'].sum(),
            }

        a_dh_st = _dh_stats(agg_dh)
        n_dh_st = _dh_stats(neu_dh)

        # ─── KPI cards ───
        st.markdown('##### 📊 주식 하방 이벤트 & 완충 KPI')
        dc1, dc2, dc3, dc4 = st.columns(4)
        dc1.metric(
            '적극 · 하방 이벤트',
            f'{a_dh_st["n_down"]}/{a_dh_st["n_total"]}회',
            delta=f'평균 주식 손실 {a_dh_st["avg_eq"]:+.1f} bps' if a_dh_st['n_down'] else None,
        )
        dc2.metric(
            '적극 · 평균 헷지 비율',
            f'{a_dh_st["avg_hr"]:+.1f}%' if a_dh_st['n_down'] else 'N/A',
            delta=f'평균 방어 완충 {a_dh_st["avg_def"]:+.1f} bps' if a_dh_st['n_down'] else None,
        )
        dc3.metric(
            '중립 · 하방 이벤트',
            f'{n_dh_st["n_down"]}/{n_dh_st["n_total"]}회',
            delta=f'평균 주식 손실 {n_dh_st["avg_eq"]:+.1f} bps' if n_dh_st['n_down'] else None,
        )
        dc4.metric(
            '중립 · 평균 헷지 비율',
            f'{n_dh_st["avg_hr"]:+.1f}%' if n_dh_st['n_down'] else 'N/A',
            delta=f'평균 방어 완충 {n_dh_st["avg_def"]:+.1f} bps' if n_dh_st['n_down'] else None,
        )

        # ─── Scatter: 주식 기여 vs 방어 기여 (전체 period) ───
        st.markdown('##### 📍 헷지 효과 산점도 (주식 기여 vs 방어 기여)')
        st.caption(
            '좌측 영역(주식 < 0) × 상단(방어 > 0) = 헷지 성공 (✅).  '
            '좌측 × 하단 = 동반 손실 (❌).  점선 **y = -x** 위쪽이면 완전 헷지.'
        )
        fig_scat = go.Figure()
        for label, color, dh in [('적극형', '#DC2626', agg_dh), ('중립형', '#2563EB', neu_dh)]:
            if dh is None or dh.empty:
                continue
            fig_scat.add_trace(go.Scatter(
                x=dh['equity_contrib_bps'], y=dh['defensive_contrib_bps'],
                mode='markers', name=label,
                marker=dict(color=color, size=9, opacity=0.7,
                            line=dict(color='white', width=1)),
                text=[pd.Timestamp(d).strftime('%Y-%m-%d') for d in dh['date']],
                customdata=dh['signal'],
                hovertemplate=(
                    '%{text} (%{customdata})<br>'
                    '주식 기여: %{x:+.1f} bps<br>'
                    '방어 기여: %{y:+.1f} bps<extra>' + label + '</extra>'
                ),
            ))
        # 완전 헷지 reference line (y = -x)
        combined_vals = []
        for dh in [agg_dh, neu_dh]:
            if dh is not None and not dh.empty:
                combined_vals.extend(dh['equity_contrib_bps'].tolist())
                combined_vals.extend(dh['defensive_contrib_bps'].tolist())
        if combined_vals:
            r_max = max(abs(min(combined_vals)), abs(max(combined_vals))) * 1.1
            fig_scat.add_trace(go.Scatter(
                x=[-r_max, r_max], y=[r_max, -r_max],
                mode='lines', name='완전 헷지 (y = -x)',
                line=dict(color='gray', width=1.2, dash='dash'),
                hoverinfo='skip',
            ))
        fig_scat.add_hline(y=0, line_dash='dot', line_color='gray', line_width=1)
        fig_scat.add_vline(x=0, line_dash='dot', line_color='gray', line_width=1)
        fig_scat.update_layout(
            height=480, hovermode='closest',
            xaxis_title='주식 슬리브 기여 (bps)',
            yaxis_title='방어 슬리브 기여 (bps)',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
            margin=dict(l=40, r=20, t=50, b=40),
        )
        st.plotly_chart(fig_scat, use_container_width=True, key='hedge_scatter')

        # ─── 하방 이벤트별 손실·완충 분해 (적극·중립 각각) ───
        st.markdown('##### 📊 하방 시점별 손실·완충 분해 (주식 음수 period 한정)')
        st.caption('빨강 = 주식 슬리브 손실, 녹색 = 방어 슬리브 완충, ◆ 검은 마커 = 순 영향 (주식+방어).')
        bar_tabs = st.tabs(['📈 적극형', '📉 중립형'])
        for tab, (prof, dh) in zip(bar_tabs, [('적극형', agg_dh), ('중립형', neu_dh)]):
            with tab:
                if dh is None or dh.empty:
                    st.info(f'{prof}: 데이터 부족')
                    continue
                down = dh[dh['is_downside']].copy()
                if down.empty:
                    st.success(f'{prof}: 주식 슬리브 음수 period가 없습니다 (전 기간 양호).')
                    continue
                fig_b = go.Figure()
                fig_b.add_trace(go.Bar(
                    x=down['date'], y=down['equity_contrib_bps'],
                    name='주식 슬리브 (손실)', marker_color='#DC2626',
                    hovertemplate='%{x|%Y-%m-%d}<br>주식: %{y:+.1f} bps<extra></extra>',
                ))
                fig_b.add_trace(go.Bar(
                    x=down['date'], y=down['defensive_contrib_bps'],
                    name='방어 슬리브 (완충)', marker_color='#10B981',
                    hovertemplate='%{x|%Y-%m-%d}<br>방어: %{y:+.1f} bps<extra></extra>',
                ))
                fig_b.add_trace(go.Scatter(
                    x=down['date'], y=down['net_contrib_bps'],
                    name='순 영향 (주식+방어)', mode='markers',
                    marker=dict(color='black', size=11, symbol='diamond',
                                line=dict(color='white', width=1.2)),
                    hovertemplate='%{x|%Y-%m-%d}<br>순: %{y:+.1f} bps<extra></extra>',
                ))
                fig_b.add_hline(y=0, line_dash='dot', line_color='black', line_width=1)
                fig_b.update_layout(
                    barmode='relative', height=400, hovermode='x unified',
                    xaxis_title='리밸 일자', yaxis_title='기여 (bps)',
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                    margin=dict(l=40, r=20, t=50, b=40),
                )
                st.plotly_chart(fig_b, use_container_width=True, key=f'hedge_bar_{prof}')

        # ─── 헷지 효과 요약 테이블 ───
        st.markdown('##### 📋 헷지 효과 요약 (전체 기간)')
        hedge_rows = []
        for prof, st_dict in [('적극형', a_dh_st), ('중립형', n_dh_st)]:
            succ_rate = (st_dict['hedge_succ'] / st_dict['n_down'] * 100) if st_dict['n_down'] else 0
            full_rate = (st_dict['full_hedge'] / st_dict['n_down'] * 100) if st_dict['n_down'] else 0
            hedge_rows.append({
                'Profile': prof,
                '주식 하방(회)':       st_dict['n_down'],
                '평균 주식 손실(bps)': st_dict['avg_eq'],
                '평균 방어 완충(bps)': st_dict['avg_def'],
                '평균 헷지 비율(%)':   st_dict['avg_hr'],
                '평균 순 영향(bps)':    st_dict['avg_net'],
                '헷지 성공률(%)':       succ_rate,
                '완전 헷지률(%)':       full_rate,
                '누적 주식 손실(bps)': st_dict['sum_eq'],
                '누적 방어 완충(bps)': st_dict['sum_def'],
            })
        hedge_df = pd.DataFrame(hedge_rows)
        hedge_styled = hedge_df.style.background_gradient(
            subset=['평균 헷지 비율(%)', '헷지 성공률(%)', '완전 헷지률(%)'],
            cmap='RdYlGn', vmin=0, vmax=100,
        ).background_gradient(
            subset=['평균 방어 완충(bps)', '평균 순 영향(bps)', '누적 방어 완충(bps)'],
            cmap='RdYlGn', vmin=-200, vmax=200,
        ).background_gradient(
            subset=['평균 주식 손실(bps)', '누적 주식 손실(bps)'],
            cmap='RdYlGn', vmin=-500, vmax=0,
        ).format({
            '평균 주식 손실(bps)': '{:+.1f}',
            '평균 방어 완충(bps)': '{:+.1f}',
            '평균 헷지 비율(%)':   '{:+.1f}',
            '평균 순 영향(bps)':    '{:+.1f}',
            '헷지 성공률(%)':       '{:.1f}',
            '완전 헷지률(%)':       '{:.1f}',
            '누적 주식 손실(bps)': '{:+.1f}',
            '누적 방어 완충(bps)': '{:+.1f}',
        })
        st.dataframe(hedge_styled, use_container_width=True, hide_index=True)

        # ─── 하방 이벤트 상세 expander ───
        with st.expander('📊 주식 하방 이벤트 상세 (모든 음수 period)'):
            detail_rows = []
            for prof, dh in [('적극형', agg_dh), ('중립형', neu_dh)]:
                if dh is None or dh.empty:
                    continue
                down = dh[dh['is_downside']].copy()
                if down.empty:
                    continue
                down.insert(0, 'Profile', prof)
                down['date'] = pd.to_datetime(down['date']).dt.strftime('%Y-%m-%d')
                down['outcome'] = np.where(down['hedge_ratio'] >= 100, '✅ 완전 헷지',
                                  np.where(down['hedge_ratio'] > 0, '⚠️ 부분 완충',
                                           '❌ 동반 손실'))
                down = down[['Profile', 'date', 'signal', 'equity_weight', 'defensive_weight',
                             'equity_contrib_bps', 'defensive_contrib_bps', 'hedge_ratio',
                             'net_contrib_bps', 'outcome']]
                down = down.rename(columns={
                    'date':                  '리밸 일자',
                    'signal':                '시그널',
                    'equity_weight':         '주식 비중(%)',
                    'defensive_weight':      '방어 비중(%)',
                    'equity_contrib_bps':    '주식 기여(bps)',
                    'defensive_contrib_bps': '방어 기여(bps)',
                    'hedge_ratio':           '헷지 비율(%)',
                    'net_contrib_bps':       '순 영향(bps)',
                    'outcome':               '판정',
                })
                detail_rows.append(down)
            if detail_rows:
                detail_df = pd.concat(detail_rows, ignore_index=True).sort_values(
                    ['리밸 일자', 'Profile']).reset_index(drop=True)
                styled = detail_df.style.background_gradient(
                    subset=['헷지 비율(%)'], cmap='RdYlGn', vmin=0, vmax=150,
                ).background_gradient(
                    subset=['방어 기여(bps)', '순 영향(bps)'], cmap='RdYlGn', vmin=-200, vmax=200,
                ).background_gradient(
                    subset=['주식 기여(bps)'], cmap='RdYlGn', vmin=-500, vmax=0,
                ).format({
                    '주식 비중(%)':    '{:.2f}',
                    '방어 비중(%)':    '{:.2f}',
                    '주식 기여(bps)':  '{:+.1f}',
                    '방어 기여(bps)':  '{:+.1f}',
                    '헷지 비율(%)':    '{:+.1f}',
                    '순 영향(bps)':    '{:+.1f}',
                }, na_rep='—')
                st.dataframe(styled, use_container_width=True, hide_index=True, height=440)
            else:
                st.info('주식 하방 이벤트가 없습니다.')

        # ─── 자동 해석 ───
        st.markdown('##### 💡 하방 완충 효과 해석')
        def _judge(succ, avg_hr):
            if succ >= 70 and avg_hr >= 30: return '✅ 효과적 완충'
            if succ >= 50 and avg_hr >= 10: return '⚠️ 부분 완충 (강화 여지)'
            return '❌ 완충 미흡 — 동반 손실 빈번 (채권 다변화·duration 점검 필요)'

        a_succ = (a_dh_st['hedge_succ'] / a_dh_st['n_down'] * 100) if a_dh_st['n_down'] else 0
        n_succ = (n_dh_st['hedge_succ'] / n_dh_st['n_down'] * 100) if n_dh_st['n_down'] else 0
        st.info(
            f'**적극형 (주식 ~89%):** 하방 {a_dh_st["n_down"]}회 중 완충 성공 {a_dh_st["hedge_succ"]}회 ({a_succ:.1f}%), '
            f'평균 헷지 비율 {a_dh_st["avg_hr"]:+.1f}% — {_judge(a_succ, a_dh_st["avg_hr"])}.  '
            f'**중립형 (주식 ~60%):** 하방 {n_dh_st["n_down"]}회 중 완충 성공 {n_dh_st["hedge_succ"]}회 ({n_succ:.1f}%), '
            f'평균 헷지 비율 {n_dh_st["avg_hr"]:+.1f}% — {_judge(n_succ, n_dh_st["avg_hr"])}.  '
            f'※ 적극형은 방어 비중(~11%)이 구조적으로 작아 헷지 비율 상한이 제한됨 — '
            f'중립형(~40%)이 본질적으로 더 큰 완충력 보유.'
        )

        st.caption(
            ':grey[**🛡️ 방어 슬리브 정의:** 채권 + 현금 (HYG/LQD/IEF/EMBD/PHYL/BKLN/GTO/KR채권 + USDKRW/KRW Cash).  '
            '**기여 계산:** 슬리브 내 ∑_t (현재 비중 × 보유 기간 종목 return) → 포트폴리오 레벨 bps.  '
            '**헷지 비율:** 방어 슬리브 기여 / |주식 슬리브 손실| (주식 음수 period 한정).  '
            '**판정 기준:** ≥100% 완전 헷지 / 0~100% 부분 완충 / <0% 동반 손실 (채권·주식 상관관계 양수화).  '
            '**해석 한계:** 같은 주식 손실 절댓값에서도 방어 비중이 크면 헷지 비율이 자연스럽게 커지므로 profile 간 직접 비교는 비중 차이를 함께 고려할 것.]'
        )

# ============================================================
# Page: Universe
# ============================================================
elif st.session_state.page == 'Universe':
    st.markdown('## 🌐 AIMVP 투자 유니버스')
    st.caption(
        'AIMVP 펀드의 **공식 투자 유니버스** — Excel `투자유니버스` 시트 기반. '
        '각 종목의 카테고리(구분) · Bloomberg ticker · yfinance 매핑 · 전략 분류 표시.'
    )

    @st.cache_data(ttl=3600)
    def load_universe():
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        if '투자유니버스' not in wb.sheetnames:
            return None
        ws = wb['투자유니버스']
        rows = []
        for i in range(3, ws.max_row + 1):  # 2 = header, 3+ = data
            cat = ws.cell(i, 2).value
            name = ws.cell(i, 3).value
            ticker = ws.cell(i, 4).value
            strategy = ws.cell(i, 5).value
            if not ticker:
                continue
            tk = std_ticker(ticker) or std_ticker(name)
            yf_sym = TICKER_MAP.get(tk, '')
            rows.append({
                '구분':           cat or '기타',
                '종목명':         name or '',
                'Bloomberg Ticker': ticker,
                'yfinance Symbol': yf_sym if yf_sym and not yf_sym.startswith('_') else (
                    '(KR proxy)' if yf_sym and yf_sym.startswith('_') else '-'),
                '전략':           strategy or '-',
                '세부 카테고리':   TICKER_CATEGORY.get(tk, '기타'),
                '_tk':            tk,
            })
        return pd.DataFrame(rows)

    # ─── Period return 계산 ───
    _end_ts_u = pd.Timestamp(end_date_str)
    _px_u = px[px.index <= _end_ts_u]
    _today_u = _px_u.index[-1] if len(_px_u) > 0 else None

    def _univ_ret(tk, days_back, annualize_years=None):
        """기간 total return(%). 데이터 부족 시 None."""
        sym = TICKER_MAP.get(tk)
        if sym is None or sym not in _px_u.columns or _today_u is None:
            return None
        if days_back == 1:  # 1D = 직전 영업일 대비
            if len(_px_u) < 2:
                return None
            start_idx = _px_u.index[-2]
        else:
            target = _today_u - pd.Timedelta(days=days_back)
            idx_before = _px_u.index[_px_u.index <= target]
            if len(idx_before) == 0:
                return None
            start_idx = idx_before[-1]
        try:
            end_v = _px_u[sym].loc[_today_u]
            start_v = _px_u[sym].loc[start_idx]
            if start_v == 0 or pd.isna(start_v) or pd.isna(end_v):
                return None
            total_ret = end_v / start_v - 1
            if annualize_years:
                if total_ret <= -1:
                    return None
                return ((1 + total_ret) ** (1 / annualize_years) - 1) * 100
            return total_ret * 100
        except Exception:
            return None

    PERIOD_SPECS = [
        ('1D(%)',     1,        None),
        ('1W(%)',     7,        None),
        ('1M(%)',     30,       None),
        ('3M(%)',     90,       None),
        ('6M(%)',     180,      None),
        ('1Y(%)',     365,      None),
        ('3Y CAGR(%)', 365 * 3, 3),
        ('5Y CAGR(%)', 365 * 5, 5),
    ]

    universe_df = load_universe()
    if universe_df is None or universe_df.empty:
        st.error('투자유니버스 시트를 읽을 수 없습니다.')
    else:
        # 기간 return 컬럼 추가 (cache 회피 위해 외부에서 적용)
        for col_name, days, ann in PERIOD_SPECS:
            universe_df[col_name] = universe_df['_tk'].apply(
                lambda tk: _univ_ret(tk, days, ann))
        # ─── KPI 카드 ───
        st.markdown('### 📊 유니버스 구성 요약')
        n_total = len(universe_df)
        n_eq = (universe_df['구분'] == '주식').sum()
        n_bd = (universe_df['구분'] == '채권').sum()
        n_other = n_total - n_eq - n_bd
        k1, k2, k3, k4 = st.columns(4)
        k1.metric('총 종목 수', f'{n_total}개')
        k2.metric('주식 ETF', f'{n_eq}개', delta=f'{n_eq/n_total*100:.0f}%')
        k3.metric('채권 ETF', f'{n_bd}개', delta=f'{n_bd/n_total*100:.0f}%')
        k4.metric('기타 (헤지펀드·원자재)', f'{n_other}개',
                   delta=f'{n_other/n_total*100:.0f}%')

        # ─── 분포 차트 (구분 + 전략) ───
        c1, c2 = st.columns(2)
        with c1:
            st.markdown('##### 🏷️ 구분별 분포')
            cat_counts = universe_df['구분'].value_counts()
            fig_cat = go.Figure(go.Pie(
                labels=cat_counts.index, values=cat_counts.values,
                hole=0.4, marker=dict(colors=['#1F3A68', '#C48D43', '#9CA3AF', '#10B981']),
                hovertemplate='%{label}: %{value}개 (%{percent})<extra></extra>',
            ))
            fig_cat.update_layout(height=280, margin=dict(t=10, b=10, l=10, r=10),
                                    showlegend=True)
            st.plotly_chart(fig_cat, use_container_width=True, key='univ_cat_pie')
        with c2:
            st.markdown('##### 🎯 전략별 분포')
            stg_counts = universe_df['전략'].value_counts()
            fig_stg = go.Figure(go.Pie(
                labels=stg_counts.index, values=stg_counts.values,
                hole=0.4, marker=dict(colors=['#1F77B4', '#FF7F0E', '#2CA02C']),
                hovertemplate='%{label}: %{value}개 (%{percent})<extra></extra>',
            ))
            fig_stg.update_layout(height=280, margin=dict(t=10, b=10, l=10, r=10),
                                    showlegend=True)
            st.plotly_chart(fig_stg, use_container_width=True, key='univ_stg_pie')

        # ─── 필터 ───
        st.markdown('---')
        st.markdown('### 📋 종목 리스트 (필터 가능)')
        f1, f2, f3 = st.columns(3)
        with f1:
            cat_filter = st.multiselect(
                '구분 필터',
                options=sorted(universe_df['구분'].unique()),
                default=sorted(universe_df['구분'].unique()),
                key='univ_cat_filter',
            )
        with f2:
            stg_filter = st.multiselect(
                '전략 필터',
                options=sorted(universe_df['전략'].unique()),
                default=sorted(universe_df['전략'].unique()),
                key='univ_stg_filter',
            )
        with f3:
            search_q = st.text_input(
                '검색 (종목명/Ticker)',
                placeholder='예: SPY, MSCI, 채권 …',
                key='univ_search',
            )

        # 필터 적용
        filtered = universe_df.copy()
        if cat_filter:
            filtered = filtered[filtered['구분'].isin(cat_filter)]
        if stg_filter:
            filtered = filtered[filtered['전략'].isin(stg_filter)]
        if search_q:
            q = search_q.lower()
            mask = (
                filtered['종목명'].str.lower().str.contains(q, na=False) |
                filtered['Bloomberg Ticker'].str.lower().str.contains(q, na=False) |
                filtered['yfinance Symbol'].str.lower().str.contains(q, na=False)
            )
            filtered = filtered[mask]

        st.caption(f'**{len(filtered)}개 / 총 {n_total}개** 표시  |  '
                    f'기간 기준: {_today_u.strftime("%Y-%m-%d") if _today_u is not None else "—"} (Report Date)')

        # ─── 메인 테이블 ───
        display_df = filtered.reset_index(drop=True).copy()
        # 내부 컬럼 제거
        if '_tk' in display_df.columns:
            display_df = display_df.drop(columns=['_tk'])
        display_df.insert(0, '#', range(1, len(display_df) + 1))

        return_cols = [c for c, _, _ in PERIOD_SPECS]
        styled_univ = display_df.style.background_gradient(
            subset=return_cols, cmap='RdYlGn', vmin=-30, vmax=30,
        ).format({c: '{:+.2f}' for c in return_cols}, na_rep='—')

        st.dataframe(
            styled_univ,
            use_container_width=True, hide_index=True,
            height=min(720, 45 + len(display_df) * 36),
        )

        # ─── 데이터 출처 / 다운로드 ───
        st.markdown('---')
        dc1, dc2 = st.columns([3, 1])
        with dc1:
            st.caption(
                ':grey[**📊 데이터 출처:** AIMVP Excel `투자유니버스` 시트 (2026-06 기준).  '
                '**yfinance Symbol** = Bloomberg ticker → yfinance 매핑 (`TICKER_MAP`).  '
                '**(KR proxy)** = KR 채권/현금 proxy 상수 수익률 적용 (yfinance 가격 데이터 부재).  '
                '**세부 카테고리** = `TICKER_CATEGORY` 분류 (광역 주식 / 국가 주식 / 채권 / 현금 / 원자재 / 헤지펀드).  '
                '**Return 컬럼:** Yahoo Finance auto-adjusted close (Total Return, 배당 재투자 반영).  '
                '**1D** = 직전 영업일 대비.  **1W/1M/3M/6M/1Y** = 캘린더 일수 기준 가장 가까운 영업일.  '
                '**3Y/5Y CAGR** = 연환산 수익률 (3년/5년 누적 → (1+r)^(1/n)-1).  '
                '데이터 부족 시 "—" 표시 (KR proxy ETF, 신규 상장 ETF 등).]'
            )
        with dc2:
            csv_df = filtered.drop(columns=['_tk'], errors='ignore')
            csv = csv_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                '📥 CSV 다운로드',
                data=csv,
                file_name=f'aimvp_universe_{report_date.strftime("%Y%m%d")}.csv',
                mime='text/csv',
                use_container_width=True,
            )

# ============================================================
# Page: 월별 포트폴리오
# ============================================================
elif st.session_state.page == '월별 포트폴리오':
    st.markdown('## 📑 월별 포트폴리오 추이 (필터링 가능)')
    st.caption(
        '운용시작 ~ 조회일까지 **모든 월별 리밸런스**의 자산 비중 + 보유 기간 ETF total return + 시그널 + ACWI 성과.  '
        '🟢/🔴 ACWI 방향 + Bull/Base/Bear 시그널 + Profile(적극/중립) 다중 필터.'
    )

    # ─── Build full monthly portfolio dataframe ───
    @st.cache_data(ttl=3600, show_spinner='월별 포트폴리오 데이터 빌드 중...')
    def _build_monthly_portfolio_df(end_date_str_arg):
        """For each rebal in both profiles, compute monthly portfolio composition + ETF returns + ACWI return."""
        end_ts = pd.Timestamp(end_date_str_arg)
        px_up = px[px.index <= end_ts]
        if len(px_up) < 2 or 'ACWI' not in px.columns:
            return pd.DataFrame()
        rows = []
        for profile_name, rebals in [('적극형', agg_rebals), ('중립형', neu_rebals)]:
            for i, (date_str, signal, weights) in enumerate(rebals):
                rb_date = pd.Timestamp(date_str)
                if rb_date > end_ts:
                    continue
                if i + 1 < len(rebals):
                    next_date = min(pd.Timestamp(rebals[i+1][0]), end_ts)
                else:
                    next_date = end_ts
                idx = px_up.index[(px_up.index >= rb_date) & (px_up.index <= next_date)]
                if len(idx) < 2:
                    continue
                p0, p1 = idx[0], idx[-1]
                # ACWI return for this period
                try:
                    acwi_ret = (px['ACWI'].loc[p1] / px['ACWI'].loc[p0] - 1) * 100
                except Exception:
                    acwi_ret = 0
                # Period label (YYYY-MM)
                month_lbl = rb_date.strftime('%Y-%m')
                # For each ETF in the portfolio
                for tk, w in weights.items():
                    if w <= 0:
                        continue
                    sym = TICKER_MAP.get(tk)
                    if sym is None:
                        continue
                    # ETF return
                    if sym.startswith('_'):
                        # Proxy assets (KR bond / KR cash / USD cash)
                        if sym == '_KRBOND':
                            etf_ret = ((1.035) ** ((p1 - p0).days / 365.25) - 1) * 100
                        elif sym == '_CASH_KR':
                            etf_ret = 0.0
                        elif sym == '_CASH_USD':
                            try:
                                etf_ret = (px['USDKRW=X'].loc[p1] / px['USDKRW=X'].loc[p0] - 1) * 100
                            except Exception:
                                etf_ret = 0.0
                        else:
                            etf_ret = 0.0
                    else:
                        if sym not in px.columns:
                            continue
                        try:
                            etf_ret = (px[sym].loc[p1] / px[sym].loc[p0] - 1) * 100
                        except Exception:
                            etf_ret = 0.0
                    contribution = w * etf_ret  # bps (w% × r%)
                    display_sym = sym if not sym.startswith('_') else tk.upper()
                    rows.append({
                        'Profile':    profile_name,
                        '리밸 일자':   date_str,
                        '월':         month_lbl,
                        '시그널':      signal,
                        'ACWI(%)':    round(acwi_ret, 2),
                        'Ticker':     display_sym,
                        '종목명':      TICKER_NAMES.get(tk, tk),
                        '구분':       TICKER_CATEGORY.get(tk, '기타'),
                        '비중(%)':    round(w, 2),
                        'Total Return(%)': round(etf_ret, 2),
                        '기여(bps)':   round(contribution, 1),
                    })
        return pd.DataFrame(rows)

    full_df = _build_monthly_portfolio_df(end_date_str)
    if full_df is None or full_df.empty:
        st.error('월별 포트폴리오 데이터 부족.')
    else:
        # ─── 🎯 Filter UI ───
        st.markdown('### 🎯 필터')
        fc1, fc2, fc3 = st.columns([1.5, 2, 2])
        with fc1:
            st.markdown('**📁 Profile**')
            profile_filter = st.multiselect(
                ' ', ['적극형', '중립형'], default=['적극형', '중립형'],
                key='mp_profile', label_visibility='collapsed',
            )
        with fc2:
            st.markdown('**🎯 시그널**')
            sc1, sc2, sc3 = st.columns(3)
            with sc1:
                f_bull = st.checkbox('🟢 Bull', value=True, key='mp_bull')
            with sc2:
                f_base = st.checkbox('🟡 Base', value=True, key='mp_base')
            with sc3:
                f_bear = st.checkbox('🔴 Bear', value=True, key='mp_bear')
            signal_filter = [s for s, on in [('Bull', f_bull), ('Base', f_base), ('Bear', f_bear)] if on]
        with fc3:
            st.markdown('**📈 ACWI 방향**')
            ac1, ac2, ac3 = st.columns(3)
            with ac1:
                f_acwi_pos = st.checkbox('🟢 + (상승)', value=True, key='mp_acwi_pos')
            with ac2:
                f_acwi_neg = st.checkbox('🔴 − (하락/0)', value=True, key='mp_acwi_neg')
            with ac3:
                if st.button('🔄 초기화', key='mp_reset', use_container_width=True):
                    for k in ['mp_bull', 'mp_base', 'mp_bear', 'mp_acwi_pos', 'mp_acwi_neg']:
                        st.session_state[k] = True
                    st.session_state['mp_profile'] = ['적극형', '중립형']
                    st.rerun()

        # ─── Apply filters ───
        filtered = full_df.copy()
        if profile_filter:
            filtered = filtered[filtered['Profile'].isin(profile_filter)]
        if signal_filter:
            filtered = filtered[filtered['시그널'].isin(signal_filter)]
        # ACWI filter
        acwi_mask = pd.Series(False, index=filtered.index)
        if f_acwi_pos:
            acwi_mask |= (filtered['ACWI(%)'] > 0)
        if f_acwi_neg:
            acwi_mask |= (filtered['ACWI(%)'] <= 0)
        filtered = filtered[acwi_mask]

        # ─── Summary KPIs ───
        st.markdown('---')
        n_months = filtered.groupby(['Profile', '리밸 일자']).ngroups
        n_etfs = filtered['Ticker'].nunique()
        avg_contrib = filtered['기여(bps)'].mean() if not filtered.empty else 0
        sum_contrib = filtered['기여(bps)'].sum() if not filtered.empty else 0
        k1, k2, k3, k4 = st.columns(4)
        k1.metric('필터된 월 수', f'{n_months}개')
        k2.metric('등장 ETF 수', f'{n_etfs}개')
        k3.metric('평균 기여', f'{avg_contrib:+.1f} bps')
        k4.metric('누적 기여(필터)', f'{sum_contrib:+.1f} bps')

        # ─── 사용자 use-case 인사이트 (Bull 대신 Base 월 등) ───
        st.markdown('---')
        st.markdown('#### 🔍 use-case: "ACWI 상승했지만 Base 시그널 발효" 월 빠른 찾기')
        uc1, uc2, uc3 = st.columns(3)
        # ACWI > 0 + Base 시그널 월
        miss_opp = full_df[(full_df['ACWI(%)'] > 0) & (full_df['시그널'] == 'Base')]
        n_miss = miss_opp.groupby(['Profile', '리밸 일자']).ngroups // max(1, miss_opp['Profile'].nunique())
        uc1.metric('🟡 ACWI↑ + Base (기회손실)', f'{n_miss}개월')
        # ACWI > 0 + Bull 시그널 월
        hit = full_df[(full_df['ACWI(%)'] > 0) & (full_df['시그널'] == 'Bull')]
        n_hit = hit.groupby(['Profile', '리밸 일자']).ngroups // max(1, hit['Profile'].nunique())
        uc2.metric('🟢 ACWI↑ + Bull (적중)', f'{n_hit}개월')
        # ACWI < 0 + Bear 시그널 월
        defense = full_df[(full_df['ACWI(%)'] <= 0) & (full_df['시그널'] == 'Bear')]
        n_defense = defense.groupby(['Profile', '리밸 일자']).ngroups // max(1, defense['Profile'].nunique())
        uc3.metric('🛡️ ACWI↓ + Bear (방어 적중)', f'{n_defense}개월')

        # ─── 📈 캡처 산점도 (월별 포트 수익률 vs ACWI 수익률) ───
        st.markdown('---')
        st.markdown('### 📈 상승을 얼마나 따라갔는가 (Capture Ratio Scatter)')
        st.caption(
            '**대각선은 100% 캡처선** (포트 수익률 = ACWI 수익률).  '
            '대각선 **위**는 ACWI보다 우위, **아래**는 열위.  '
            '필터링된 월만 표시 — Profile별 산점도로 시그널 효과 즉시 가시화.'
        )

        if not filtered.empty:
            # 월별 포트 수익률 계산 (Profile × 리밸 일자별 weighted sum)
            # 포트 월간 return(%) = Σ(weight × Total Return / 100) = Σ(기여) / 100
            scatter_df = filtered.groupby(['Profile', '리밸 일자', '시그널', 'ACWI(%)']).agg(
                port_return=('기여(bps)', lambda s: s.sum() / 100),
                n_etfs=('Ticker', 'count'),
            ).reset_index()
            scatter_df = scatter_df.rename(columns={'port_return': '포트 월간(%)',
                                                      'ACWI(%)': 'ACWI 월간(%)'})

            # Profile 선택 (적극/중립 한 번에 또는 분리)
            sc_view = st.radio(
                '뷰 모드',
                ['📊 통합 (적극+중립)', '🔀 Profile 분리'],
                horizontal=True, key='mp_scatter_view',
            )

            SIG_COLORS = {'Bull': '#C48D43', 'Base': '#5B7BA7', 'Bear': '#DC2626'}

            def _build_scatter(df, title, key_suffix):
                if df.empty:
                    return None
                fig = go.Figure()
                # 대각선 (y = x) — ACWI 범위 기반
                x_min = min(df['ACWI 월간(%)'].min(), df['포트 월간(%)'].min(), 0) - 1
                x_max = max(df['ACWI 월간(%)'].max(), df['포트 월간(%)'].max(), 0) + 1
                fig.add_trace(go.Scatter(
                    x=[x_min, x_max], y=[x_min, x_max],
                    mode='lines', name='100% 캡처 (= ACWI)',
                    line=dict(color='#9E9E9E', width=1.5, dash='dash'),
                    hoverinfo='skip',
                ))
                # Signal별 점
                for sig in ['Bull', 'Base', 'Bear']:
                    sub = df[df['시그널'] == sig]
                    if sub.empty:
                        continue
                    fig.add_trace(go.Scatter(
                        x=sub['ACWI 월간(%)'], y=sub['포트 월간(%)'],
                        mode='markers',
                        name=f'{sig} (n={len(sub)})',
                        marker=dict(color=SIG_COLORS[sig], size=12, opacity=0.85,
                                    line=dict(color='white', width=1.2)),
                        text=[f"{p} · {d}" for p, d in zip(sub['Profile'], sub['리밸 일자'])],
                        hovertemplate=('%{text}<br>' + sig +
                                       '<br>ACWI: %{x:+.2f}%<br>포트: %{y:+.2f}%' +
                                       '<extra></extra>'),
                    ))
                fig.add_hline(y=0, line_dash='dot', line_color='#CCCCCC', line_width=1)
                fig.add_vline(x=0, line_dash='dot', line_color='#CCCCCC', line_width=1)
                # 100% 캡처 annotation (우측 상단)
                fig.add_annotation(
                    x=x_max * 0.95, y=x_max * 0.95, text='100% 캡처 (= ACWI)',
                    showarrow=False, font=dict(size=11, color='#666666'),
                    xanchor='right', yanchor='bottom',
                )
                fig.update_layout(
                    height=520, hovermode='closest',
                    title=dict(text=title, font=dict(size=14)) if title else None,
                    xaxis=dict(title='MSCI ACWI 월간 수익률 (%)',
                               zeroline=False, gridcolor='#EEEEEE'),
                    yaxis=dict(title='포트 월간 수익률 (%)',
                               zeroline=False, gridcolor='#EEEEEE',
                               scaleanchor='x', scaleratio=1),
                    legend=dict(orientation='h', yanchor='bottom', y=-0.18,
                                xanchor='left', x=0, font=dict(size=10)),
                    margin=dict(t=40, b=70, l=60, r=30),
                    plot_bgcolor='white',
                )
                return fig

            # ─── ACWI 상승 월만 필터 ───
            scatter_up = scatter_df[scatter_df['ACWI 월간(%)'] > 0]
            scatter_dn = scatter_df[scatter_df['ACWI 월간(%)'] <= 0]

            st.markdown(f'#### 🟢 ACWI 상승 월 ({len(scatter_up)}개월)')
            if scatter_up.empty:
                st.info('ACWI 상승 월 데이터 없음')
            elif sc_view.startswith('📊'):
                fig = _build_scatter(scatter_up, None, 'up_all')
                if fig is not None:
                    st.plotly_chart(fig, use_container_width=True, key='mp_scatter_up_all')
            else:
                sc1, sc2 = st.columns(2)
                with sc1:
                    fig_a = _build_scatter(scatter_up[scatter_up['Profile'] == '적극형'],
                                             '적극형', 'up_agg')
                    if fig_a is not None:
                        st.plotly_chart(fig_a, use_container_width=True, key='mp_scatter_up_agg')
                    else:
                        st.info('적극형 ACWI 상승 데이터 없음')
                with sc2:
                    fig_n = _build_scatter(scatter_up[scatter_up['Profile'] == '중립형'],
                                             '중립형', 'up_neu')
                    if fig_n is not None:
                        st.plotly_chart(fig_n, use_container_width=True, key='mp_scatter_up_neu')
                    else:
                        st.info('중립형 ACWI 상승 데이터 없음')

            # ─── ACWI 하락 월 시각화 ───
            st.markdown('---')
            st.markdown('### 🛡️ 하락을 얼마나 방어했는가 (Downside Capture Scatter)')
            st.caption(
                '**대각선은 100% 캡처선** (포트 수익률 = ACWI 수익률).  '
                '하락 구간에서는 **대각선 위(덜 떨어진)** 가 방어 우위.  '
                'Bull 시그널이 하락 월에 발효되면 risk-on 노출로 더 떨어질 수 있고, '
                'Bear 시그널 발효 시는 방어자산 비중으로 덜 떨어지는 효과 기대.'
            )

            st.markdown(f'#### 🔴 ACWI 하락 월 ({len(scatter_dn)}개월)')
            if scatter_dn.empty:
                st.info('ACWI 하락 월 데이터 없음')
            elif sc_view.startswith('📊'):
                fig = _build_scatter(scatter_dn, None, 'dn_all')
                if fig is not None:
                    st.plotly_chart(fig, use_container_width=True, key='mp_scatter_dn_all')
            else:
                sc1, sc2 = st.columns(2)
                with sc1:
                    fig_a = _build_scatter(scatter_dn[scatter_dn['Profile'] == '적극형'],
                                             '적극형', 'dn_agg')
                    if fig_a is not None:
                        st.plotly_chart(fig_a, use_container_width=True, key='mp_scatter_dn_agg')
                    else:
                        st.info('적극형 ACWI 하락 데이터 없음')
                with sc2:
                    fig_n = _build_scatter(scatter_dn[scatter_dn['Profile'] == '중립형'],
                                             '중립형', 'dn_neu')
                    if fig_n is not None:
                        st.plotly_chart(fig_n, use_container_width=True, key='mp_scatter_dn_neu')
                    else:
                        st.info('중립형 ACWI 하락 데이터 없음')

            # 캡처 비율 KPI 요약
            with st.expander('📊 캡처 비율 요약 (Profile × 시그널)'):
                cap_rows = []
                for prof in scatter_df['Profile'].unique():
                    for sig in ['Bull', 'Base', 'Bear']:
                        sub = scatter_df[(scatter_df['Profile'] == prof) &
                                          (scatter_df['시그널'] == sig)]
                        n = len(sub)
                        if n == 0:
                            continue
                        # 상승 시 (ACWI > 0) 평균 캡처 비율
                        up = sub[sub['ACWI 월간(%)'] > 0]
                        if len(up) > 0:
                            cap_up = (up['포트 월간(%)'].sum() / up['ACWI 월간(%)'].sum() * 100
                                       if up['ACWI 월간(%)'].sum() != 0 else 0)
                        else:
                            cap_up = float('nan')
                        # 하락 시 (ACWI ≤ 0) 평균 캡처 비율
                        dn = sub[sub['ACWI 월간(%)'] <= 0]
                        if len(dn) > 0 and dn['ACWI 월간(%)'].sum() != 0:
                            cap_dn = (dn['포트 월간(%)'].sum() / dn['ACWI 월간(%)'].sum() * 100)
                        else:
                            cap_dn = float('nan')
                        cap_rows.append({
                            'Profile': prof, '시그널': sig, '이벤트(개월)': n,
                            'ACWI↑ 캡처(%)': round(cap_up, 1) if up.shape[0] > 0 else None,
                            'ACWI↓ 캡처(%)': round(cap_dn, 1) if dn.shape[0] > 0 else None,
                            '포트 평균(%)': round(sub['포트 월간(%)'].mean(), 2),
                            'ACWI 평균(%)': round(sub['ACWI 월간(%)'].mean(), 2),
                        })
                if cap_rows:
                    cap_df = pd.DataFrame(cap_rows)
                    cap_styled = cap_df.style.background_gradient(
                        subset=['ACWI↑ 캡처(%)'], cmap='RdYlGn', vmin=0, vmax=120,
                    ).background_gradient(
                        subset=['ACWI↓ 캡처(%)'], cmap='RdYlGn_r', vmin=0, vmax=120,
                    ).format({
                        'ACWI↑ 캡처(%)':  '{:.1f}',
                        'ACWI↓ 캡처(%)':  '{:.1f}',
                        '포트 평균(%)':    '{:+.2f}',
                        'ACWI 평균(%)':   '{:+.2f}',
                    }, na_rep='—')
                    st.dataframe(cap_styled, use_container_width=True, hide_index=True)
                    st.caption(
                        ':grey[**캡처 비율(%) 정의:** 해당 시그널의 (포트 월별 수익률 합) / (ACWI 월별 수익률 합) × 100.  '
                        '**ACWI↑ 캡처:** 100% 초과 = 상승 시 ACWI 보다 더 따라감 (긍정).  '
                        '**ACWI↓ 캡처:** 100% 미만 = 하락 시 ACWI 보다 덜 떨어짐 (방어, 긍정) — 컬러 reversed.]'
                    )

        # ─── Display table ───
        st.markdown('---')
        st.markdown(f'### 📋 필터링된 월별 종목 성과 ({len(filtered)}행)')
        if filtered.empty:
            st.info('필터 조건에 맞는 데이터 없음 — 필터 옵션을 조정해주세요.')
        else:
            # Sort by 월 descending, then ETF
            filtered_sorted = filtered.sort_values(['Profile', '리밸 일자', 'Ticker'],
                                                     ascending=[True, False, True]).reset_index(drop=True)
            # Style with gradient
            display_cols = ['Profile', '리밸 일자', '시그널', 'ACWI(%)',
                             'Ticker', '종목명', '구분', '비중(%)', 'Total Return(%)', '기여(bps)']
            styled = filtered_sorted[display_cols].style.background_gradient(
                subset=['ACWI(%)', 'Total Return(%)'], cmap='RdYlGn', vmin=-15, vmax=15,
            ).background_gradient(
                subset=['기여(bps)'], cmap='RdYlGn', vmin=-200, vmax=200,
            ).background_gradient(
                subset=['비중(%)'], cmap='Blues', vmin=0, vmax=30,
            ).format({
                'ACWI(%)':           '{:+.2f}',
                '비중(%)':            '{:.2f}',
                'Total Return(%)':   '{:+.2f}',
                '기여(bps)':          '{:+.1f}',
            })
            st.dataframe(styled, use_container_width=True, hide_index=True, height=620)

            # 월별 요약 expander
            with st.expander('📊 월별 portfolio summary (Profile × 월 단위 합계)'):
                summary = filtered.groupby(['Profile', '리밸 일자', '시그널', 'ACWI(%)']).agg(
                    ETF수=('Ticker', 'count'),
                    총비중=('비중(%)', 'sum'),
                    평균_TR=('Total Return(%)', 'mean'),
                    누적기여=('기여(bps)', 'sum'),
                ).reset_index().sort_values(['Profile', '리밸 일자'], ascending=[True, False])
                summary_styled = summary.style.background_gradient(
                    subset=['ACWI(%)'], cmap='RdYlGn', vmin=-15, vmax=15,
                ).background_gradient(
                    subset=['평균_TR'], cmap='RdYlGn', vmin=-10, vmax=10,
                ).background_gradient(
                    subset=['누적기여'], cmap='RdYlGn', vmin=-500, vmax=500,
                ).format({
                    'ACWI(%)': '{:+.2f}',
                    '총비중':   '{:.1f}',
                    '평균_TR': '{:+.2f}',
                    '누적기여': '{:+.1f}',
                })
                st.dataframe(summary_styled, use_container_width=True, hide_index=True, height=420)

            # CSV download
            csv_data = filtered_sorted[display_cols].to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                '📥 필터링된 결과 CSV',
                data=csv_data,
                file_name=f'aimvp_monthly_portfolio_filter_{report_date.strftime("%Y%m%d")}.csv',
                mime='text/csv',
            )

        # ─── 데이터 출처 caption ───
        st.caption(
            ':grey[**📊 데이터 출처:** Yahoo Finance — auto-adjusted (Total Return).  '
            '**ACWI(%):** 해당 리밸 시점 → 다음 리밸(또는 조회일) 보유 기간 MSCI ACWI total return.  '
            '**Total Return(%):** 동일 보유 기간 ETF total return.  '
            '**기여(bps):** 비중 × Total Return → 포트폴리오 레벨 기여 (월별 독립 산출).  '
            '**KR 채권·현금:** proxy 상수 수익률 적용 (yfinance 데이터 부재).  '
            '**Profile:** 적극형(89.55% 주식) / 중립형(59.70% 주식).]'
        )

# ============================================================
# Footer
# ============================================================
st.markdown('---')
foot1, foot2 = st.columns(2)
with foot1:
    st.caption(
        f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M KST")}  |  '
        f'AIMVP Phase 1 Dashboard MVP  |  Data: Yahoo Finance'
    )
with foot2:
    st.caption(f'**다음 발행:** {(report_date + timedelta(days=7)).strftime("%Y-%m-%d (Fri)")}')
